from __future__ import annotations

import argparse
import copy
import json
import re
import shutil
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "新闻格式.xlsx"
GROUP_ORDER = {"巴西": 0, "印度": 1, "泰国": 2, "中国": 3, "其他国家": 4}
IMPACT_PREFIXES = ("偏多糖价：", "偏空糖价：", "利多：", "利空：", "中性：", "影响有限：")
PLACEHOLDERS = (
    "暂无新闻",
    "暂无最新数据",
    "暂无最新对比数据",
    "暂无可比数据",
    "暂无最新",
    "暂未更新",
    "数据尚未公布",
)
INDIA_MAIN_CANE_REGIONS = (
    "北方邦", "Uttar Pradesh", "UP",
    "马哈拉施特拉邦", "Maharashtra",
    "卡纳塔克邦", "Karnataka",
    "泰米尔纳德邦", "Tamil Nadu",
    "古吉拉特邦", "Gujarat",
    "比哈尔邦", "Bihar",
    "旁遮普邦", "Punjab",
    "哈里亚纳邦", "Haryana",
    "北阿坎德邦", "Uttarakhand",
)
INDIA_WEATHER_TERMS = (
    "降雨", "雨", "季风", "天气", "气象", "预警", "强降雨", "暴雨", "干旱", "洪涝", "积水",
    "rain", "rainfall", "monsoon", "weather", "alert", "warning", "heavy rain", "flood", "drought",
)
INDIA_RAIN_BENEFIT_TERMS = (
    "适量降雨", "降雨增加", "降雨增多", "未来降雨", "强降雨", "暴雨预报", "墒情改善",
    "季风活跃", "季风增强", "widespread rainfall", "heavy rainfall", "rainfall forecast",
    "monsoon revival", "active monsoon",
)
INDIA_DROUGHT_TERMS = (
    "干旱", "降雨不足", "季风偏弱", "降雨减少", "雨量不足", "deficient rainfall",
    "weak monsoon", "dry spell", "rainfall deficit",
)
INDIA_DAMAGE_TERMS = (
    "已造成", "洪涝", "农田被淹", "甘蔗倒伏", "道路中断", "作物受损", "预计减产",
    "受灾", "损失", "flood damage", "crop damage", "waterlogging", "lodging", "road disruption",
)
INDIA_HARVEST_TERMS = ("收割", "压榨", "运输", "入榨", "开榨", "砍蔗", "harvest", "crushing", "transport")
THAI_MAIN_CANE_PROVINCES = (
    "乌隆他尼", "Udon Thani",
    "孔敬", "Khon Kaen",
    "呵叻", "那空叻差是玛", "Nakhon Ratchasima",
    "猜也蓬", "Chaiyaphum",
    "加拉信", "Kalasin",
    "黎府", "Loei",
    "那空沙旺", "Nakhon Sawan",
    "甘烹碧", "Kamphaeng Phet",
    "素可泰", "Sukhothai",
    "彭世洛", "Phitsanulok",
    "北碧", "Kanchanaburi",
    "华富里", "Lopburi",
    "素攀武里", "Suphanburi",
    "猜纳", "Chai Nat",
    "沙缴", "Sa Kaeo",
    "春武里", "Chonburi",
)
THAI_WEATHER_TERMS = (
    "天气", "气象", "降雨", "雨", "雷阵雨", "干旱", "洪涝", "积水",
    "rain", "rainfall", "thunderstorm", "flood", "drought",
)
THAI_RAIN_INCREASE_TERMS = (
    "降雨增加", "降雨将增加", "雨量增加", "降雨增多", "降雨明显增多", "降雨改善",
    "墒情", "有利于改善", "强降雨", "强到很强降雨", "暴雨预警", "降雨范围", "降雨强度",
)
THAI_LOW_COVERAGE_TERMS = ("20%", "约20%", "较分散", "分散", "覆盖率较低", "覆盖率低", "局地", "少量")
THAI_DAMAGE_TERMS = (
    "已造成", "造成严重洪涝", "严重洪涝", "甘蔗倒伏", "农田被淹",
    "作物受损", "预计减产", "受灾", "损失", "根系受损",
)
THAI_DROUGHT_TERMS = ("降雨减少", "降雨不足", "持续干旱", "干旱", "偏干")
THAI_HARVEST_TERMS = ("收割", "压榨", "运输", "入榨", "开榨", "收榨")
THAI_WEATHER_EVENT_TERMS = (
    "降雨", "雨量", "雷阵雨", "干旱", "洪涝", "积水",
    "rain", "rainfall", "thunderstorm", "flood", "drought",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write audited Sugar News Excel report.")
    parser.add_argument(
        "--date",
        dest="target_date",
        help="News target date in YYYY-MM-DD. Defaults to current local date minus one day.",
    )
    return parser.parse_args()


def target_date_from_args(value: str | None) -> date:
    if value:
        return datetime.strptime(value, "%Y-%m-%d").date()
    return date.today() - timedelta(days=1)


def load_verified_news(target: date) -> dict:
    path = (
        ROOT
        / "data"
        / "verified_news"
        / f"{target:%Y}"
        / f"{target:%m}"
        / f"sugar_news_{target:%Y-%m-%d}.json"
    )
    if not path.exists():
        raise FileNotFoundError(f"Missing verified news file: {path}")
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if data.get("target_date") != f"{target:%Y-%m-%d}":
        raise ValueError(f"Verified news target_date mismatch in {path}")
    return data


def _contains_any(text: str, terms: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(term.lower() in lowered for term in terms)


def validate_india_weather_impact(item: dict, idx: int) -> None:
    if item.get("country_group") != "印度":
        return

    fact_text = " ".join(str(item.get(field, "")) for field in ("title", "news"))
    text = f"{fact_text} {item.get('impact', '')}"
    if not _contains_any(text, INDIA_WEATHER_TERMS):
        return

    in_main_area = _contains_any(fact_text, INDIA_MAIN_CANE_REGIONS)
    if not in_main_area:
        if not item["impact"].startswith("影响有限："):
            raise ValueError(f"India weather item {idx + 1} is outside main cane regions and should be impact-limited")
        return

    if _contains_any(fact_text, INDIA_HARVEST_TERMS):
        if not item["impact"].startswith("偏多糖价："):
            raise ValueError(f"India weather item {idx + 1} indicates harvest/crushing disruption and should be bullish")
        return

    if _contains_any(fact_text, INDIA_DAMAGE_TERMS):
        if not item["impact"].startswith("偏多糖价："):
            raise ValueError(f"India weather item {idx + 1} indicates confirmed damage and should be bullish")
        return

    if _contains_any(fact_text, INDIA_DROUGHT_TERMS):
        if not item["impact"].startswith("偏多糖价："):
            raise ValueError(f"India weather item {idx + 1} indicates drought/rain shortage and should be bullish")
        return

    if _contains_any(fact_text, INDIA_RAIN_BENEFIT_TERMS):
        if not item["impact"].startswith("偏空糖价："):
            raise ValueError(f"India weather item {idx + 1} indicates growing-season rainfall support and should be bearish")


def validate_thai_weather_impact(item: dict, idx: int) -> None:
    if item.get("country_group") != "泰国":
        return

    fact_text = " ".join(str(item.get(field, "")) for field in ("title", "news"))
    text = f"{fact_text} {item.get('impact', '')}"
    if not _contains_any(text, THAI_WEATHER_TERMS):
        return
    if not _contains_any(fact_text, THAI_WEATHER_EVENT_TERMS):
        return

    in_main_area = (
        _contains_any(fact_text, THAI_MAIN_CANE_PROVINCES)
        or "东北部" in fact_text
        or "中部" in fact_text
    )
    if not in_main_area:
        if not item["impact"].startswith("影响有限："):
            raise ValueError(f"Thai weather item {idx + 1} is outside main cane areas and should be impact-limited")
        return

    is_bearish = item["impact"].startswith(("偏空糖价：", "利空："))
    if _contains_any(fact_text, THAI_HARVEST_TERMS):
        if not item["impact"].startswith("偏多糖价："):
            raise ValueError(f"Thai weather item {idx + 1} indicates harvest disruption and should be bullish")
        return

    if _contains_any(fact_text, THAI_DAMAGE_TERMS):
        return

    if _contains_any(fact_text, THAI_DROUGHT_TERMS):
        if not item["impact"].startswith("偏多糖价："):
            raise ValueError(f"Thai weather item {idx + 1} indicates drought/rain shortage and should be bullish")
        return

    if _contains_any(fact_text, THAI_RAIN_INCREASE_TERMS) or _contains_any(fact_text, ("雷阵雨", "阵雨", "大雨", "强降雨")):
        if not is_bearish:
            raise ValueError(f"Thai weather item {idx + 1} indicates growing-season rainfall improvement and should be bearish")


def normalized_items(data: dict) -> list[dict]:
    items = data.get("items") or []
    seen = set()
    output = []
    for idx, item in enumerate(items):
        for field in ("country_group", "country", "news", "impact", "source_url", "published_date_local"):
            if not item.get(field):
                raise ValueError(f"Item {idx + 1} missing required field: {field}")
        if item["published_date_local"] != data["target_date"] and item.get("date_status") != "continuing_impact":
            raise ValueError(f"Item {idx + 1} has unaccepted date: {item['published_date_local']}")
        if not any(item["impact"].startswith(prefix) for prefix in IMPACT_PREFIXES):
            raise ValueError(f"Item {idx + 1} impact must start with one of {IMPACT_PREFIXES}")
        if any(text in item["news"] or text in item["impact"] for text in PLACEHOLDERS):
            raise ValueError(f"Item {idx + 1} contains placeholder/missing-data wording")
        if "LMT" in item["news"].upper() or " LMT" in item["impact"].upper() or "lmt" in item["news"]:
            raise ValueError(f"Item {idx + 1} contains raw LMT unit")
        if "来源：" not in item["news"] or item["source_url"] not in item["news"]:
            raise ValueError(f"Item {idx + 1} source must be included at the end of B-column news")
        if item["country_group"] == "其他国家" and item["country"] == "其他":
            raise ValueError("Other-country rows must use concrete country/region names, not 其他")
        if item["country"] == "中国" and item["country_group"] != "中国":
            raise ValueError("China news must use country_group=中国 and must not be stored as other-country news")
        if item["country_group"] == "中国" and item["country"] != "中国":
            raise ValueError("country_group=中国 rows must use country=中国")
        validate_india_weather_impact(item, idx)
        validate_thai_weather_impact(item, idx)
        key = item.get("dedupe_key") or re.sub(r"\s+", "", item["news"][:80])
        if key in seen:
            raise ValueError(f"Duplicate dedupe_key: {key}")
        seen.add(key)
        item["_input_order"] = idx
        output.append(item)

    return sorted(
        output,
        key=lambda x: (
            GROUP_ORDER.get(x["country_group"], 3),
            -int(x.get("importance", 0)),
            x["_input_order"],
        ),
    )


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, 4):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.protection:
            target.protection = copy.copy(source.protection)
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)
        if source.font:
            target.font = copy.copy(source.font)


def write_excel(target: date, items: list[dict]) -> Path:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Missing template: {TEMPLATE}")

    report_dir = ROOT / "reports" / f"{target:%Y}" / f"{target:%m}"
    report_dir.mkdir(parents=True, exist_ok=True)
    output = report_dir / f"Sugar News {target:%Y-%m-%d}.xlsx"
    shutil.copy2(TEMPLATE, output)

    wb = load_workbook(output)
    if len(wb.sheetnames) != 1:
        raise ValueError("Template must contain exactly one worksheet")
    ws = wb.active
    if [ws.cell(1, col).value for col in range(1, 4)] != ["国家", "新闻", "影响"]:
        raise ValueError("Template headers must be 国家/新闻/影响")

    max_row = max(ws.max_row, 2)
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)

    template_wb = load_workbook(TEMPLATE)
    template_ws = template_wb.active
    source_row = 2 if template_ws.max_row >= 2 else 1

    for row_offset, item in enumerate(items, start=2):
        copy_row_style(template_ws, source_row, row_offset)
        ws.cell(row_offset, 1).value = item["country"]
        ws.cell(row_offset, 2).value = item["news"]
        ws.cell(row_offset, 3).value = item["impact"]
        ws.cell(row_offset, 1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.cell(row_offset, 2).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.cell(row_offset, 3).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.row_dimensions[row_offset].height = max(
            72,
            min(180, 24 + 0.55 * max(len(item["news"]), len(item["impact"]))),
        )

    for col in ("B", "C"):
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, 55)

    wb.save(output)
    return output


def validate_saved_excel(output: Path, items: list[dict]) -> dict:
    wb = load_workbook(output)
    ws = wb.active
    rows = []
    for row in range(2, ws.max_row + 1):
        country = ws.cell(row, 1).value
        news = ws.cell(row, 2).value
        impact = ws.cell(row, 3).value
        if country or news or impact:
            rows.append({"row": row, "country": country, "news": news, "impact": impact})

    expected_pairs = {(item["country"], item["news"]) for item in items}
    actual_pairs = {(row["country"], row["news"]) for row in rows}
    missing_pairs = expected_pairs - actual_pairs

    group_positions = []
    for row in rows:
        group = row["country"] if row["country"] in GROUP_ORDER and row["country"] != "其他国家" else None
        if row["country"] == "巴西":
            group_positions.append(0)
        elif row["country"] == "印度":
            group_positions.append(1)
        elif row["country"] == "泰国":
            group_positions.append(2)
        elif row["country"] == "中国":
            group_positions.append(3)
        else:
            group_positions.append(4)

    checks = {
        "headers_ok": [ws.cell(1, col).value for col in range(1, 4)] == ["国家", "新闻", "影响"],
        "row_count_matches_verified": len(rows) == len(items),
        "all_verified_news_written": not missing_pairs,
        "country_order_ok": group_positions == sorted(group_positions),
        "no_extra_columns": ws.max_column == 3,
        "no_merged_cells": len(ws.merged_cells.ranges) == 0,
        "impact_prefixes_ok": all(
            isinstance(row["impact"], str) and row["impact"].startswith(IMPACT_PREFIXES)
            for row in rows
        ),
        "sources_ok": all(isinstance(row["news"], str) and "来源：" in row["news"] for row in rows),
        "no_placeholders": all(
            not any(text in (row["news"] or "") or text in (row["impact"] or "") for text in PLACEHOLDERS)
            for row in rows
        ),
        "wrap_ok": all(
            ws.cell(row["row"], col).alignment.wrap_text for row in rows for col in (2, 3)
        ),
        "other_countries_not_collapsed": all(row["country"] != "其他" for row in rows),
        "china_not_in_other_group": True,
    }
    if any(not value for value in checks.values()):
        raise ValueError(f"Excel validation failed: {checks}; missing={list(missing_pairs)}")

    return {
        "output": str(output),
        "excel_rows": len(rows),
        "countries_written": dict(Counter(row["country"] for row in rows)),
        "checks": checks,
        "saved_rows": rows,
    }


def write_log(target: date, data: dict, items: list[dict], validation: dict) -> Path:
    log_dir = ROOT / "logs" / f"{target:%Y}" / f"{target:%m}"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"write_log_{target:%Y-%m-%d}.json"
    payload = {
        "target_date": f"{target:%Y-%m-%d}",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "verified_news_file_count": len(data.get("items") or []),
        "passed_to_excel_count": len(items),
        "excel_saved_count": validation["excel_rows"],
        "verified_count_by_country": dict(Counter(item["country"] for item in items)),
        "excel_count_by_country": validation["countries_written"],
        "checks": validation["checks"],
    }
    with log_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return log_path


def main() -> None:
    target = target_date_from_args(parse_args().target_date)
    data = load_verified_news(target)
    items = normalized_items(data)
    output = write_excel(target, items)
    validation = validate_saved_excel(output, items)
    log_path = write_log(target, data, items, validation)
    print(json.dumps({
        "target_date": f"{target:%Y-%m-%d}",
        "output": str(output),
        "write_log": str(log_path),
        "excel_saved_count": validation["excel_rows"],
        "countries_written": validation["countries_written"],
        "checks": validation["checks"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
