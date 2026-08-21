from __future__ import annotations

import argparse
import codecs
import hashlib
from html import unescape
import json
import os
import re
import shutil
import subprocess
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from urllib.parse import quote_plus, urlencode
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TASK_ROOT = PROJECT_ROOT
PUBLIC_ROOT = PROJECT_ROOT / "public" / "sugar-news"
PUBLIC_DATA_ROOT = PUBLIC_ROOT / "data"
EDITORIAL_SKILL_PATH = PROJECT_ROOT / ".codex" / "skills" / "sugar-news-editorial-rules" / "SKILL.md"
RSS_AUTOGEN_TIMEOUT_SECONDS = int(os.getenv("SUGAR_NEWS_RSS_TIMEOUT", "4"))
RSS_AUTOGEN_MAX_QUERIES_PER_COUNTRY = int(os.getenv("SUGAR_NEWS_RSS_MAX_QUERIES_PER_COUNTRY", "32"))
RSS_AUTOGEN_MAX_TOTAL_QUERIES = int(os.getenv("SUGAR_NEWS_RSS_MAX_TOTAL_QUERIES", "180"))
RSS_AUTOGEN_MAX_ITEMS_PER_QUERY = int(os.getenv("SUGAR_NEWS_RSS_MAX_ITEMS_PER_QUERY", "20"))
RSS_AUTOGEN_PUBLICATION_WINDOW_HOURS = int(os.getenv("SUGAR_NEWS_RSS_PUBLICATION_WINDOW_HOURS", "36"))
TMD_DAILY_FORECAST_URL = "https://tmd.go.th/en/forecast/daily"
OPEN_METEO_FORECAST_URL = "https://api.open-meteo.com/v1/forecast"
METRIC_REFRESH_TIMEOUT_SECONDS = int(os.getenv("SUGAR_NEWS_METRIC_REFRESH_TIMEOUT", "240"))
try:
    SHANGHAI = ZoneInfo("Asia/Shanghai")
except Exception:
    SHANGHAI = timezone(timedelta(hours=8), name="Asia/Shanghai")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
GROUP_ORDER = {"巴西": 0, "印度": 1, "泰国": 2, "其他国家": 3, "中国": 4}
COUNTRY_ALIASES = {
    "中国": ("china", "中国", "广西", "云南", "郑糖"),
    "巴西": ("brazil", "brasil", "brazilian", "巴西", "sao paulo", "centro-sul", "caarapó", "caarapo", "raízen", "raizen", "adecoagro"),
    "印度": ("india", "indian", "uttar pradesh", "maharashtra", "karnataka", "bihar", "shamli", "belagavi", "amaravathi", "印度", "北方邦", "卡纳塔克", "比哈尔"),
    "泰国": ("thailand", "thai", "ประเทศไทย", "泰国"),
    "印度尼西亚": ("indonesia", "indonesian", "印尼", "印度尼西亚"),
    "巴基斯坦": ("pakistan", "pakistani", "巴基斯坦"),
    "菲律宾": ("philippines", "philippine", "菲律宾"),
    "孟加拉国": ("bangladesh", "bangladeshi", "tangail", "孟加拉", "唐盖尔"),
    "肯尼亚": ("kenya", "kenyan", "naivas", "cleanshelf", "quickmart", "carrefour", "soko directory", "肯尼亚"),
    "斐济": ("fiji", "fijian", "fbc news", "斐济"),
    "南非": ("south africa", "south african", "kwazulu", "african farming", "南非"),
    "越南": ("vietnam", "vietnamese", "越南"),
    "俄罗斯": ("russia", "russian", "俄罗斯"),
    "英国": ("british sugar", "united kingdom", "uk sugar", "cantley", "英国"),
    "斐济": ("fiji", "fijian", "fsc", "斐济"),
    "喀麦隆": ("cameroon", "cameroon's", "cameroonian", "喀麦隆"),
    "欧盟": ("european union", " eu ", "欧盟"),
    "美国": ("united states", " u.s.", " us ", "美国"),
    "墨西哥": ("mexico", "mexican", "墨西哥"),
}
MEDICAL_SUGAR_TERMS = (
    "blood sugar", "glucose", "diabetes", "diabetic", "insulin", "glycemic",
    "hyperglycemia", "hypoglycemia", "glucose monitoring", "diabetes treatment",
    "blood glucose", "low blood sugar", "high blood sugar", "血糖", "糖尿病",
    "胰岛素", "降糖", "低血糖", "高血糖", "血糖监测",
    "dementia", "alzheimer", "health risk", "health benefits", "nutrition",
    "dietary", "calorie", "obesity", "sweetener health", "early life",
)
NON_INDUSTRY_SUGAR_TERMS = (
    "video game", "game launch", "launches on windows", "windows pc",
    "steam", "nintendo", "playstation", "xbox", "novel", "book launch",
    "author", "debut novel", "fiction", "film", "album", "song",
    "restaurant", "dessert recipe", "cake recipe", "horoscope", "zodiac",
    "love forecast", "weekly love", "weekly health", "sugar bytes", "synth",
    "plugin", "plugins", "sound pack", "mega bundle", "monsoon session",
    "assembly session", "parliament session", "legislative session",
)
IMPACT_PREFIXES = ("偏多糖价：", "偏空糖价：", "利多：", "利空：", "中性：", "影响有限：")
PLACEHOLDERS = (
    "暂无新闻",
    "暂无最新数据",
    "暂无最新对比数据",
    "暂无可比数据",
    "暂无最新",
    "暂未更新",
    "数据尚未公布",
)
VAGUE_SUMMARY_PHRASES = (
    "关键数据包括",
    "指标包括",
    "相关数值为",
    "糖价相关数值为",
    "数据为",
    "已披露具体事件方向但标题缺少数值",
    "具体幅度未披露",
    "涉及食糖价格或市场流通变化",
    "涉及巴西食糖价格或市场流通变化",
    "对市场具有参考意义",
    "对糖价具有参考意义",
    "该消息可能影响市场情绪",
    "可能影响市场情绪",
    "将影响贸易商采购和终端补库",
    "价格变化会影响贸易商采购和终端补库",
    "市场关注相关变化",
    "对糖价走势产生一定影响",
    "相关消息值得关注",
    "行业发展值得持续关注",
    "行业发展迎来新变化",
    "供需格局可能发生变化",
    "相关政策可能对市场产生影响",
    "后续影响仍需观察",
    "市场仍需关注后续变化",
    "该事项对食糖供应、需求或价格的影响仍需结合后续政策、产量和贸易数据继续跟踪",
    "该信息需要继续跟踪，短期对当期糖产量和出口量的直接影响有限",
    "相关变化可能影响糖料供应、压榨节奏或加工能力",
    "后续需跟踪对食糖产量和现货供应的实际影响",
    "该变化会改变甘蔗、糖蜜或糖浆在制糖和制醇之间的分配",
    "进而影响食糖供应",
    "事件归属为",
    "事件归属国家",
    "公开标题显示",
    "标题显示",
    "原文标题未披露可量化幅度",
    "该事件会影响糖业供应、需求、库存或产业运行预期",
    "该事件属于糖业产业链信息",
    "标题未给出足以判断单边方向",
    "该供需数据需要结合产量、库存和贸易流向判断",
    "产量、库存、销量或消费变化会直接改变食糖供需平衡",
    "糖厂运行变化会直接影响甘蔗入榨、压榨节奏和阶段性食糖产量",
    "报价变化会反映现货供需松紧和贸易商补库意愿",
    "进口、出口、关税或配额变化会改变国内外可用糖源和贸易流向",
    "印度价格上行通常会抬高补库成本，价格下行则说明供应压力或需求走弱正在传导到现货端",
    "产区天气或病虫害变化会影响甘蔗生长、收割和糖料供应稳定性",
    "主产区农业或气象机构预警甘蔗产区天气、干旱或病虫害",
)
NEWS_IMPACT_MARKER_RE = re.compile(r"\s*。?影响：(?:利多糖价|利空糖价|中性)\s*$")
VAGUE_SUMMARY_PATTERNS = (
    re.compile(r"[^。！？]{0,50}消息涉及[^。！？]{0,80}(?:变化|安排|运行|市场|糖业)"),
    re.compile(r"(?:该事项|该信息|相关变化)[^。！？]{0,80}(?:继续跟踪|参考意义|影响有限)"),
    re.compile(r"(?:改变|影响)[^。！？]{0,30}(?:甘蔗|糖蜜|糖浆)[^。！？]{0,40}(?:制糖|制醇)[^。！？]{0,40}(?:分配|食糖供应)"),
    re.compile(r"(?:数据为|具体幅度未披露)[^。！？]{0,40}"),
)
NEWS_ACTION_TERMS = (
    "宣布", "公布", "发布", "批准", "要求", "计划", "拟", "预计", "预报", "预测",
    "提高", "上调", "下调", "降低", "上涨", "下跌", "增加", "减少", "增长", "下降",
    "达到", "转向", "改为", "支持",
    "扩大", "收紧", "放宽", "限制", "禁止", "取消", "暂停", "恢复", "关闭", "启动",
    "开榨", "收榨", "压榨", "生产", "产糖", "进口", "出口", "销售", "采购", "库存",
    "报价", "收购价", "配额", "关税", "补贴", "融资", "收购", "出售", "扩建", "停产",
    "复产", "受损", "预警", "降雨", "暴雨", "干旱", "洪涝", "雷阵雨", "大雨",
    "said", "announced", "reported", "forecast", "estimated", "raised", "cut", "increased",
    "decreased", "fell", "rose", "approved", "restricted", "resumed", "halted",
)
NEWS_DIRECTION_TERMS = (
    "上调", "下调", "提高", "降低", "上涨", "下跌", "增加", "减少", "增长", "下降",
    "同比", "环比", "由", "至", "达到", "为", "超过", "不足", "偏高", "偏低",
    "暂停", "恢复", "禁止", "批准", "限制", "启动", "关闭", "开榨", "收榨",
    "预计", "预报", "预测", "大雨", "暴雨", "干旱", "洪涝", "短缺", "过剩",
    "扩张", "扩大", "扩散", "蔓延", "收缩", "改善", "恶化", "受损", "支撑", "压制", "压低",
    "increase", "decrease", "rise", "fall", "raise", "cut", "lower", "higher", "lower",
)
NEWS_DETAIL_TERMS = (
    "万吨", "吨", "千吨", "公担", "卢比", "雷亚尔", "美元", "美分", "元/吨", "mm",
    "榨季", "配额", "关税", "政策", "禁令", "预警", "大雨", "暴雨", "干旱", "洪涝",
    "甘蔗", "甜菜", "糖厂", "产糖", "压榨", "库存", "报价", "进口", "出口", "乙醇",
    "糖蜜", "糖浆", "燃料级乙醇", "E5", "E10", "E20",
)
IMPACT_TARGET_TERMS = (
    "供应", "需求", "库存", "进口", "出口", "贸易", "糖料", "甘蔗", "甜菜", "糖厂",
    "压榨", "产量", "产糖", "制糖", "乙醇", "糖蜜", "糖浆", "成本", "现货", "糖价",
    "原糖", "白糖", "郑糖", "供需",
)
IMPACT_CAUSAL_TERMS = (
    "因此", "从而", "将", "会", "有助于", "可能", "导致", "使", "压制", "支撑",
    "削弱", "提高", "降低", "增加", "减少", "稳定", "扩大", "缓解", "扰动",
    "形成压力", "限制", "补充", "改变", "反映", "利多", "利空", "中性",
)
CONCRETE_DETAIL_RE = re.compile(
    r"\d+(?:[.,]\d+)?\s*(?:%|万吨|吨|千吨|公担|卢比|雷亚尔|美元|美分|元/吨|mm|KL|kl|升|家|座|万亿|亿|万)?"
)

def project_display_path(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(PROJECT_ROOT.resolve())).replace("\\", "/")
    except ValueError:
        return str(resolved).replace("\\", "/")


def load_editorial_skill_metadata() -> dict:
    if not EDITORIAL_SKILL_PATH.exists():
        raise FileNotFoundError(f"Missing Sugar News editorial skill: {EDITORIAL_SKILL_PATH}")
    content = EDITORIAL_SKILL_PATH.read_text(encoding="utf-8")
    required_groups = {
        "summary_2_3": ("2-3", "sentences"),
        "date_expression": ("publication dates", "YYYY-MM-DD"),
        "country_assignment": ("Country Assignment", "Indonesia"),
        "medical_filter": ("blood sugar", "血糖"),
        "china_daily_monitoring": (
            "中国糖业新闻每日重点监测",
            "食糖进口",
            "郑糖",
            "糖料产区",
            "China search is mandatory",
            "public placeholder",
        ),
        "brazil_metrics_daily": ("巴西糖价与库存每日刷新", "brazil_sugar_metrics.py", "Vercel"),
        "pre_publish": ("Pre-Publish Quality Checks", "Stop publication"),
        "concrete_news_summary": ("who did what", "concrete change", "消息涉及", "media outlet as the event subject"),
        "summary_style_anchor": ("2026-08-16", "2026-08-17", "2026-08-18", "standing style anchor", "automatic-vs-rewrite corrections"),
        "global_highlights": ("全球糖业新闻重点", "15美分/磅", "2-3 Chinese sentences"),
        "brazil_hedging_monitoring": ("Brazil sugar hedging progress", "巴西糖厂套保", "巴西糖套保进度"),
        "india_impact_overrides": ("100%` sugar import duty", "record highs", "Sugarcane acreage increases are bearish"),
        "impact_marker": ("影响：利多糖价", "影响：利空糖价", "影响：中性"),
        "country_order": ("巴西", "印度", "泰国", "其他国家", "中国"),
        "two_stage_search": ("Two-Stage Search And Candidate Verification", "Country source matrix", "latest 36 hours", "Required regression topics"),
    }
    missing = [
        name
        for name, phrases in required_groups.items()
        if not all(phrase in content for phrase in phrases)
    ]
    if missing:
        raise ValueError(f"Sugar News editorial skill missing required rules: {missing}")
    return {
        "path": project_display_path(EDITORIAL_SKILL_PATH),
        "sha256": hashlib.sha256(content.encode("utf-8")).hexdigest(),
    }


INDIA_MAIN_CANE_REGIONS = (
    "北方邦", "Uttar Pradesh", "UP",
    "马哈拉施特拉邦", "Maharashtra",
    "卡纳塔克邦", "Karnataka",
    "泰米尔纳德邦", "Tamil Nadu",
    "古吉拉特邦", "Gujarat",
    "比哈尔邦", "Bihar",
    "旁遮普邦", "Punjab",
    "哈里亚纳邦", "Haryana",
    "北阿坎德邦", "Uttarakhand",
)
INDIA_WEATHER_TERMS = (
    "降雨", "雨", "季风", "天气", "气象", "预警", "强降雨", "暴雨", "干旱", "洪涝", "积水",
    "rain", "rainfall", "monsoon", "weather", "alert", "warning", "heavy rain", "flood", "drought",
)
INDIA_RAIN_BENEFIT_TERMS = (
    "适量降雨", "降雨增加", "降雨增多", "未来降雨", "强降雨", "暴雨预报", "墒情改善",
    "季风活跃", "季风增强", "widespread rainfall", "heavy rainfall", "rainfall forecast",
    "monsoon revival", "active monsoon",
)
INDIA_DROUGHT_TERMS = (
    "干旱", "降雨不足", "季风偏弱", "降雨减少", "雨量不足", "deficient rainfall",
    "weak monsoon", "dry spell", "rainfall deficit",
)
INDIA_WATER_STRESS_TERMS = (
    "水资源压力", "水分约束", "缺水", "水资源约束", "灌溉不足", "半干旱",
    "water stress", "water risks", "water constraint", "water scarcity", "semi-arid",
)
INDIA_DAMAGE_TERMS = (
    "已造成", "洪涝", "农田被淹", "甘蔗倒伏", "道路中断", "作物受损", "预计减产",
    "受灾", "损失", "红腐病", "病害", "虫害", "flood damage", "crop damage",
    "waterlogging", "lodging", "road disruption", "red rot", "pest", "disease",
)
INDIA_HARVEST_TERMS = ("收割", "压榨", "运输", "入榨", "开榨", "砍蔗", "harvest", "crushing", "transport")
INDIA_INDIRECT_ETHANOL_POLICY_TERMS = (
    "e20", "ethanol blend", "ethanol blending", "ethanol mix", "blend in gasoline",
    "gasoline", "petrol", "biofuel", "oil ministry", "omc", "oil marketing company",
    "ethanol procurement", "ethanol tender", "distillery",
)
INDIA_INDIRECT_FEEDSTOCK_TERMS = (
    "sugarcane juice", "cane juice", "cane-based", "sugar syrup", "syrup",
    "molasses", "b-heavy", "c-heavy", "maize", "corn", "grain", "rice",
    "broken rice", "feedstock", "गन्ना", "शीरा", "मक्का", "इथेनॉल",
)
INDIA_PRICE_INVENTORY_SEARCH_TEMPLATES = (
    ("en", "India sugar S-grade M-grade domestic price {readable}"),
    ("en", "India sugar prices today S grade M grade {readable}"),
    ("en", "Maharashtra Uttar Pradesh Karnataka sugar price {readable}"),
    ("en", "Uttar Pradesh sugar ex-mill price {readable}"),
    ("en", "UP sugar ex-mill rate {readable}"),
    ("en", "Uttar Pradesh mill sugar price {readable}"),
    ("en", "M-grade sugar ex-mill Uttar Pradesh {readable}"),
    ("en", "North India sugar ex-mill price {readable}"),
    ("hi", "मुजफ्फरनगर चीनी मिल भाव {day} जुलाई {year}"),
    ("hi", "उत्तर प्रदेश चीनी एक्स मिल कीमत {day} जुलाई {year}"),
    ("en", "India sugar carryover stock ending stock {readable}"),
    ("en", "India sugar closing stock ISMA NFCSF {readable}"),
    ("en", "India sugar ending stocks consumption ratio {readable}"),
)
INDIA_PRICE_INVENTORY_SOURCE_GUIDE = (
    "ChiniMandi domestic sugar prices",
    "ISMA / Indian Sugar & Bio-energy Manufacturers Association",
    "NFCSF / National Federation of Cooperative Sugar Factories",
    "Department of Food and Public Distribution, Government of India",
    "reliable commodity and agriculture media with dated market quotes",
)
DATE_FORMAT_EXAMPLES = (
    "July 19, 2026",
    "19 July 2026",
    "19/07/2026",
    "2026-07-19",
    "19 de julho de 2026",
    "19 जुलाई 2026",
    "19 กรกฎาคม 2569",
    "19 جولائی 2026",
    "Hulyo 19 2026",
    "ngày 19 tháng 7 năm 2026",
    "19 июля 2026",
    "19 Juli 2026",
)
THAI_MAIN_CANE_PROVINCES = (
    "\u4e4c\u9686\u4ed6\u5c3c", "Udon Thani",
    "\u5b54\u656c", "Khon Kaen",
    "\u5475\u53fb", "\u90a3\u7a7a\u53fb\u5dee\u662f\u739b", "Nakhon Ratchasima",
    "\u731c\u4e5f\u84ec", "Chaiyaphum",
    "\u52a0\u62c9\u4fe1", "Kalasin",
    "\u9ece\u5e9c", "Loei",
    "\u90a3\u7a7a\u6c99\u65fa", "Nakhon Sawan",
    "\u7518\u70f9\u78a7", "Kamphaeng Phet",
    "\u7d20\u53ef\u6cf0", "Sukhothai",
    "\u5f6d\u4e16\u6d1b", "Phitsanulok",
    "\u5317\u78a7", "Kanchanaburi",
    "\u534e\u5bcc\u91cc", "Lopburi",
    "\u7d20\u6500\u6b66\u91cc", "Suphanburi",
    "\u731c\u7eb3", "Chai Nat",
    "\u6c99\u7f34", "Sa Kaeo",
    "\u6625\u6b66\u91cc", "Chonburi",
)
THAI_WEATHER_TERMS = (
    "\u5929\u6c14", "\u6c14\u8c61", "\u964d\u96e8", "\u96e8", "\u96f7\u9635\u96e8",
    "\u5e72\u65f1", "\u6d2a\u6d9d", "\u79ef\u6c34", "rain", "rainfall", "thunderstorm", "flood", "drought",
)
THAI_RAIN_INCREASE_TERMS = (
    "\u964d\u96e8\u589e\u52a0", "\u964d\u96e8\u5c06\u589e\u52a0", "\u96e8\u91cf\u589e\u52a0",
    "\u964d\u96e8\u589e\u591a", "\u964d\u96e8\u660e\u663e\u589e\u591a", "\u964d\u96e8\u6539\u5584",
    "\u5892\u60c5", "\u6709\u5229\u4e8e\u6539\u5584", "\u5f3a\u964d\u96e8",
    "\u5f3a\u5230\u5f88\u5f3a\u964d\u96e8", "\u66b4\u96e8\u9884\u8b66", "\u964d\u96e8\u8303\u56f4",
    "\u964d\u96e8\u5f3a\u5ea6",
)
THAI_LOW_COVERAGE_TERMS = (
    "20%", "\u7ea620%", "\u8f83\u5206\u6563", "\u5206\u6563", "\u8986\u76d6\u7387\u8f83\u4f4e",
    "\u8986\u76d6\u7387\u4f4e", "\u5c40\u5730", "\u5c11\u91cf",
)
THAI_DAMAGE_TERMS = (
    "\u5df2\u9020\u6210", "\u9020\u6210\u4e25\u91cd\u6d2a\u6d9d", "\u4e25\u91cd\u6d2a\u6d9d",
    "\u7518\u8517\u5012\u4f0f", "\u519c\u7530\u88ab\u6df9", "\u4f5c\u7269\u53d7\u635f",
    "\u9884\u8ba1\u51cf\u4ea7", "\u53d7\u707e", "\u635f\u5931", "\u6839\u7cfb\u53d7\u635f",
)
THAI_DROUGHT_TERMS = ("\u964d\u96e8\u51cf\u5c11", "\u964d\u96e8\u4e0d\u8db3", "\u6301\u7eed\u5e72\u65f1", "\u5e72\u65f1", "\u504f\u5e72")
THAI_HARVEST_TERMS = ("\u6536\u5272", "\u538b\u69a8", "\u8fd0\u8f93", "\u5165\u69a8", "\u5f00\u69a8", "\u6536\u69a8")
THAI_WEATHER_EVENT_TERMS = (
    "\u964d\u96e8", "\u96e8\u91cf", "\u96f7\u9635\u96e8", "\u5e72\u65f1", "\u6d2a\u6d9d", "\u79ef\u6c34",
    "rain", "rainfall", "thunderstorm", "flood", "drought",
)
GLOBAL_SEARCH_TEMPLATES = (
    "global sugar industry news {readable}",
    "sugar production export policy {readable}",
    "sugarcane ethanol mills {readable}",
    "sugar import export tariff quota {readable}",
    "sugar price government policy {readable}",
    "sugar industry news {day} {month_name} {year}",
    "sugarcane news {day} {month_name} {year}",
    "ethanol sugar mills {day} {month_name} {year}",
)
PORTUGUESE_MONTH_NAMES = {
    1: "janeiro",
    2: "fevereiro",
    3: "março",
    4: "abril",
    5: "maio",
    6: "junho",
    7: "julho",
    8: "agosto",
    9: "setembro",
    10: "outubro",
    11: "novembro",
    12: "dezembro",
}
OTHER_COUNTRY_SEARCH_TEMPLATES = (
    ("en", "Indonesia sugar industry {readable}"),
    ("en", "Pakistan sugar industry {readable}"),
    ("en", "Philippines sugar industry {readable}"),
    ("en", "Vietnam sugar industry {readable}"),
    ("en", "Russia sugar beet sugar {readable}"),
    ("en", "Cameroon sugar industry {readable}"),
    ("en", "EU sugar beet production {readable}"),
    ("en", "United States sugar beet cane {readable}"),
    ("en", "Mexico sugar production export {readable}"),
)
COUNTRY_SEARCH_TEMPLATES = {
    "巴西": (
        ("en", "Brazil sugar industry news {readable}"),
        ("en", "Brazil sugarcane ethanol export {readable}"),
        ("en", "Brazil corn ethanol Renovabio {readable}"),
        ("en", "Brazil ethanol gasoline corn ethanol {readable}"),
        ("en", "Brazil sugar hedging progress mills hedged sugar {readable}"),
        ("zh-CN", "巴西糖厂套保 食糖套保比例 {year}年{month}月{day}日"),
        ("zh-CN", "巴西 玉米乙醇 Renovabio {year}年{month}月{day}日"),
        ("zh-CN", "巴西 乙醇 汽油 {year}年{month}月{day}日"),
        ("pt-BR", "Brasil açúcar etanol {day} {month_name_pt} {year}"),
        ("pt-BR", "Brasil setor sucroenergético {day} de {month_name_pt} de {year}"),
        ("pt-BR", "Brasil etanol de milho RenovaBio {day} de {month_name_pt} de {year}"),
        ("pt-BR", "Brasil etanol gasolina etanol de milho {day} de {month_name_pt} de {year}"),
        ("pt-BR", "usinas cana açúcar etanol {date_slash}"),
    ),
    "印度": (
        ("en", "India sugar industry news {readable}"),
        ("en", "India sugarcane ethanol mills {readable}"),
        ("en", "India sugar news {day} {month_name} {year}"),
        ("en", "India sugar production {readable}"),
        ("en", "India sugar stocks {readable}"),
        ("en", "India sugar prices {readable}"),
        ("en", "India sugar ex-mill price {readable}"),
        ("en", "India sugar export policy {readable}"),
        ("en", "India sugar import {readable}"),
        ("en", "India sugar sales quota {readable}"),
        ("en", "India sugar shortage {readable}"),
        ("en", "India sugar mills {readable}"),
        ("en", "India sugarcane production {readable}"),
        ("en", "India sugarcane acreage {readable}"),
        ("en", "India sugarcane FRP {readable}"),
        ("en", "India ethanol policy {readable}"),
        ("en", "India ethanol blending {readable}"),
        ("en", "India E20 petrol {readable}"),
        ("en", "India E20 ethanol target {readable}"),
        ("en", "India ethanol above 20 percent {readable}"),
        ("en", "India sugarcane ethanol {readable}"),
        ("en", "India molasses ethanol {readable}"),
        ("en", "India sugar syrup ethanol {readable}"),
        ("en", "India grain ethanol {readable}"),
        ("en", "India maize ethanol {readable}"),
        ("en", "India ethanol feedstock {readable}"),
        ("en", "India oil ministry ethanol {readable}"),
        ("en", "India OMC ethanol tender {readable}"),
        ("en", "India cane-based distillery {readable}"),
        ("en", "site:reuters.com India sugar {readable}"),
        ("en", "site:reuters.com India ethanol {readable}"),
        ("en", "site:reuters.com India E20 {readable}"),
        ("en", "site:reuters.com India sugarcane {readable}"),
        ("en", "site:reuters.com India molasses {readable}"),
        ("en", "India sugarcane rainfall {readable}"),
        ("en", "India sugar belt rainfall forecast {readable}"),
        ("en", "Uttar Pradesh sugarcane rain forecast {readable}"),
        ("en", "Maharashtra sugarcane rainfall {readable}"),
        ("en", "Karnataka sugarcane rainfall {readable}"),
        ("en", "India monsoon sugar production {readable}"),
        ("en", "IMD rainfall forecast sugarcane states {readable}"),
        ("en", "heavy rainfall sugarcane India {readable}"),
        ("en", "excess rainfall cane crop India {readable}"),
        ("en", "deficient rainfall sugarcane India {readable}"),
        ("hi", "भारत चीनी उद्योग {day} जुलाई {year}"),
        ("hi", "गन्ना चीनी मिल इथेनॉल {day} जुलाई {year}"),
        ("hi", "भारत चीनी उत्पादन {day} जुलाई {year}"),
        ("hi", "भारत चीनी कीमत {day} जुलाई {year}"),
        ("hi", "ई20 पेट्रोल {day} जुलाई {year}"),
        ("hi", "भारत इथेनॉल ब्लेंडिंग {day} जुलाई {year}"),
        ("hi", "मक्के से इथेनॉल {day} जुलाई {year}"),
        ("hi", "शीरा इथेनॉल {day} जुलाई {year}"),
        ("hi", "उत्तर प्रदेश गन्ना बारिश {day} जुलाई {year}"),
        ("hi", "महाराष्ट्र गन्ना बारिश {day} जुलाई {year}"),
        ("hi", "कर्नाटक गन्ना बारिश {day} जुलाई {year}"),
    ),
    "泰国": (
        ("en", "Thailand sugar industry news {readable}"),
        ("en", "Thailand sugarcane mills ethanol {readable}"),
        ("en", "Thailand sugar news {day} {month_name} {year}"),
        ("en", "Thailand sugarcane rainfall forecast {readable}"),
        ("en", "Udon Thani Khon Kaen Nakhon Ratchasima sugarcane rain forecast {readable}"),
        ("en", "Nakhon Sawan Kanchanaburi Lopburi Chai Nat rainfall forecast {readable}"),
        ("en", "Thailand cane growing areas thunderstorm heavy rain drought {readable}"),
        ("th", "ประเทศไทย น้ำตาล อ้อย {day} กรกฎาคม {buddhist_year}"),
        ("th", "ข่าวอ้อย น้ำตาล {day} กรกฎาคม {buddhist_year}"),
        ("th", "อุตสาหกรรมอ้อยและน้ำตาล {day} กรกฎาคม {buddhist_year}"),
        ("th", "โรงงานน้ำตาล เอทานอล {day} กรกฎาคม {buddhist_year}"),
    ),
    "中国": (
        ("zh-CN", "中国糖业新闻 {year}年{month}月{day}日"),
        ("zh-CN", "中国食糖 {year}年{month}月{day}日"),
        ("zh-CN", "中国白糖 {year}年{month}月{day}日"),
        ("zh-CN", "中国甘蔗 {year}年{month}月{day}日"),
        ("zh-CN", "中国甜菜糖 {year}年{month}月{day}日"),
        ("zh-CN", "广西糖业 甘蔗 {year}年{month}月{day}日"),
        ("zh-CN", "云南糖业 甘蔗 {year}年{month}月{day}日"),
        ("zh-CN", "广东 甘蔗 糖业 {year}年{month}月{day}日"),
        ("zh-CN", "海南 甘蔗 糖业 {year}年{month}月{day}日"),
        ("zh-CN", "内蒙古 新疆 黑龙江 甜菜糖 {year}年{month}月{day}日"),
        ("zh-CN", "食糖产销数据 {year}年{month}月{day}日"),
        ("zh-CN", "食糖库存 产销率 {year}年{month}月{day}日"),
        ("zh-CN", "食糖进口 {year}年{month}月{day}日"),
        ("zh-CN", "原糖进口 到港量 进口成本 {year}年{month}月{day}日"),
        ("zh-CN", "食糖进口配额 {year}年{month}月{day}日"),
        ("zh-CN", "糖浆预混粉进口 {year}年{month}月{day}日"),
        ("zh-CN", "加工糖厂 开机 {year}年{month}月{day}日"),
        ("zh-CN", "广西糖业 {year}年{month}月{day}日"),
        ("zh-CN", "云南糖业 {year}年{month}月{day}日"),
        ("zh-CN", "郑州白糖期货 {year}年{month}月{day}日"),
        ("zh-CN", "郑糖主力合约 {year}年{month}月{day}日"),
        ("zh-CN", "白糖现货价格 {year}年{month}月{day}日"),
        ("zh-CN", "甘蔗收购价 种植补贴 {year}年{month}月{day}日"),
        ("zh-CN", "糖厂开榨 收榨 {year}年{month}月{day}日"),
        ("zh-CN", "糖料产区 降雨 暴雨 干旱 台风 {year}年{month}月{day}日"),
        ("zh-CN", "制糖集团公告 {year}年{month}月{day}日"),
        ("en", "China sugar industry {readable}"),
        ("en", "China sugar production {readable}"),
        ("en", "China sugar imports {readable}"),
        ("en", "China sugarcane beet sugar {readable}"),
        ("en", "China white sugar futures {readable}"),
        ("en", "China sugar syrup imports {readable}"),
    ),
    "巴基斯坦": (
        ("en", "Pakistan sugar industry {readable}"),
        ("en", "Pakistan sugarcane sugar mills {readable}"),
        ("en", "Pakistan sugar export import price {readable}"),
        ("en", "Pakistan sugar policy {day} {month_name} {year}"),
        ("ur", "پاکستان چینی صنعت {day} جولائی {year}"),
        ("ur", "گنا چینی ملز {day} جولائی {year}"),
        ("ur", "چینی برآمد درآمد {day} جولائی {year}"),
    ),
    "菲律宾": (
        ("en", "Philippines sugar industry {readable}"),
        ("en", "Philippines sugar production import {readable}"),
        ("en", "Philippines Sugar Regulatory Administration {readable}"),
        ("en", "Philippines sugarcane mills {day} {month_name} {year}"),
        ("fil", "industriya ng asukal Hulyo {day} {year}"),
        ("fil", "produksyon ng tubo at asukal Hulyo {day} {year}"),
        ("fil", "importasyon ng asukal Hulyo {day} {year}"),
    ),
    "越南": (
        ("en", "Vietnam sugar industry {readable}"),
        ("en", "Vietnam sugar import tariff {readable}"),
        ("en", "Vietnam sugarcane production {readable}"),
        ("en", "Vietnam sugar anti-dumping {day} {month_name} {year}"),
        ("vi", "ngành đường Việt Nam ngày {day} tháng 7 năm {year}"),
        ("vi", "mía đường Việt Nam {date_slash}"),
        ("vi", "nhập khẩu đường {date_slash}"),
        ("vi", "thuế chống bán phá giá đường {date_slash}"),
    ),
    "俄罗斯": (
        ("en", "Russia sugar industry {readable}"),
        ("en", "Russia sugar beet production {readable}"),
        ("en", "Russia sugar export price {readable}"),
        ("en", "Russian sugar market {day} {month_name} {year}"),
        ("ru", "сахарная промышленность России {day} июля {year}"),
        ("ru", "сахарная свекла {day} июля {year}"),
        ("ru", "производство сахара Россия {day} июля {year}"),
        ("ru", "экспорт сахара Россия {day} июля {year}"),
    ),
    "印度尼西亚": (
        ("en", "Indonesia sugar industry {readable}"),
        ("en", "Indonesia sugar import production {readable}"),
        ("en", "Indonesia sugar self-sufficiency {readable}"),
        ("en", "Indonesia sugarcane mills {day} {month_name} {year}"),
        ("id", "industri gula Indonesia {day} Juli {year}"),
        ("id", "produksi gula dan tebu {day} Juli {year}"),
        ("id", "impor gula Indonesia {day} Juli {year}"),
        ("id", "swasembada gula {day} Juli {year}"),
        ("id", "pabrik gula {day} Juli {year}"),
    ),
}

SOURCE_MATRIX = {
    "巴西": (
        "Ministério de Minas e Energia / CNPE",
        "MAPA",
        "ANP",
        "UNICA",
        "Datagro",
        "Archer Consulting",
        "StoneX",
        "Hedgepoint",
        "Czarnikow / Czapp",
        "Conab",
        "NovaCana",
        "CanaOnline",
        "Reuters",
        "Brazil port and trade data",
    ),
    "印度": (
        "Press Information Bureau",
        "Lok Sabha / Rajya Sabha questions",
        "Department of Food and Public Distribution",
        "Ministry of Petroleum and Natural Gas",
        "Ministry of Agriculture",
        "ISMA",
        "NFCSF",
        "state cane commissioner offices",
        "IMD",
        "Vasantdada Sugar Institute",
        "ChiniMandi",
        "local newspapers",
    ),
    "泰国": (
        "OCSB",
        "Ministry of Industry Thailand",
        "Thai cane grower and sugar associations",
        "Thai Meteorological Department",
        "Thai government announcements",
        "The Nation Thailand",
        "Bangkok Post",
        "Thai agricultural media",
    ),
    "中国": (
        "农业农村部 / CASDE",
        "海关总署",
        "国家统计局",
        "商务部",
        "中国糖业协会",
        "广西糖业协会",
        "云南省糖业协会",
        "云糖网",
        "泛糖科技",
        "沐甜科技",
        "郑州商品交易所",
        "淀粉糖行业数据",
        "制糖集团公告",
    ),
    "其他国家": (
        "Indonesia government / agriculture / industry ministries",
        "USDA / EIA / American Sugar Alliance",
        "Philippines Sugar Regulatory Administration",
        "Pakistan government / PSMA",
        "Vietnam sugar association",
        "Russia beet and sugar statistics",
        "Fiji Sugar Corporation",
        "Nepal sugar mills and cane dues sources",
        "European Commission and beet sugar bodies",
    ),
}

ADDITIONAL_COUNTRY_SEARCH_TEMPLATES = {
    "巴西": (
        ("pt-BR", "site:gov.br etanol gasolina mistura anidro {readable}"),
        ("pt-BR", "CNPE mistura etanol anidro gasolina {readable}"),
        ("pt-BR", "ANP etanol gasolina Brasil {readable}"),
        ("pt-BR", "ANP estoque de etanol hidratado anidro {readable}"),
        ("pt-BR", "UNICA moagem cana produção de açúcar etanol {readable}"),
        ("pt-BR", "Datagro consumo global de açúcar {readable}"),
        ("pt-BR", "Conab cana-de-açúcar produção açúcar {readable}"),
        ("pt-BR", "NovaCana etanol usina cana-de-açúcar {readable}"),
        ("pt-BR", "CanaOnline usina moagem cana açúcar etanol {readable}"),
        ("pt-BR", "preço da cana usina açúcar etanol {readable}"),
        ("pt-BR", "exportação de açúcar porto Brasil {readable}"),
        ("pt-BR", "hedge açúcar usinas fixação contratos exportação {readable}"),
        ("en", "Brazil sugar hedge ratio fixed-price sales export contracts {readable}"),
        ("en", "Brazil ethanol blending E30 E32 sugarcane {readable}"),
        ("en", "Brazil ethanol gasoline RenovaBio corn ethanol {readable}"),
        ("en", "Brazil UNICA sugarcane crushing sugar production ethanol {readable}"),
        ("en", "Brazil Datagro global sugar consumption forecast {readable}"),
    ),
    "印度": (
        ("en", "site:pib.gov.in ethanol blending sugarcane molasses {readable}"),
        ("en", "site:pib.gov.in sugarcane ethanol foreign exchange savings {readable}"),
        ("en", "site:pib.gov.in ethanol petrol pumps coverage {readable}"),
        ("en", "site:pib.gov.in ethanol procurement price sugarcane juice {readable}"),
        ("en", "Lok Sabha sugar ethanol molasses question {readable}"),
        ("en", "Rajya Sabha ethanol capacity molasses sugarcane {readable}"),
        ("en", "India cane price cane dues sugar mill {readable}"),
        ("en", "India sugar mill operational factories cane requirement {readable}"),
        ("en", "Vasantdada Sugar Institute sugarcane variety yield {readable}"),
        ("en", "India red rot white grub sugarcane {readable}"),
        ("en", "India drought pest sugarcane acreage {readable}"),
        ("en", "Uttar Pradesh cane price SAP sugar mill {readable}"),
        ("en", "Maharashtra cane acreage sugarcane drought pest {readable}"),
        ("en", "Karnataka sugarcane red rot white grub {readable}"),
    ),
    "泰国": (
        ("en", "OCSB Thailand sugarcane area cane price {readable}"),
        ("en", "Thailand cane price sugar mill crushing season {readable}"),
        ("en", "Thailand cassava replaces sugarcane planting area {readable}"),
        ("en", "Thailand sugarcane white leaf disease {readable}"),
        ("en", "Thailand sugar export OCSB {readable}"),
        ("th", "สอน. พื้นที่ปลูกอ้อย ราคาอ้อย {day} กรกฎาคม {buddhist_year}"),
        ("th", "โรคใบขาวอ้อย โรงงานน้ำตาล {day} กรกฎาคม {buddhist_year}"),
        ("th", "ราคาอ้อย เปิดหีบ โรงงานน้ำตาล {day} กรกฎาคม {buddhist_year}"),
    ),
    "中国": (
        ("zh-CN", "淀粉糖 玉米消耗量 产能利用率 {year}年{month}月{day}日"),
        ("zh-CN", "玉米糖浆 果葡糖浆 白糖价差 {year}年{month}月{day}日"),
        ("zh-CN", "糖浆 白砂糖预混粉 进口 {year}年{month}月{day}日"),
        ("zh-CN", "广西 食糖销量 工业库存 产销率 {year}年{month}月{day}日"),
        ("zh-CN", "云南 甘蔗 食糖 产量 {year}年{month}月{day}日"),
        ("zh-CN", "中国糖业协会 食糖产销库存 {year}年{month}月{day}日"),
        ("zh-CN", "CASDE 食糖 供需 月报 {year}年{month}月{day}日"),
        ("zh-CN", "海关 食糖进口 糖浆预混粉 {year}年{month}月{day}日"),
    ),
    "印度尼西亚": (
        ("en", "Indonesia raw sugar import allocation refinery regulation {readable}"),
        ("id", "Peraturan Presiden impor gula mentah industri rafinasi {day} Juli {year}"),
        ("id", "Kementerian Perindustrian gula rafinasi impor gula mentah {day} Juli {year}"),
        ("id", "kebijakan etanol tebu Indonesia {day} Juli {year}"),
    ),
    "美国": (
        ("en", "EIA ethanol production stocks week ending {readable}"),
        ("en", "USDA sugar beet cane sugar production imports {readable}"),
        ("en", "United States sugar demand corn syrup consumption {readable}"),
        ("en", "American Sugar Alliance beet cane sugar {readable}"),
    ),
    "菲律宾": (
        ("en", "Philippines sugarcane pest provinces planters {readable}"),
        ("en", "Sugar Regulatory Administration sugar import policy {readable}"),
        ("en", "Philippines cane farmers aid sugarcane {readable}"),
        ("en", "Negros sugar mill sugarcane pest {readable}"),
    ),
    "尼泊尔": (
        ("en", "Nepal sugar mills cane dues import sugar {readable}"),
        ("en", "Nepal sugarcane farmers payment arrears {readable}"),
    ),
    "斐济": (
        ("en", "Fiji Sugar Corporation crushing cane price {readable}"),
        ("en", "Fiji sugarcane growers cane payment {readable}"),
    ),
    "欧洲": (
        ("en", "European Commission sugar beet production exports {readable}"),
        ("en", "France sugar beet production disease {readable}"),
        ("en", "UK British Sugar beet crop {readable}"),
    ),
}

CASE_REGRESSION_TOPICS = (
    "brazil_ethanol_blend_policy",
    "brazil_global_sugar_consumption_forecast",
    "brazil_sugar_hedging_progress",
    "india_ethanol_tax_or_fiscal_data",
    "india_ethanol_crude_import_savings",
    "india_ethanol_pump_coverage",
    "india_ethanol_capacity",
    "india_ethanol_procurement_price",
    "india_sugarcane_variety_trial",
    "india_local_cane_drought_or_pest",
    "india_mill_count_running_and_cane_requirement",
    "indonesia_raw_sugar_import_policy",
    "us_eia_ethanol_production_and_stocks",
    "philippines_cane_farmer_aid",
    "yunnan_cane_and_sugar_output",
    "china_starch_sugar_corn_use_capacity",
    "guangxi_sugar_sales_and_industrial_inventory",
)

NEWS_TOPIC_RULES = (
    ("ethanol_capacity", ("ethanol capacity", "ethanol programme", "ethanol program", "foreign exchange savings", "crude import", "distillery", "ethanol production", "ethanol stocks", "乙醇产能", "乙醇产量", "乙醇库存", "采购价", "procurement price", "molasses", "sugarcane juice")),
    ("ethanol_policy", ("e20", "e27", "e30", "e32", "blend", "blending", "gasoline", "petrol", "biofuel", "ethanol policy", "etanol", "乙醇政策", "掺混", "加油站")),
    ("mill_operations", ("sugar mill", "mills", "factory", "crushing", "operational", "running", "cane requirement", "开榨", "收榨", "糖厂", "压榨", "运行")),
    ("cane_farming", ("cane price", "cane dues", "cane acreage", "sugarcane area", "farmer", "planter", "甘蔗收购价", "种植面积", "蔗农", "蔗款")),
    ("variety_research", ("variety", "trial", "yield", "sugar content", "vasantdada", "research institute", "新品种", "试验", "单产", "糖分")),
    ("weather_pest", ("rain", "rainfall", "monsoon", "drought", "flood", "pest", "disease", "red rot", "white grub", "white leaf", "降雨", "干旱", "病虫害", "红腐病", "白叶病", "白蛴螬")),
    ("supply_demand", ("production", "output", "stocks", "inventory", "sales quota", "consumption", "forecast", "产量", "库存", "销量", "产销率", "消费", "预测")),
    ("trade_policy", ("import", "export", "tariff", "quota", "raw sugar", "refinery", "进口", "出口", "配额", "关税", "原糖", "精炼")),
    ("starch_sugar_substitute", ("starch sugar", "corn syrup", "hfcs", "glucose syrup", "淀粉糖", "玉米糖浆", "果葡糖浆", "预混粉", "玉米消耗")),
    ("price_market", ("price", "prices", "ex-mill", "wholesale", "retail", "futures", "basis", "hedge", "hedged", "hedging", "fixed-price", "fixação", "fixacao", "export contracts", "报价", "现货", "期货", "出厂价", "套保", "固定价销售")),
)

TOPIC_LABELS = {
    "ethanol_policy": "乙醇掺混或燃料政策",
    "ethanol_capacity": "乙醇产能、产量或采购价格",
    "mill_operations": "糖厂运行和压榨安排",
    "cane_farming": "甘蔗种植、蔗价或蔗农现金流",
    "variety_research": "甘蔗品种试验和技术推广",
    "weather_pest": "甘蔗产区天气、干旱或病虫害",
    "supply_demand": "食糖产量、库存、销量或消费预测",
    "trade_policy": "食糖贸易、进口、出口或配额政策",
    "starch_sugar_substitute": "中国替代糖源和淀粉糖供需",
    "price_market": "食糖价格和市场流通",
    "general_industry": "糖业运行",
}

COUNTRY_TRUSTED_SOURCE_MARKERS = {
    "巴西": (
        "unica", "datagro", "novacana", "canaonline", "conab", "anp",
        "gov.br", "reuters", "ethanol producer", "cepea", "archerconsulting",
        "archer consulting", "hedgepoint", "stonex", "czapp", "czarnikow",
    ),
    "印度": (
        "chinimandi", "pib", "press information bureau", "the economic times",
        "times of india", "business standard", "financial express",
        "hindustan times", "moneycontrol", "agrospectrum india", "isma",
        "nfcsf", "vasantdada", "deccan herald", "the hindu",
    ),
    "泰国": (
        "ocsb", "thai meteorological", "bangkok post", "the nation thailand",
        "prachachat", "thaiger",
    ),
    "中国": (
        "云糖网", "沐甜科技", "泛糖科技", "中国糖业协会", "海关总署",
        "农业农村部", "郑州商品交易所", "广西糖业", "云南糖业",
    ),
    "美国": ("eia", "usda", "american sugar alliance", "u.s. energy information"),
    "菲律宾": ("inquirer", "sugar regulatory administration", "sra", "philstar", "businessworld"),
    "印度尼西亚": ("antaranews", "kompas", "tempo", "kementerian", "indonesia business post"),
    "巴基斯坦": ("psma", "dawn", "business recorder", "profit by pakistan today"),
    "越南": ("vietnam", "vietnamnet", "vietnamplus", "viet nam news"),
    "俄罗斯": ("interfax", "tass", "sugar.ru", "ikar"),
    "斐济": ("fiji sugar corporation", "fbc news", "fiji times"),
    "尼泊尔": ("kathmandu post", "my republica", "nepal"),
    "欧洲": ("european commission", "british sugar", "franceagrimer"),
}

RSS_TITLE_ONLY_REJECT_TERMS = (
    "latest results", "quarterly results", "pat rises", "profit after tax",
    "net profit", "share price", "shares", "stock market", "stock to watch",
    "buy rating", "target price", "market cap",
)

LOW_SIGNAL_ETHANOL_DISCUSSION_TERMS = (
    "choice between pure petrol", "petrol-e20 choice", "pure petrol",
    "lower fuel prices", "town hall", "rice millers", "benefits, challenges, concerns",
)

COUNTRY_ALIASES.update({
    "美国": COUNTRY_ALIASES.get("美国", ()) + ("eia", "usda", "american sugar alliance", "u.s. sugar", "us sugar"),
    "菲律宾": COUNTRY_ALIASES.get("菲律宾", ()) + ("mindanao", "negros", "sugar regulatory administration", "sra", "planters", "palace monitoring sugar"),
    "尼泊尔": COUNTRY_ALIASES.get("尼泊尔", ()) + ("nepal", "nepali", "cane dues"),
    "斐济": COUNTRY_ALIASES.get("斐济", ()) + ("fiji sugar corporation", "fsc"),
    "欧洲": COUNTRY_ALIASES.get("欧洲", ()) + ("european commission", "france sugar beet", "british sugar", "uk beet"),
    "乌克兰": COUNTRY_ALIASES.get("乌克兰", ()) + ("ukraine", "ukrainian", "agroportal.ua"),
})

def merge_search_templates(
    base: dict[str, tuple[tuple[str, str], ...]],
    extra: dict[str, tuple[tuple[str, str], ...]],
) -> dict[str, tuple[tuple[str, str], ...]]:
    merged = {country: list(templates) for country, templates in base.items()}
    for country, templates in extra.items():
        existing = merged.setdefault(country, [])
        seen = {template for _language, template in existing}
        for language, template in templates:
            if template not in seen:
                existing.append((language, template))
                seen.add(template)
    return {country: tuple(templates) for country, templates in merged.items()}


COUNTRY_SEARCH_TEMPLATES = merge_search_templates(COUNTRY_SEARCH_TEMPLATES, ADDITIONAL_COUNTRY_SEARCH_TEMPLATES)
OTHER_COUNTRY_SEARCH_TEMPLATES = OTHER_COUNTRY_SEARCH_TEMPLATES + (
    ("en", "United States EIA ethanol production stocks {readable}"),
    ("en", "Philippines sugarcane pest cane farmers aid {readable}"),
    ("en", "Indonesia raw sugar import refinery regulation {readable}"),
    ("en", "Fiji Sugar Corporation crushing cane price {readable}"),
    ("en", "Nepal sugarcane farmers cane dues {readable}"),
    ("en", "European Commission sugar beet production export {readable}"),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Sugar News Excel and dashboard data.")
    parser.add_argument("--date", help="Target news date in YYYY-MM-DD. Defaults to Beijing yesterday.")
    parser.add_argument("--task-root", help="Sugar News task root. Defaults to the current project root.")
    parser.add_argument("--skip-if-success", action="store_true", help="Skip if public status already marks target date successful.")
    parser.add_argument("--offline-only", action="store_true", help="Do not attempt fallback online discovery; require verified JSON.")
    parser.add_argument("--allow-rss-autogen", action="store_true", help="Generate a conservative verified JSON from RSS if no curated verified JSON exists.")
    parser.add_argument("--force-rss-autogen", action="store_true", help="Regenerate verified news from the upgraded RSS discovery path even when a verified JSON already exists.")
    parser.add_argument(
        "--skip-metric-refresh",
        action="store_true",
        help="Do not run metric scrapers in-process; consume the latest metric snapshots already refreshed by the workflow.",
    )
    parser.add_argument(
        "--preserve-existing-metrics",
        action="store_true",
        help="Explicit news-only repair: retain the target report's existing Brazil and India dashboard blocks.",
    )
    return parser.parse_args()


def beijing_now() -> datetime:
    fixed = os.getenv("SUGAR_NEWS_NOW")
    if fixed:
        return datetime.fromisoformat(fixed).astimezone(SHANGHAI)
    return datetime.now(SHANGHAI)


def target_date(value: str | None) -> str:
    if value:
        return datetime.strptime(value, "%Y-%m-%d").date().isoformat()
    return (beijing_now().date() - timedelta(days=1)).isoformat()


def task_root_from_args(value: str | None) -> Path:
    root = Path(value or os.getenv("SUGAR_NEWS_ROOT", str(DEFAULT_TASK_ROOT))).resolve()
    return root


def date_parts(date_text: str) -> tuple[str, str]:
    yyyy, mm, _ = date_text.split("-")
    return yyyy, mm


def ensure_task_dirs(task_root: Path, date_text: str) -> None:
    yyyy, mm = date_parts(date_text)
    for rel in (
        Path("data") / "verified_news" / yyyy / mm,
        Path("logs") / yyyy / mm,
        Path("reports") / yyyy / mm,
    ):
        (task_root / rel).mkdir(parents=True, exist_ok=True)


def verified_json_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "data" / "verified_news" / yyyy / mm / f"sugar_news_{date_text}.json"


def search_log_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "logs" / yyyy / mm / f"search_log_{date_text}.json"


def write_log_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "logs" / yyyy / mm / f"write_log_{date_text}.json"


def excel_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "reports" / yyyy / mm / f"Sugar News {date_text}.xlsx"


def public_report_path(date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return PUBLIC_DATA_ROOT / "reports" / yyyy / mm / f"{date_text}.json"


def public_index_path() -> Path:
    return PUBLIC_DATA_ROOT / "index.json"


def public_status_path() -> Path:
    return PUBLIC_DATA_ROOT / "status.json"


def success_exists(date_text: str) -> bool:
    status_path = public_status_path()
    index_path = public_index_path()
    report_path = public_report_path(date_text)
    if not status_path.exists() or not index_path.exists() or not report_path.exists():
        return False
    with status_path.open("r", encoding="utf-8") as f:
        status = json.load(f)
    if status.get("latestNewsDate") != date_text or status.get("lastRunStatus") != "success":
        return False
    with index_path.open("r", encoding="utf-8") as f:
        index = json.load(f)
    if index.get("latestNewsDate") != date_text:
        return False
    expected_report = "/" + str(report_path.relative_to(PROJECT_ROOT)).replace("\\", "/")
    return any(report.get("newsDate") == date_text and report.get("path") == expected_report for report in index.get("reports", []))


def google_news_rss_url(query: str) -> str:
    return "https://news.google.com/rss/search?q=" + quote_plus(query) + "&hl=en-US&gl=US&ceid=US:en"


def fetch_rss(query: str, timeout: int = 15) -> list[dict]:
    req = Request(google_news_rss_url(query), headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=timeout) as resp:
        body = resp.read()
    root = ET.fromstring(body)
    items = []
    for node in root.findall("./channel/item"):
        title = node.findtext("title") or ""
        link = node.findtext("link") or ""
        published = node.findtext("pubDate") or ""
        desc = node.findtext("description") or ""
        items.append({"title": title, "link": link, "published": published, "description": desc})
    return items


def google_news_article_id(link: str) -> str | None:
    match = re.search(r"https?://news\.google\.com/rss/articles/([^?/#]+)", link or "")
    return match.group(1) if match else None


def decode_google_news_source_url(link: str, timeout: int | None = None) -> tuple[str, dict]:
    """Resolve Google News RSS article links to the original publisher URL."""
    article_id = google_news_article_id(link)
    if not article_id:
        return link, {"status": "not_google_news"}
    timeout = timeout or max(8, RSS_AUTOGEN_TIMEOUT_SECONDS * 2)
    article_url = f"https://news.google.com/rss/articles/{article_id}?oc=5"
    try:
        req = Request(article_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=timeout) as resp:
            html = resp.read().decode("utf-8", "ignore")
        ts_match = re.search(r'data-n-a-ts="([^"]+)"', html)
        sig_match = re.search(r'data-n-a-sg="([^"]+)"', html)
        if not ts_match or not sig_match:
            return link, {"status": "failed", "reason": "missing_google_news_signature"}

        request_payload = [
            "garturlreq",
            [
                [
                    "en-US",
                    "US",
                    ["FINANCE_TOP_INDICES", "WEB_TEST_1_0_0"],
                    None,
                    None,
                    1,
                    1,
                    "US:en",
                    None,
                    180,
                    None,
                    None,
                    None,
                    None,
                    None,
                    0,
                    None,
                    None,
                    [1608992183, 723341000],
                ],
                "en-US",
                "US",
                1,
                [2, 3, 4, 8],
                1,
                0,
                "655000234",
                0,
                0,
                None,
                0,
            ],
            article_id,
            int(ts_match.group(1)),
            sig_match.group(1),
        ]
        batched = [[["Fbv4je", json.dumps(request_payload, separators=(",", ":")), None, "generic"]]]
        body = urlencode({"f.req": json.dumps(batched, separators=(",", ":"))}).encode()
        req = Request(
            "https://news.google.com/_/DotsSplashUi/data/batchexecute?rpcids=Fbv4je",
            data=body,
            headers={
                "Content-Type": "application/x-www-form-urlencoded;charset=utf-8",
                "Referer": "https://news.google.com/",
                "User-Agent": "Mozilla/5.0",
            },
        )
        with urlopen(req, timeout=timeout) as resp:
            text = resp.read().decode("utf-8", "ignore")
        match = re.search(r'\[\\"garturlres\\",\\"(.*?)\\",', text)
        if not match:
            return link, {"status": "failed", "reason": "decode_response_without_source_url"}
        resolved = codecs.decode(match.group(1), "unicode_escape")
        return resolved, {"status": "resolved", "googleNewsUrl": link}
    except Exception as exc:
        return link, {"status": "failed", "reason": str(exc)[:300]}


def fallback_discovery(date_text: str, task_root: Path) -> None:
    """Record auditable search attempts.

    The cloud job needs a durable trail even when a fully verified newsroom-style
    dataset cannot be produced automatically. This fallback intentionally does
    not publish unverified RSS items as facts.
    """
    dt = datetime.strptime(date_text, "%Y-%m-%d")
    buddhist_year = dt.year + 543
    readable = dt.strftime("%B %-d %Y") if os.name != "nt" else dt.strftime("%B %#d %Y")
    context = {
        "readable": readable,
        "day": dt.day,
        "month_name": dt.strftime("%B"),
        "month_name_pt": PORTUGUESE_MONTH_NAMES[dt.month],
        "month": dt.month,
        "year": dt.year,
        "date_slash": dt.strftime("%d/%m/%Y"),
        "buddhist_year": buddhist_year,
    }
    searches = []
    searches.extend(("全球", "en", template.format(**context)) for template in GLOBAL_SEARCH_TEMPLATES)
    for country, templates in COUNTRY_SEARCH_TEMPLATES.items():
        searches.extend((country, language, template.format(**context)) for language, template in templates)
    searches.extend(("印度指标", language, template.format(**context)) for language, template in INDIA_PRICE_INVENTORY_SEARCH_TEMPLATES)
    log = {
        "target_date": date_text,
        "run_date": beijing_now().date().isoformat(),
        "search_tool": "Google News RSS fallback via urllib",
        "note": "RSS search results are logged for audit. Items are not published unless a verified JSON is created.",
        "date_format_examples": DATE_FORMAT_EXAMPLES,
        "india_price_inventory_sources": INDIA_PRICE_INVENTORY_SOURCE_GUIDE,
        "other_country_rule": "Other-country news is unlimited; each concrete country keeps an independent object/list and must never be collapsed into a single 其他 key.",
        "searches": [],
        "pipeline_counts": {
            "global_initial_candidates": 0,
            "country_supplement_candidates": 0,
            "candidate_news_after_search": 0,
            "date_verified_or_continuing_impact": 0,
            "relevance_passed": 0,
            "importance_passed": 0,
            "deduped": 0,
            "structured_data_count": 0,
            "passed_to_excel": 0,
        },
    }
    total = 0
    global_total = 0
    country_total = 0
    for country, language, query in searches:
        entry = {
            "country": country,
            "language": language,
            "keywords": query,
            "request_status": "pending",
            "returned_count": 0,
            "retained_count": 0,
            "filtered": [],
        }
        try:
            items = fetch_rss(query)
            entry["request_status"] = "executed"
            entry["returned_count"] = len(items)
            total += len(items)
            if country == "全球":
                global_total += len(items)
            else:
                country_total += len(items)
            entry["sample_results"] = items[:5]
            for result in items[:5]:
                entry["filtered"].append({
                    "country": country,
                    "title": result.get("title"),
                    "news_date": result.get("published"),
                    "source": "Google News RSS",
                    "url": result.get("link"),
                    "stage": "source_page_verification",
                    "reason": "RSS result requires source-page date/body verification before publication.",
                })
        except Exception as exc:
            entry["request_status"] = "failed"
            entry["error"] = str(exc)[:500]
        log["searches"].append(entry)
    log["pipeline_counts"]["global_initial_candidates"] = global_total
    log["pipeline_counts"]["country_supplement_candidates"] = country_total
    log["pipeline_counts"]["candidate_news_after_search"] = total
    path = search_log_path(task_root, date_text)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)


def rss_source_from_title(title: str) -> tuple[str, str]:
    parts = title.rsplit(" - ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return title.strip(), "Google News RSS"


def rss_item_date(item: dict) -> str | None:
    try:
        value = parsedate_to_datetime(item.get("published", ""))
    except Exception:
        return None
    return value.astimezone(SHANGHAI).date().isoformat()


def rss_item_datetime(item: dict) -> datetime | None:
    try:
        return parsedate_to_datetime(item.get("published", "")).astimezone(SHANGHAI)
    except Exception:
        return None


def publication_window_for_target(date_text: str) -> tuple[datetime, datetime, str]:
    target = datetime.strptime(date_text, "%Y-%m-%d").date()
    end = datetime.combine(target + timedelta(days=1), datetime.min.time(), tzinfo=SHANGHAI) + timedelta(hours=6)
    if end.weekday() == 0:
        start = datetime.combine(target - timedelta(days=2), datetime.min.time(), tzinfo=SHANGHAI) + timedelta(hours=16)
        rule = "monday_weekend_window_friday_16_to_monday_06"
    else:
        start = end - timedelta(hours=RSS_AUTOGEN_PUBLICATION_WINDOW_HOURS)
        rule = f"rolling_{RSS_AUTOGEN_PUBLICATION_WINDOW_HOURS}_hour_window"
    return start, end, rule


def rss_item_in_publication_window(item: dict, date_text: str) -> tuple[bool, str, str | None]:
    published_at = rss_item_datetime(item)
    if not published_at:
        return False, "publication time cannot be parsed", None
    start, end, rule = publication_window_for_target(date_text)
    if start <= published_at <= end:
        return True, rule, published_at.date().isoformat()
    return False, f"publication time outside {rule}: {start.isoformat()} to {end.isoformat()}", published_at.date().isoformat()


def classify_sugar_topic(text: str) -> str:
    lowered = text.lower()
    if re.search(r"(?i)\bsugar\s+prices?\b|\bprices?\s+(?:jump|surge|rise|rose|fall|fell)\b|ex-mill|wholesale|retail", lowered):
        return "price_market"
    for topic, terms in NEWS_TOPIC_RULES:
        if any_phrase(lowered, terms):
            return topic
    return "general_industry"


def extract_metrics(text: str) -> list[str]:
    patterns = (
        r"\d+(?:[.,]\d+)?\s*(?:%|percent|percentage points|bps)",
        r"\d+(?:[.,]\d+)?\s*(?:lakh|crore|million|billion|tonnes?|tons?|MT|LMT|quintals?|litres?|liters?|barrels?|bpd|KL|kl|hectares?|acres?|provinces?|states?|mills?|factories?|days?)",
        r"(?:₹|Rs\.?|INR|R\$|\$)\s*\d+(?:[.,]\d+)?",
        r"E(?:10|20|27|30|32)",
        r"\d+(?:[.,]\d+)?\s*(?:万吨|吨|千吨|公担|卢比|美元|美分|亿升|万公顷|公顷|家|个省|座|天)",
    )
    found: list[str] = []
    for pattern in patterns:
        for match in re.findall(pattern, text, flags=re.IGNORECASE):
            value = match if isinstance(match, str) else "".join(match)
            value = value.strip()
            if re.fullmatch(r"(?i)rs\s*[0-9]", value):
                continue
            if re.fullmatch(r"(?i)\dMT", value):
                continue
            if value.lower() in {"rs2", "rs 2"}:
                continue
            if value and value not in found:
                found.append(value)
    return found[:8]


def localize_metric_for_summary(value: str) -> str:
    replacements = (
        (r"(?i)(?<![\d.])(\d+(?:[.,]\d+)?)\s*provinces?\b", r"\1个省"),
        (r"(?i)(?<![\d.])(\d+(?:[.,]\d+)?)\s*states?\b", r"\1个州"),
        (r"(?i)(?<![\d.])(\d+(?:[.,]\d+)?)\s*mills?\b", r"\1家糖厂"),
        (r"(?i)(?<![\d.])(\d+(?:[.,]\d+)?)\s*factories?\b", r"\1家糖厂"),
        (r"(?i)(?<![\d.])(\d+(?:[.,]\d+)?)\s*days?\b", r"\1天"),
        (r"(?i)(?<![\d.])(\d+(?:[.,]\d+)?)\s*hectares?\b", r"\1公顷"),
        (r"(?i)(?<![\d.])(\d+(?:[.,]\d+)?)\s*acres?\b", r"\1英亩"),
    )
    localized = value
    for pattern, replacement in replacements:
        localized = re.sub(pattern, replacement, localized)
    return localized


def polish_verified_summary_text(item: dict) -> dict:
    row = dict(item)
    for text_field in ("news", "impact"):
        if row.get(text_field):
            row[text_field] = localize_metric_for_summary(str(row[text_field]))

    title_text = str(row.get("title") or row.get("source_title") or "")
    metadata = row.get("metadata")
    if isinstance(metadata, dict):
        title_text = f"{title_text} {metadata.get('source_title') or ''}"
    if re.search(r"(?i)rs\.?\s*4263\s*/\s*mt", title_text):
        row["news"] = re.sub(r"以(Rs\s*4263)(?!/MT)销售PDM", r"以\1/MT销售PDM", row.get("news", ""))
    return row


def infer_event_actor(country: str, topic: str, title: str, source: str) -> str:
    lowered = title.lower()
    if country == "印度" and topic == "price_market" and "mills expect" in lowered:
        return "印度糖厂"
    if country == "菲律宾" and topic == "weather_pest":
        if "negros oriental" in lowered and "national aid" in lowered:
            return "菲律宾东内格罗斯省政府"
        if "negocc task force" in lowered or ("task force" in lowered and "spray" in lowered):
            return "菲律宾西内格罗斯省虫害防控工作组"
    if "pib" in source.lower() or "ministry" in lowered or "government" in lowered or "govt" in lowered or "centre" in lowered:
        if country == "印度":
            return "印度政府"
        if country == "巴西":
            return "巴西政府"
        if country == "中国":
            return "中国政府部门"
        return f"{country}政府"
    if "eia" in source.lower() or "eia" in lowered:
        return "美国能源信息署EIA"
    if "usda" in source.lower() or "usda" in lowered:
        return "美国农业部USDA"
    if "unica" in source.lower() or "unica" in lowered:
        return "UNICA"
    if "datagro" in source.lower() or "datagro" in lowered:
        return "Datagro"
    if "vasantdada" in lowered:
        return "Vasantdada糖业研究所"
    if "sugar regulatory administration" in lowered or source.upper() == "SRA":
        return "菲律宾糖业监管署"
    if "planter" in lowered or "farmers" in lowered or "cane farmers" in lowered:
        return f"{country}蔗农组织"
    if "mill" in lowered or "factory" in lowered:
        return f"{country}糖厂"
    if topic == "weather_pest":
        return f"{country}主产区农业或气象机构"
    if topic == "price_market":
        return f"{country}糖业市场"
    return f"{country}糖业相关机构"


def infer_event_action(topic: str, title: str) -> str:
    lowered = title.lower()
    if "seeks national aid" in lowered:
        return "寻求国家援助"
    if "spray" in lowered and "fungi" in lowered:
        return "使用真菌处理"
    if any(term in lowered for term in ("raise", "increase", "hike", "提高", "上调")):
        return "提高"
    if any(term in lowered for term in ("cut", "lower", "reduce", "下调", "降低", "减少")):
        return "下调"
    if any(term in lowered for term in ("approve", "approved", "order", "impose", "limit", "quota", "tariff", "ban", "policy", "regulation")):
        return "发布或调整"
    if any(term in lowered for term in ("forecast", "predict", "estimate", "seen", "预计", "预测")):
        return "预测"
    if any(term in lowered for term in ("report", "reports", "公布", "发布")):
        return "公布"
    if topic == "weather_pest":
        return "预警"
    if topic == "mill_operations":
        return "披露"
    return "公布"


def structured_candidate_from_rss(country_bucket: str, rss: dict, date_text: str, title_clean: str, source: str) -> dict:
    link = rss.get("link", "").strip()
    haystack = f"{title_clean} {rss.get('description', '')}"
    concrete_country, country_group = infer_core_country(haystack, country_bucket)
    if country_bucket not in {"其他国家", "印度指标"} and country_group != country_bucket:
        assigned_country = concrete_country
        assigned_group = country_group
    elif country_bucket == "其他国家":
        assigned_country = concrete_country
        assigned_group = country_group
    else:
        assigned_country = concrete_country
        assigned_group = country_group
    if assigned_group not in GROUP_ORDER:
        assigned_group = "其他国家"
    topic = classify_sugar_topic(haystack)
    metrics = extract_metrics(haystack)
    actor = infer_event_actor(assigned_country, topic, title_clean, source)
    action = infer_event_action(topic, title_clean)
    in_window, window_reason, item_date = rss_item_in_publication_window(rss, date_text)
    return {
        "source_title": title_clean,
        "source_url": link,
        "publisher": source,
        "publication_time": rss.get("published"),
        "event_date": item_date or date_text,
        "event_country": assigned_country,
        "event_region": None,
        "event_actor": actor,
        "event_action": action,
        "metrics": metrics,
        "comparison_period": None,
        "topic": topic,
        "sugar_relevance": "pending",
        "impact_direction": None,
        "impact_logic": None,
        "verification_status": "待核实",
        "publication_window_status": "in_window" if in_window else "out_of_window",
        "publication_window_reason": window_reason,
    }


def event_fingerprint(candidate: dict) -> str:
    title = str(candidate.get("source_title") or "").lower()
    topic = str(candidate.get("topic") or "")
    country = str(candidate.get("event_country") or "")
    metrics = " ".join(candidate.get("metrics") or [])
    if (
        country == "印度"
        and topic == "price_market"
        and "17%" in metrics
        and "sugar price" in title
        and any_phrase(title, ("stock limit", "stock limits", "curbs", "curb"))
    ):
        return "印度|糖价上涨17|库存限制或后续调控"
    parts = [
        country,
        str(candidate.get("event_actor") or ""),
        str(candidate.get("event_action") or ""),
        metrics,
        str(candidate.get("event_date") or ""),
        topic,
    ]
    return re.sub(r"\W+", "", "|".join(parts).lower())[:160]


def is_india_indirect_sugar_relevant(text: str) -> bool:
    """Detect India ethanol-policy stories that affect sugar without saying sugar.

    E20, OMC procurement, and ethanol-feedstock policies can change the split
    between cane-derived ethanol and food sugar. These are high-relevance India
    sugar stories even when the headline is framed as fuel or energy policy.
    """
    lowered = text.lower()
    has_policy = any(term in lowered for term in INDIA_INDIRECT_ETHANOL_POLICY_TERMS)
    has_feedstock = any(term in lowered for term in INDIA_INDIRECT_FEEDSTOCK_TERMS)
    has_above_e20_context = "above 20" in lowered or "beyond e20" in lowered or "over 20" in lowered
    return has_policy and (has_feedstock or has_above_e20_context)


def has_phrase(text: str, phrase: str) -> bool:
    escaped = re.escape(phrase.lower())
    if re.search(r"[a-z0-9]", phrase.lower()):
        return re.search(rf"(?<![a-z0-9]){escaped}(?![a-z0-9])", text.lower()) is not None
    return phrase.lower() in text.lower()


def any_phrase(text: str, phrases: tuple[str, ...]) -> bool:
    return any(has_phrase(text, phrase) for phrase in phrases)


def matched_core_countries(text: str) -> list[str]:
    padded = f" {text.lower()} "
    matches = []
    for country, aliases in COUNTRY_ALIASES.items():
        if any_phrase(padded, aliases):
            matches.append(country)
    return matches


def trusted_source_for_country(country: str, source: str, text: str = "") -> bool:
    markers = COUNTRY_TRUSTED_SOURCE_MARKERS.get(country, ())
    if not markers:
        return False
    return any_phrase(f"{source} {text}".lower(), markers)


def candidate_country_confidence(candidate: dict, search_bucket: str, haystack: str, source: str) -> str | None:
    country = candidate.get("event_country") or search_bucket
    matches = matched_core_countries(f"{haystack} {source}")
    if country in matches:
        return "explicit_country_or_region_in_title_or_source"
    if trusted_source_for_country(country, source, haystack):
        return "trusted_country_source"
    if search_bucket == "印度指标":
        return "metric_search_bucket"
    return None


def is_title_only_low_quality_context(text: str) -> bool:
    lowered = text.lower()
    if is_medical_sugar_context(lowered) or is_non_industry_sugar_context(lowered):
        return True
    if any_phrase(lowered, RSS_TITLE_ONLY_REJECT_TERMS):
        return True
    if "sugar rationing" in lowered and any_phrase(lowered, ("dementia", "early life", "health", "risk")):
        return True
    return False


def candidate_has_verifiable_industry_fact(candidate: dict) -> tuple[bool, str]:
    title = str(candidate.get("source_title", ""))
    lowered = title.lower()
    topic = candidate.get("topic") or "general_industry"
    metrics = candidate.get("metrics") or []
    if is_title_only_low_quality_context(title):
        return False, "title-only candidate is finance, health, entertainment, or otherwise not a sugar-industry event"
    if topic == "general_industry":
        return False, "RSS title lacks a specific sugar-industry topic; kept in candidate log only"
    if topic == "price_market" and not metrics:
        return False, "price-market RSS title lacks price level, change amount, or comparison metric"
    if topic in {"supply_demand", "cane_farming", "variety_research", "ethanol_capacity"} and not metrics:
        return False, "candidate topic requires a concrete numeric metric before publication"
    if topic == "weather_pest" and not (
        metrics
        or any_phrase(lowered, ("rain", "rainfall", "monsoon", "drought", "flood", "pest", "disease", "red rot", "white grub", "white leaf", "降雨", "干旱", "病虫害"))
    ):
        return False, "weather or pest candidate lacks affected area, hazard, or magnitude"
    if topic == "ethanol_policy":
        if not metrics:
            return False, "ethanol-policy candidate lacks blend level, price, saving, or other concrete metric"
        if any_phrase(lowered, LOW_SIGNAL_ETHANOL_DISCUSSION_TERMS) and not any_phrase(
            lowered,
            ("sugarcane", "cane", "molasses", "sugar syrup", "government", "govt", "ministry", "pib", "procurement", "foreign exchange", "crude import", "加油站", "糖蜜", "甘蔗"),
        ):
            return False, "ethanol candidate is a consumer or political fuel-price discussion without sugar-feedstock linkage"
    if topic == "trade_policy" and not (metrics or any_phrase(lowered, ("ban", "restriction", "curb", "import", "export", "tariff", "regulation", "allocation", "进口", "出口", "禁令"))):
        return False, "trade-policy candidate lacks policy direction or quantity"
    return True, "candidate has publishable title-level sugar-industry facts"


def is_medical_sugar_context(text: str) -> bool:
    return any_phrase(text, MEDICAL_SUGAR_TERMS)


def is_non_industry_sugar_context(text: str) -> bool:
    return any_phrase(text, NON_INDUSTRY_SUGAR_TERMS)


SOURCE_SUFFIX_RE = re.compile(
    r"\s*来源：[^（]+（https?://[^）]+）(?:。?影响：(?:利多糖价|利空糖价|中性))?\s*$"
)
PUBLICATION_LEAD_RE = re.compile(
    r"^\s*(?:\d{4}-\d{2}-\d{2}\s+)?[^。！？]{0,40}(?:报道|消息|发布|称)[:：]"
)
ORDINARY_PUBLICATION_RE = re.compile(
    r"(?:今日|今天|本日)(?:发布|消息|报道)|\d{1,2}月\d{1,2}日(?:消息|报道)"
)


def news_body_without_source(news: str) -> str:
    return SOURCE_SUFFIX_RE.sub("", news or "").strip()


def impact_marker_from_impact(impact: str) -> str:
    if impact.startswith(("偏多糖价：", "利多：")):
        return "影响：利多糖价"
    if impact.startswith(("偏空糖价：", "利空：")):
        return "影响：利空糖价"
    return "影响：中性"


def ensure_news_impact_marker(news: str, impact: str) -> str:
    label = impact_marker_from_impact(impact)
    stripped = NEWS_IMPACT_MARKER_RE.sub("", (news or "").strip()).rstrip("。")
    return f"{stripped}。{label}"


def validate_news_impact_marker(item: dict, idx: int) -> None:
    expected = impact_marker_from_impact(str(item.get("impact", "")))
    if not str(item.get("news", "")).strip().endswith(expected):
        raise ValueError(f"Verified item {idx} must end B-column news with {expected}")


def split_cn_sentences(text: str) -> list[str]:
    pieces = re.split(r"[。！？]+", text)
    return [piece.strip() for piece in pieces if piece.strip()]


def has_chinese_text(text: str) -> bool:
    return re.search(r"[\u4e00-\u9fff]", text or "") is not None


def core_item_text(item: dict) -> str:
    return " ".join(
        str(value)
        for value in (
            item.get("title", ""),
            news_body_without_source(str(item.get("news", ""))),
            item.get("impact", ""),
        )
    )


def is_china_monitoring_note(item: dict) -> bool:
    return (
        item.get("date_status") == "monitoring_completed"
        or item.get("title") == "中国糖业每日监测"
        or str(item.get("dedupe_key", "")).startswith("china_daily_monitoring_")
    )


def validate_no_vague_summary(item: dict, idx: int, body: str, impact: str) -> None:
    quality_text = f"{body} {impact}"
    for phrase in VAGUE_SUMMARY_PHRASES:
        if phrase in quality_text:
            raise ValueError(f"Verified item {idx} contains vague summary phrase: {phrase}")
    for pattern in VAGUE_SUMMARY_PATTERNS:
        if pattern.search(quality_text):
            raise ValueError(f"Verified item {idx} contains source-led vague fallback wording")


def validate_media_is_not_event_subject(item: dict, idx: int, body: str) -> None:
    first_sentence = split_cn_sentences(body)[0] if split_cn_sentences(body) else body
    source_name = str(item.get("source_name", "")).strip()
    if not source_name:
        return
    lead = first_sentence.strip()
    source_lead_markers = ("消息涉及", "消息称", "报道称", "报道，", "报道指出", "发布消息")
    if lead.lower().startswith(source_name.lower()):
        suffix = lead[len(source_name):].lstrip()
        if suffix.startswith(source_lead_markers):
            raise ValueError(f"Verified item {idx} uses media source as the event subject")


def validate_concrete_event_and_impact(item: dict, idx: int, body: str, impact: str) -> None:
    if is_china_monitoring_note(item):
        return
    if not any_phrase(body, NEWS_ACTION_TERMS):
        raise ValueError(f"Verified item {idx} lacks a clear event action")
    if not any_phrase(body, NEWS_DIRECTION_TERMS):
        raise ValueError(f"Verified item {idx} lacks a clear change direction")
    if not (CONCRETE_DETAIL_RE.search(body) or any_phrase(body, NEWS_DETAIL_TERMS)):
        raise ValueError(f"Verified item {idx} lacks concrete data, policy, market, production, trade, or weather detail")
    if not any_phrase(f"{body} {impact}", IMPACT_TARGET_TERMS):
        raise ValueError(f"Verified item {idx} lacks a sugar supply-demand impact target")
    if not any_phrase(impact, IMPACT_CAUSAL_TERMS):
        raise ValueError(f"Verified item {idx} lacks supply-demand-to-price transmission logic")


def normalize_country_fields(item: dict) -> dict:
    row = dict(item)
    concrete_country, country_group = infer_core_country(core_item_text(row), row.get("country") or "")
    if concrete_country and concrete_country != row.get("country"):
        row["country"] = concrete_country
        row["country_group"] = country_group
    if row.get("country_group") not in GROUP_ORDER:
        row["country_group"] = "其他国家"
    if row.get("country_group") == "其他国家" and row.get("country") in {"其他", "其他国家"}:
        concrete_country, country_group = infer_core_country(core_item_text(row), "其他国家")
        if country_group == "其他国家" and concrete_country not in {"其他", "其他国家"}:
            row["country"] = concrete_country
        else:
            raise ValueError("Other-country rows must use a concrete country/region name")
    if row.get("country") in GROUP_ORDER and row.get("country") != "其他国家":
        row["country_group"] = row["country"]
    return row


def validate_editorial_quality(item: dict, idx: int) -> None:
    body = news_body_without_source(str(item.get("news", "")))
    impact = str(item.get("impact", ""))
    quality_text = f"{body} {impact}"
    if is_medical_sugar_context(quality_text):
        raise ValueError(f"Verified item {idx} is medical/health sugar content")
    if is_non_industry_sugar_context(quality_text):
        raise ValueError(f"Verified item {idx} is non-industry sugar content")
    if not has_chinese_text(body):
        raise ValueError(f"Verified item {idx} summary must be written in Chinese")
    sentences = split_cn_sentences(body)
    if not 2 <= len(sentences) <= 3:
        raise ValueError(f"Verified item {idx} summary must be 2-3 Chinese sentences, got {len(sentences)}")
    if PUBLICATION_LEAD_RE.search(body) and not body.startswith(("周末延续消息：", "近期重要消息：")):
        raise ValueError(f"Verified item {idx} starts with source/publication-date reporting formula")
    if ORDINARY_PUBLICATION_RE.search(body):
        raise ValueError(f"Verified item {idx} repeats ordinary publication date wording")
    validate_no_vague_summary(item, idx, body, impact)
    validate_media_is_not_event_subject(item, idx, body)
    validate_concrete_event_and_impact(item, idx, body, impact)
    inferred_country, inferred_group = infer_core_country(core_item_text(item), item.get("country") or "")
    if inferred_country in GROUP_ORDER and inferred_country != "其他国家" and item.get("country_group") != inferred_group:
        raise ValueError(f"Verified item {idx} country_group={item.get('country_group')} conflicts with core country {inferred_country}")
    if inferred_group == "其他国家" and inferred_country not in {"其他", "其他国家"} and item.get("country") in {"其他", "其他国家"}:
        raise ValueError(f"Verified item {idx} must label other-country item as {inferred_country}")


def validate_global_summary(summary: str) -> str:
    text = re.sub(r"\s+", " ", (summary or "").strip())
    if not text:
        raise ValueError("Global sugar highlights summary is required")
    if not has_chinese_text(text):
        raise ValueError("Global sugar highlights must be written in Chinese")
    sentences = split_cn_sentences(text)
    if not 2 <= len(sentences) <= 3:
        raise ValueError(f"Global sugar highlights must be 2-3 Chinese sentences, got {len(sentences)}")
    validate_no_vague_summary({"title": "全球糖业新闻重点"}, 0, text, "")
    if "震荡偏强" in text and not re.search(r"15\s*美分/磅", text):
        raise ValueError("Global summary may use 震荡偏强 only with the ICE raw sugar 15 cents/lb condition")
    if not any(country in text for country in ("巴西", "印度", "泰国", "中国", "菲律宾", "全球", "国际")):
        raise ValueError("Global summary must name at least one country or global market subject")
    if not any(term in text for term in ("利多", "支撑", "减少", "收紧", "利空", "压制", "增加", "供应", "需求", "库存", "天气", "乙醇")):
        raise ValueError("Global summary must describe concrete bullish/bearish supply-demand factors")
    return text


def item_primary_fact(item: dict) -> str:
    body = news_body_without_source(str(item.get("news", "")))
    sentences = split_cn_sentences(body)
    return sentences[0] if sentences else body


def item_impact_fact(item: dict) -> str:
    body = news_body_without_source(str(item.get("news", "")))
    sentences = split_cn_sentences(body)
    if len(sentences) >= 2:
        return sentences[1]
    return item_primary_fact(item)


def compact_fact(text: str, limit: int = 88) -> str:
    text = re.sub(r"\s+", "", text.strip("。；; "))
    if len(text) <= limit:
        return text
    return text[:limit].rstrip("，、；;") + "…"


def build_global_summary_from_items(items: list[dict]) -> str:
    if not items:
        raise ValueError("Cannot build global summary without verified news items")
    event_bits = [compact_fact(item_primary_fact(item), 78) for item in items[:3]]
    first = "全球糖业新闻重点集中在" + "；".join(event_bits) + "。"

    bullish = [compact_fact(item_impact_fact(item), 70) for item in items if item["impact"].startswith(("偏多糖价：", "利多："))]
    bearish = [compact_fact(item_impact_fact(item), 70) for item in items if item["impact"].startswith(("偏空糖价：", "利空："))]
    neutral = [compact_fact(item_impact_fact(item), 70) for item in items if item["impact"].startswith(("中性：", "影响有限："))]
    factor_parts = []
    if bullish:
        factor_parts.append("利多因素是" + bullish[0])
    if bearish:
        factor_parts.append("利空因素是" + bearish[0])
    if not factor_parts and neutral:
        factor_parts.append("中性因素是" + neutral[0])
    second = "；".join(factor_parts) + "。"

    if bullish and bearish:
        third = "国际糖价的主要矛盾在供应收缩预期与阶段性供应或需求压力之间，ICE原糖持续站上15美分/磅前以震荡判断为宜。"
    elif bullish:
        third = "国际糖价主要受供应收缩或乙醇分流预期支撑，但ICE原糖持续站上15美分/磅前不使用震荡偏强表述。"
    elif bearish:
        third = "国际糖价主要受供应改善、天气恢复或流通增加压力牵制，短期以震荡判断为宜。"
    else:
        third = "国际糖价缺少新的单边供需冲击，短期以震荡判断为宜。"
    return validate_global_summary(first + second + third)


def global_summary_for_report(items: list[dict], verified_data: dict | None = None) -> str:
    verified_data = verified_data or {}
    provided = verified_data.get("global_summary") or verified_data.get("globalHighlights")
    if provided:
        return validate_global_summary(str(provided))
    return build_global_summary_from_items(items)


def infer_core_country(text: str, fallback_country: str) -> tuple[str, str]:
    matches = matched_core_countries(text)
    if not matches:
        return fallback_country, fallback_country if fallback_country in GROUP_ORDER else "其他国家"
    priority = ["巴西", "印度", "泰国", "中国"]
    for country in priority:
        if country in matches:
            return country, country
    country = matches[0]
    return country, "其他国家"


def rss_sugar_relevant(country: str, text: str) -> bool:
    if is_medical_sugar_context(text):
        return False
    if is_non_industry_sugar_context(text):
        return False
    domain_terms = (
        "sugar", "sugarcane", "cane", "molasses", "raw sugar", "white sugar",
        "hedge", "hedged", "hedging", "fixed-price", "fixação", "fixacao",
        "biofuel", "syrup", "distillery",
        "frp", "sap", "aista", "isma", "nfcsf", "ex-mill", "sales quota",
        "sucroenergético", "açúcar", "cana", "etanol", "น้ำตาล", "อ้อย",
        "เอทานอล", "食糖", "白糖", "甘蔗", "甜菜糖", "郑糖",
    )
    ethanol_terms = ("ethanol", "e20", "e10", "blend", "blending", "bioethanol")
    sugar_feedstock_terms = ("sugarcane", "cane", "molasses", "sugar syrup", "sugar self-sufficiency", "distillery", "गन्ना", "इथेनॉल", "甘蔗", "糖蜜")
    weather_terms = ("rain", "rainfall", "monsoon", "drought", "flood", "weather", "降雨", "季风", "干旱", "洪涝")
    cane_regions = (
        "uttar pradesh", "maharashtra", "karnataka", "tamil nadu", "gujarat",
        "bihar", "punjab", "haryana", "uttarakhand", "khon kaen",
        "nakhon ratchasima", "chaiyaphum", "udon thani", "sao paulo",
        "centro-sul", "guangxi", "yunnan", "广西", "云南", "北方邦",
        "马哈拉施特拉", "卡纳塔克",
    )
    if any_phrase(text, domain_terms):
        return True
    if country == "印度" and is_india_indirect_sugar_relevant(text):
        return True
    if any_phrase(text, (
        "anhydrous ethanol", "ethanol blend", "ethanol blending", "ethanol programme",
        "ethanol program", "foreign exchange", "crude import", "petrol pump",
        "ethanol capacity", "ethanol procurement price", "乙醇掺混", "乙醇产能",
        "节省外汇", "替代原油", "加油站", "乙醇采购价",
    )):
        return True
    if country == "巴西" and any_phrase(text, (
        "corn ethanol", "maize ethanol", "grain ethanol", "etanol de milho",
        "ethanol gasoline", "gasoline ethanol", "etanol gasolina", "etanol e gasolina",
        "sugar hedging", "mills hedged", "hedge ratio", "fixed-price sales",
        "fixação de açúcar", "contratos de exportação", "usinas hedge",
        "糖厂套保", "食糖套保比例", "糖套保进度",
    )):
        return True
    if country in {"美国", "其他国家"} and any_phrase(text, ("eia", "ethanol production", "ethanol stocks")):
        return True
    if any_phrase(text, ethanol_terms) and any_phrase(text, sugar_feedstock_terms):
        return True
    return any_phrase(text, weather_terms) and any_phrase(text, cane_regions)


def ethanol_feedstock_impact(candidate: dict) -> tuple[str, str]:
    text = " ".join(
        str(part)
        for part in (
            candidate.get("source_title", ""),
            " ".join(str(metric) for metric in (candidate.get("metrics") or [])),
            candidate.get("event_action", ""),
        )
        if part
    )
    if any_phrase(text, ("maize", "corn", "grain", "broken rice", "rice", "玉米", "碎米", "粮食")):
        return (
            "利空",
            "粮食乙醇供应增加会降低油销公司对甘蔗汁、B重糖蜜和糖浆乙醇的边际采购压力，使更多甘蔗糖源留在结晶产糖环节，增加白糖供应预期。",
        )
    if any_phrase(text, ("c-heavy", "c heavy", "c heavy molasses", "c-heavy molasses", "c重糖蜜", "c 重糖蜜")):
        return (
            "中性",
            "C重糖蜜制醇主要消化白糖结晶后的副产品，对当期白糖产出挤出较小，但会改善糖厂副产品变现和压榨现金流。",
        )
    if any_phrase(text, ("b-heavy", "b heavy", "b heavy molasses", "b-heavy molasses", "b重糖蜜", "b 重糖蜜")):
        return (
            "利多",
            "B重糖蜜制醇会在糖浆继续结晶前截留部分蔗糖，糖厂提高B重糖蜜制醇比例时，可结晶成白糖的蔗糖量减少。",
        )
    if any_phrase(text, ("sugar syrup", "syrup", "糖浆")):
        return (
            "利多",
            "糖浆制醇会把尚可继续结晶产糖的糖源直接送入乙醇装置，糖厂提高糖浆制醇比例时，白糖产出和可售糖源减少。",
        )
    if any_phrase(text, ("cane juice", "sugarcane juice", "甘蔗汁")):
        return (
            "利多",
            "甘蔗汁制醇会把原本可进入澄清和结晶流程的蔗糖直接送入乙醇装置，糖厂提高甘蔗汁制醇比例时，白糖产出被挤出。",
        )
    return (
        "利多",
        "若新增乙醇需求由B重糖蜜、糖浆或甘蔗汁满足，糖厂会把这些可发酵糖源送入乙醇装置；其中B重糖蜜、糖浆和甘蔗汁会减少可结晶成白糖的蔗糖量。",
    )


EN_MONTHS_CN = {
    "january": "1月",
    "jan": "1月",
    "february": "2月",
    "feb": "2月",
    "march": "3月",
    "mar": "3月",
    "april": "4月",
    "apr": "4月",
    "may": "5月",
    "june": "6月",
    "jun": "6月",
    "july": "7月",
    "jul": "7月",
    "august": "8月",
    "aug": "8月",
    "september": "9月",
    "sep": "9月",
    "sept": "9月",
    "october": "10月",
    "oct": "10月",
    "november": "11月",
    "nov": "11月",
    "december": "12月",
    "dec": "12月",
}


def format_cn_number(value: float) -> str:
    if abs(value - round(value)) < 0.005:
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def parse_float(value: str) -> float | None:
    try:
        return float(value.replace(",", ""))
    except ValueError:
        return None


def extract_supply_demand_cutoff(text: str) -> str:
    cn_match = re.search(r"(?:截至|截止|至)\s*((?:\d{4}年)?\d{1,2}月\d{1,2}日)", text)
    if cn_match:
        return f"截至{cn_match.group(1)}"
    en_month = "|".join(re.escape(month) for month in sorted(EN_MONTHS_CN, key=len, reverse=True))
    match = re.search(
        rf"(?i)\b(?:as of|as at|through|up to|until|by)\s+({en_month})\.?\s+(\d{{1,2}})(?:,\s*\d{{4}})?",
        text,
    )
    if match:
        month = EN_MONTHS_CN[match.group(1).lower().rstrip(".")]
        return f"截至{month}{int(match.group(2))}日"
    match = re.search(
        rf"(?i)\b(?:as of|as at|through|up to|until|by)\s+(\d{{1,2}})\s+({en_month})\.?(?:\s+\d{{4}})?",
        text,
    )
    if match:
        month = EN_MONTHS_CN[match.group(2).lower().rstrip(".")]
        return f"截至{month}{int(match.group(1))}日"
    return ""


def infer_supply_demand_metric_kind(text: str) -> str:
    lowered = text.lower()
    has_production = any_phrase(lowered, ("production", "output", "produced", "crop", "产量", "产糖", "产出"))
    has_inventory = any_phrase(lowered, ("inventory", "inventories", "stock", "stocks", "库存"))
    has_sales = any_phrase(lowered, ("sales", "sold", "sale", "销量", "销售", "产销率"))
    has_consumption = any_phrase(lowered, ("consumption", "demand", "consume", "消费", "需求"))
    if has_sales and has_inventory:
        return "sales_inventory"
    if has_production:
        return "production"
    if has_inventory:
        return "inventory"
    if has_sales:
        return "sales"
    if has_consumption:
        return "consumption"
    return "supply_demand"


def supply_demand_metric_label(kind: str) -> str:
    return {
        "production": "糖产量",
        "inventory": "库存",
        "sales": "销量",
        "consumption": "消费",
        "sales_inventory": "销量和库存",
        "supply_demand": "供需指标",
    }.get(kind, "供需指标")


def infer_supply_demand_direction_key(text: str) -> str:
    lowered = text.lower()
    if any_phrase(lowered, ("slow", "slows", "slowed", "slowing", "放缓", "趋缓")):
        return "slowing"
    if any_phrase(lowered, ("shortage", "deficit", "tight", "短缺", "缺口", "偏紧")):
        return "shortage"
    if any_phrase(lowered, ("surplus", "overhang", "过剩")):
        return "surplus"
    if any_phrase(lowered, ("lower", "down", "decline", "declines", "declined", "fall", "falls", "fell", "drop", "drops", "dropped", "decrease", "decreased", "reduce", "reduced", "shrink", "shrinks", "下降", "减少", "下滑", "降低", "收缩")):
        return "down"
    if any_phrase(lowered, ("higher", "record", "up", "increase", "increases", "increased", "rise", "rises", "rose", "growth", "grow", "grows", "grown", "增加", "提高", "增长", "上升")):
        return "up"
    return ""


def supply_demand_direction_label(direction: str) -> str:
    return {
        "down": "下降",
        "up": "增加",
        "slowing": "增速放缓",
        "shortage": "缺口扩大",
        "surplus": "过剩",
    }.get(direction, "")


def extract_supply_demand_percent(texts: list[str]) -> str:
    for text in texts:
        match = re.search(r"(?i)(\d+(?:[.,]\d+)?)\s*(?:%|percent)", text)
        if match:
            return f"{match.group(1).replace(',', '.')}%"
    return ""


def extract_supply_demand_volume(text: str) -> str:
    cn_match = re.search(r"(\d+(?:[.,]\d+)?)\s*万吨", text)
    if cn_match:
        return f"{cn_match.group(1).replace(',', '.')}万吨"
    patterns = (
        (r"(?i)(\d+(?:[.,]\d+)?)\s*million\s*(?:metric\s*)?(?:tonnes?|tons?|mt)\b", 100.0),
        (r"(?i)(\d+(?:[.,]\d+)?)\s*(?:lmt|lakh\s*(?:metric\s*)?tonnes?|lakh\s*tons?)\b", 10.0),
        (r"(?i)(\d+(?:[.,]\d+)?)\s*(?:metric\s*)?(?:tonnes?|tons?|mt)\b", 0.0001),
    )
    for pattern, multiplier in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        value = parse_float(match.group(1))
        if value is not None:
            return f"{format_cn_number(value * multiplier)}万吨"
    return ""


def supply_demand_metric_text(candidate: dict, metrics: list[str]) -> str:
    title = str(candidate.get("source_title", ""))
    metric_haystack = " ".join([title, *(str(metric) for metric in candidate.get("metrics") or []), *metrics])
    kind = infer_supply_demand_metric_kind(metric_haystack)
    label = supply_demand_metric_label(kind)
    direction = infer_supply_demand_direction_key(metric_haystack)
    direction_label = supply_demand_direction_label(direction)
    cutoff = extract_supply_demand_cutoff(metric_haystack)
    percent = extract_supply_demand_percent([metric_haystack, *metrics])
    volume = extract_supply_demand_volume(metric_haystack)
    prefix = cutoff
    if direction_label:
        phrase = f"{prefix}{label}{direction_label}"
        if percent:
            phrase += percent
        if volume:
            phrase += f"至{volume}"
        return phrase
    if volume:
        return f"{prefix}{label}为{volume}"
    if percent:
        return f"{prefix}{label}变化幅度为{percent}"
    if metrics:
        return f"{prefix}{label}披露{'、'.join(metrics[:4])}"
    raise ValueError(f"{label}RSS候选缺少可核验数值或方向")


def metric_text_for_candidate(candidate: dict, metrics: list[str]) -> str:
    if candidate.get("topic") == "supply_demand":
        return supply_demand_metric_text(candidate, metrics)
    title = str(candidate.get("source_title", ""))
    lowered_title = title.lower()
    topic = candidate.get("topic")
    if topic == "price_market":
        percent = next((metric for metric in metrics if "%" in metric), "")
        if percent:
            period = "一个月" if any_phrase(lowered_title, ("month", "monthly")) else ""
            if any_phrase(lowered_title, ("jump", "surge", "rise", "rose", "higher", "上涨")):
                return f"糖价{period}上涨{percent}"
            if any_phrase(lowered_title, ("fall", "fell", "drop", "lower", "down", "下跌")):
                return f"糖价{period}下跌{percent}"
            return f"糖价变动幅度为{percent}"
        if metrics:
            return f"糖价或相关政策基准为{'、'.join(metrics[:4])}"
        raise ValueError("price-market RSS title lacks price level or change amount")
    if topic == "trade_policy":
        volume = extract_supply_demand_volume(title + " " + " ".join(metrics))
        if volume:
            if any_phrase(lowered_title, ("import", "imports", "进口")):
                return f"进口数量为{volume}"
            if any_phrase(lowered_title, ("export", "exports", "出口")):
                return f"出口数量为{volume}"
            return f"贸易政策数量为{volume}"
        if metrics:
            return f"贸易政策披露{'、'.join(metrics[:4])}"
        raise ValueError("trade-policy RSS title lacks import/export/quota/tariff detail")
    if topic == "weather_pest":
        area_metric = next((metric for metric in metrics if "公顷" in metric or "hectare" in metric.lower()), "")
        lgu_match = re.search(r"(?i)(\d+(?:[.,]\d+)?)\s+more\s+LGUs?", title)
        if "spray" in lowered_title and "fungi" in lowered_title and area_metric:
            local_area = localize_metric_for_summary(area_metric)
            if lgu_match:
                return f"已用真菌处理{local_area}甘蔗虫害地块，并计划推进到另外{int(float(lgu_match.group(1).replace(',', '')))}个地方政府辖区"
            return f"已用真菌处理{local_area}甘蔗虫害地块"
        if "seeks national aid" in lowered_title and any_phrase(lowered_title, ("pest", "sugarcane pest")):
            return "因甘蔗虫害扩散寻求国家援助"
        if metrics:
            return f"病虫害或天气事实披露{'、'.join(metrics[:4])}"
        return "已说明甘蔗虫害扩散或产区天气风险"
    if metrics:
        return "披露" + "、".join(metrics[:4])
    raise ValueError("RSS候选缺少可核验数值、政策条款、生产、贸易、价格或天气事实")


def supply_demand_transmission(candidate: dict) -> str:
    title = str(candidate.get("source_title", ""))
    metric_haystack = " ".join([title, *(str(metric) for metric in candidate.get("metrics") or [])])
    kind = infer_supply_demand_metric_kind(metric_haystack)
    direction = infer_supply_demand_direction_key(metric_haystack)
    if kind == "production":
        if direction in {"down", "shortage"}:
            return "糖产量下降会减少当期新增可售糖源，收紧供应端并支撑糖价。"
        if direction in {"up", "surplus"}:
            return "糖产量增加会提高当期新增可售糖源，缓解供应压力并压制糖价。"
    if kind == "inventory":
        if direction in {"down", "shortage"}:
            return "库存下降会削弱现货供应缓冲，放大补库需求对价格的支撑。"
        if direction in {"up", "surplus"}:
            return "库存增加会提高现货供应缓冲，压制补库需求和糖价。"
    if kind in {"sales", "consumption"}:
        if direction == "up":
            return "销量或消费增加会加快库存消化并扩大需求端吸收，降低库存缓冲并支撑糖价。"
        if direction in {"down", "slowing"}:
            return "销量或消费下降或增速放缓会放慢库存消化，削弱需求端支撑并压制糖价。"
    if kind == "sales_inventory":
        return "销量决定需求端消化速度，库存决定现货供应缓冲；两项指标需按各自方向判断对糖价的拉动或压制。"
    if direction in {"down", "shortage"}:
        return "供需指标走弱或供应缺口扩大说明可用糖源收紧，容易支撑糖价。"
    if direction in {"up", "surplus"}:
        return "供需指标走强或供应过剩说明可用糖源增加，容易压制糖价。"
    return "该供需指标需要明确对应产量、库存、销量或消费方向后，才能判断供给端或需求端影响。"


def supply_demand_impact(candidate: dict) -> str:
    logic = supply_demand_transmission(candidate).rstrip("。")
    direction = infer_supply_demand_direction_key(
        " ".join([str(candidate.get("source_title", "")), *(str(metric) for metric in candidate.get("metrics") or [])])
    )
    kind = infer_supply_demand_metric_kind(str(candidate.get("source_title", "")))
    if kind in {"sales", "consumption"}:
        if direction == "up":
            return f"利多：{logic}。"
        if direction in {"down", "slowing"}:
            return f"利空：{logic}。"
    if direction in {"down", "shortage"}:
        return f"利多：{logic}。"
    if direction in {"up", "surplus"}:
        return f"利空：{logic}。"
    return f"中性：{logic}。"


def impact_for_candidate(candidate: dict) -> str:
    topic = candidate.get("topic") or "general_industry"
    text = str(candidate.get("source_title", "")).lower()
    if topic in {"ethanol_policy", "ethanol_capacity"}:
        direction, logic = ethanol_feedstock_impact(candidate)
        logic = logic.rstrip("。")
        if direction == "利多":
            return f"{direction}：{logic}，从而减少食糖供应预期并支撑糖价。"
        if direction == "利空":
            return f"{direction}：{logic}，从而压制糖价。"
        return f"{direction}：{logic}，对糖价方向暂不单边。"
    if topic == "weather_pest":
        if any_phrase(text, ("rain", "rainfall", "monsoon", "thunderstorm", "降雨", "大雨")) and not any_phrase(text, ("damage", "flood damage", "drought", "deficit", "loss", "干旱", "受灾", "损失")):
            return "利空：生长期降雨增加有利于补充甘蔗产区土壤水分并改善单产预期，从而提高后续糖料供应。"
        return "利多：干旱、洪涝或病虫害会压低甘蔗单产并削弱糖料供应稳定性，从而支撑糖价。"
    if topic == "cane_farming":
        if any_phrase(text, ("pest", "insect", "disease", "red rot", "white grub", "spreads", "spread", "infestation", "病虫害", "虫害", "病害", "扩散")):
            return "利多：病虫害扩散会降低受害地块甘蔗单产和可入榨糖料，若防治不及时将下调食糖产量预期并支撑糖价。"
        if any_phrase(text, ("cane dues", "payment", "arrears", "蔗款", "欠款")):
            return "利空：蔗款支付改善会缓解蔗农现金流压力并稳定交蔗积极性，有利于下一季糖料供应。"
        if any_phrase(text, ("acreage", "area", "planting", "planted", "面积", "种植")) and any_phrase(text, ("increase", "rose", "rises", "up", "higher", "扩大", "增加", "增长", "提高")):
            return "利空：甘蔗种植面积扩大将提高下一季可入榨糖料和食糖产量预期，对糖价形成压力。"
        return "中性：原文只说明甘蔗生产或蔗农经营变化，尚未给出可判断糖料供应增减的明确方向。"
    if topic == "variety_research":
        return "利空：高单产或抗病甘蔗品种推广会改善中长期糖料供应潜力，增加未来食糖产量预期。"
    if topic == "mill_operations":
        if any_phrase(text, ("shutdown", "halt", "closed", "accident", "shortage", "停产", "关闭", "原料不足")):
            return "利多：糖厂停产、事故或原料不足会拖慢压榨和产糖节奏，减少阶段性食糖供应。"
        return "利空：糖厂运行、压榨能力或原料供应改善会提高食糖生产节奏，增加阶段性供应。"
    if topic == "trade_policy":
        candidate_country = str(candidate.get("country_group") or candidate.get("event_country") or candidate.get("country") or "")
        full_text = " ".join(str(candidate.get(field, "")) for field in ("source_title", "event_action", "impact_logic")).lower()
        if (
            candidate_country == "印度"
            and any_phrase(full_text, ("100% sugar import duty", "100% import duty", "import tax", "import duty", "进口税", "关税"))
            and any_phrase(full_text, ("record", "high", "price", "shortage", "创纪录", "创新高", "价格", "短缺", "紧张"))
        ):
            return "利多：印度在国内糖价创纪录或供应紧张背景下评估削减进口税，说明本地可用糖源偏紧并强化供应紧张预期；若实际进口到港，后续才会补充供应并压制现货。"
        if any_phrase(text, ("import", "进口")):
            return "利空：进口政策放宽或进口量增加会补充国内可用糖源，对本地糖价形成压力。"
        return "利多：出口限制、配额收紧或贸易成本上升会减少可流通糖源并支撑国际糖价。"
    if topic == "starch_sugar_substitute":
        return "利空：淀粉糖、玉米糖浆或预混粉供应增加会替代部分食糖消费，削弱白糖需求。"
    if topic == "supply_demand":
        return supply_demand_impact(candidate)
    if topic == "price_market":
        if any_phrase(text, ("rise", "rises", "higher", "up", "上涨", "上调")):
            return "利多：现货或出厂报价上涨反映阶段性供应偏紧或采购需求增强，会支撑短期糖价。"
        if any_phrase(text, ("fall", "lower", "down", "下跌", "下调")):
            return "利空：现货或出厂报价下跌反映供应压力或需求转弱，会压制短期糖价。"
        return "中性：价格信息缺少明确涨跌幅或区域基准，暂不改变供需判断。"
    return "中性：原文未说明产量、库存、贸易、价格或糖料变化的方向，对糖价方向暂不单边。"


def rss_summary_for_publication(candidate: dict) -> tuple[str, str]:
    country = candidate.get("event_country") or "相关地区"
    actor = candidate.get("event_actor") or f"{country}糖业相关机构"
    action = candidate.get("event_action") or "公布"
    topic = candidate.get("topic") or "general_industry"
    label = TOPIC_LABELS.get(topic, TOPIC_LABELS["general_industry"])
    title = str(candidate.get("source_title", "")).strip()
    lowered_title = title.lower()
    metrics = [localize_metric_for_summary(str(metric)) for metric in (candidate.get("metrics") or [])]
    source_suffix = f"来源：{candidate.get('publisher') or '原始来源'}（{candidate.get('source_url') or ''}）"
    if country == "印度" and "sugar export ban" in lowered_title and "smuggling" in lowered_title:
        impact = "利多：出口禁令限制印度正规糖源外流，节前需求转向非正规贸易会加剧周边市场供应紧张。"
        candidate["impact_direction"] = "利多"
        candidate["impact_logic"] = impact.split("：", 1)[1]
        news = (
            "印度食糖出口禁令继续影响南亚节前贸易流向，尼泊尔市场出现走私增加的报道。"
            "正规出口受限会减少周边市场可获得糖源，并使节前补库需求转向非正规渠道，从而支撑区域糖价。"
            f"{source_suffix}"
        )
        return news, impact
    if country == "印度" and "pdm" in lowered_title and "rs" in lowered_title and "sugar mills" in lowered_title:
        metric = "、".join(metrics[:2]) if metrics else "公开价格"
        if re.search(r"(?i)rs\.?\s*4263\s*/\s*mt", title):
            metric = "Rs 4263/MT"
        impact = "利空：糖厂副产品销售价格明确有助于改善现金流，间接支持压榨和制糖经营稳定。"
        candidate["impact_direction"] = "利空"
        candidate["impact_logic"] = impact.split("：", 1)[1]
        news = (
            f"印度中央政府促成糖厂以{metric}销售PDM，交易主体为印度糖厂。"
            "副产品销售变现有助于改善糖厂现金流和运营稳定性，降低压榨季资金压力。"
            f"{source_suffix}"
        )
        return news, impact
    if country == "印度" and "without ethanol" in lowered_title and any("Rs 125" in metric for metric in metrics):
        impact = "利多：乙醇掺混扩大燃料端需求；若新增需求由甘蔗汁、B重糖蜜或糖浆满足，糖厂用于结晶产糖的蔗糖量会减少，从而支撑糖价。"
        candidate["impact_direction"] = "利多"
        candidate["impact_logic"] = impact.split("：", 1)[1]
        news = (
            "印度政府称，如果没有乙醇掺混，汽油价格可能达到每升Rs 125。"
            "乙醇掺混降低燃料成本的同时扩大燃料乙醇需求；当新增需求由甘蔗汁、B重糖蜜或糖浆满足时，这些可发酵糖源会进入乙醇装置而不是继续结晶产糖，食糖供应预期下降。"
            f"{source_suffix}"
        )
        return news, impact
    if country == "印度" and topic == "price_market" and "sugar price" in lowered_title and any("17%" in metric for metric in metrics):
        impact = "利空：库存限制会压缩贸易商和终端囤货空间，削弱价格上涨后的补库需求，短期压制印度现货糖价。"
        candidate["impact_direction"] = "利空"
        candidate["impact_logic"] = impact.split("：", 1)[1]
        if "mills expect" in lowered_title or "curb" in lowered_title:
            news = (
                "印度糖价一个月上涨17%，糖厂预计政府会在库存限制落地后继续加强市场调控。"
                "库存限制压缩贸易商和终端囤货空间，削弱价格上涨后的补库需求，短期利空印度现货糖价。"
                f"{source_suffix}"
            )
        else:
            news = (
                "印度糖价上涨17%，新的库存限制已开始执行。"
                "库存限制压缩贸易商和终端囤货空间，削弱价格上涨后的补库需求，短期利空印度现货糖价。"
                f"{source_suffix}"
            )
        return news, impact
    if country == "菲律宾" and topic == "weather_pest" and "negros oriental" in lowered_title and "national aid" in lowered_title:
        impact = "利多：东内格罗斯甘蔗虫害扩散并需要国家援助，说明地方防控和蔗农现金流压力上升，若处置不及时将削弱糖料供应稳定性。"
        candidate["impact_direction"] = "利多"
        candidate["impact_logic"] = impact.split("：", 1)[1]
        news = (
            "菲律宾东内格罗斯省因甘蔗虫害扩散寻求国家援助，地方政府希望中央支持受灾蔗农和防控行动。"
            "虫害扩散会压低受害地块单产并增加蔗农补救成本，若防控资金不到位，内格罗斯糖料供应稳定性下降，利多糖价。"
            f"{source_suffix}"
        )
        return news, impact
    if country == "菲律宾" and topic == "weather_pest" and "spray" in lowered_title and "fungi" in lowered_title:
        area_metric = next((metric for metric in metrics if "公顷" in metric or "hectare" in metric.lower()), "")
        area_text = localize_metric_for_summary(area_metric) if area_metric else "部分"
        lgu_match = re.search(r"(?i)(\d+(?:[.,]\d+)?)\s+more\s+LGUs?", title)
        lgu_text = f"另外{int(float(lgu_match.group(1).replace(',', '')))}个地方政府辖区" if lgu_match else "更多辖区"
        impact = "利多：西内格罗斯已出现需要集中处理的甘蔗虫害地块，若真菌防治未能及时控制扩散，将压低受害区域单产并削弱糖料供应稳定性。"
        candidate["impact_direction"] = "利多"
        candidate["impact_logic"] = impact.split("：", 1)[1]
        news = (
            f"菲律宾西内格罗斯省虫害防控工作组已用真菌处理{area_text}甘蔗虫害地块，并计划把防治推进到{lgu_text}。"
            "生物防治扩大有助于压低虫害继续蔓延的风险，但已处理面积说明内格罗斯蔗区仍存在虫害压力，短期利多糖价。"
            f"{source_suffix}"
        )
        return news, impact
    if country == "菲律宾" and "mindanao" in lowered_title and "sugarcane pest" in lowered_title:
        metric = "、".join(metrics[:2]) if metrics else "多个府省"
        impact = "利多：病虫害扩散会压低甘蔗单产并削弱糖料供应稳定性，从而支撑糖价。"
        candidate["impact_direction"] = "利多"
        candidate["impact_logic"] = impact.split("：", 1)[1]
        news = (
            f"菲律宾棉兰老岛蔗农组织预警，甘蔗病虫害已扩散至{metric}。"
            "病虫害扩大将影响甘蔗生长和可收获糖料，若防控不及时，后续食糖产量预期会下调。"
            f"{source_suffix}"
        )
        return news, impact
    metric_text = metric_text_for_candidate(candidate, metrics)
    if topic == "price_market" and not metrics:
        raise ValueError("price-market RSS title lacks price level or change amount")
    if topic == "general_industry":
        raise ValueError("general RSS candidate needs a specific sugar supply, demand, trade, price, weather, or ethanol topic before publication")
    location_clause = "" if country and country in actor else f"，涉及{country}"
    summary_label = label
    if topic == "supply_demand":
        summary_label = f"{supply_demand_metric_label(infer_supply_demand_metric_kind(title + ' ' + ' '.join(metrics)))}数据"
    first = f"{actor}{action}{summary_label}，{metric_text}{location_clause}。"
    ethanol_transmission = ethanol_feedstock_impact(candidate)[1] if topic in {"ethanol_policy", "ethanol_capacity"} else ""
    transmission = {
        "ethanol_policy": ethanol_transmission,
        "ethanol_capacity": ethanol_transmission,
        "mill_operations": f"{country}糖厂若增加开工或压榨，会加快当季食糖产出；若停产或延后开榨，则阶段性供应收紧。",
        "cane_farming": "甘蔗价格、面积或蔗款变化会影响蔗农种植意愿和下一季糖料供应。",
        "variety_research": "新品种单产、糖分或抗病性变化会影响中长期甘蔗供应潜力。",
        "weather_pest": f"{country}产区若出现虫害、干旱或洪涝，会压低受影响地块单产并削弱糖料供应稳定性。",
        "supply_demand": supply_demand_transmission(candidate),
        "trade_policy": "进口放宽会补充进口国国内糖源并压制本地现货；出口配额增加会提高国际可流通糖源，出口限制则减少国际供应。",
        "starch_sugar_substitute": "替代糖源供应变化会影响白糖消费替代和终端需求。",
        "price_market": f"{country}食糖价格上涨说明现货供应偏紧或节前需求增强，价格下跌则说明供应压力或需求转弱正在传导到现货端。",
        "general_industry": "候选新闻缺少可发布的具体糖业主题，已退回核实。",
    }[topic]
    impact = impact_for_candidate(candidate)
    candidate["impact_direction"] = impact.split("：", 1)[0]
    candidate["impact_logic"] = impact.split("：", 1)[1] if "：" in impact else impact
    return f"{first}{transmission}{source_suffix}", impact


THAI_TMD_CANE_PROVINCES = (
    "Udon Thani",
    "Khon Kaen",
    "Nakhon Ratchasima",
    "Chaiyaphum",
    "Kalasin",
    "Loei",
    "Nakhon Sawan",
    "Kamphaeng Phet",
    "Sukhothai",
    "Phitsanulok",
    "Kanchanaburi",
    "Lopburi",
    "Suphanburi",
    "Chai Nat",
    "Sa Kaeo",
    "Chon Buri",
    "Chonburi",
)

THAI_TMD_PROVINCE_CN = {
    "Udon Thani": "乌隆他尼",
    "Khon Kaen": "孔敬",
    "Nakhon Ratchasima": "呵叻",
    "Chaiyaphum": "猜也蓬",
    "Kalasin": "加拉信",
    "Loei": "黎府",
    "Nakhon Sawan": "那空沙旺",
    "Kamphaeng Phet": "甘烹碧",
    "Sukhothai": "素可泰",
    "Phitsanulok": "彭世洛",
    "Kanchanaburi": "北碧",
    "Lopburi": "华富里",
    "Suphanburi": "素攀武里",
    "Chai Nat": "猜纳",
    "Sa Kaeo": "沙缴",
    "Chon Buri": "春武里",
    "Chonburi": "春武里",
}

THAI_TMD_REGION_CANE_AREAS = {
    "northeastern": {
        "label": "东北部核心甘蔗产区",
        "areas": ("呵叻", "孔敬", "乌隆他尼", "猜也蓬", "加拉信", "黎府"),
    },
    "northern": {
        "label": "北部甘蔗产区",
        "areas": ("那空沙旺", "甘烹碧", "素可泰", "彭世洛"),
    },
    "central": {
        "label": "中部及西部甘蔗产区",
        "areas": ("北碧", "华富里", "素攀武里", "猜纳"),
    },
    "eastern": {
        "label": "东部补充产区",
        "areas": ("沙缴", "春武里"),
    },
}

THAI_TMD_RAIN_PHRASES = (
    "rain",
    "rains",
    "showers",
    "shower",
    "thundershower",
    "thundershowers",
    "thunderstorm",
    "thunderstorms",
    "heavy rain",
    "heavy rains",
)

THAI_OPEN_METEO_POINTS = (
    {"region": "东北部", "province": "呵叻", "lat": 14.9799, "lon": 102.0977},
    {"region": "东北部", "province": "孔敬", "lat": 16.4419, "lon": 102.8350},
    {"region": "东北部", "province": "乌隆他尼", "lat": 17.4138, "lon": 102.7872},
    {"region": "东北部", "province": "猜也蓬", "lat": 15.8068, "lon": 102.0315},
    {"region": "东北部", "province": "加拉信", "lat": 16.4322, "lon": 103.5066},
    {"region": "东北部", "province": "黎府", "lat": 17.4860, "lon": 101.7223},
    {"region": "北部", "province": "那空沙旺", "lat": 15.7047, "lon": 100.1372},
    {"region": "北部", "province": "甘烹碧", "lat": 16.4828, "lon": 99.5227},
    {"region": "北部", "province": "素可泰", "lat": 17.0078, "lon": 99.8230},
    {"region": "北部", "province": "彭世洛", "lat": 16.8211, "lon": 100.2659},
    {"region": "中部及西部", "province": "北碧", "lat": 14.0228, "lon": 99.5328},
    {"region": "中部及西部", "province": "华富里", "lat": 14.7995, "lon": 100.6534},
    {"region": "中部及西部", "province": "素攀武里", "lat": 14.4745, "lon": 100.1177},
    {"region": "中部及西部", "province": "猜纳", "lat": 15.1852, "lon": 100.1251},
    {"region": "东部", "province": "沙缴", "lat": 13.8240, "lon": 102.0646},
    {"region": "东部", "province": "春武里", "lat": 13.3611, "lon": 100.9847},
)


def plain_text_from_html(html_text: str) -> str:
    text = re.sub(r"(?is)<script.*?</script>|<style.*?</style>", " ", html_text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", unescape(text)).strip()


def tmd_region_rainfall_matches(text: str) -> list[dict]:
    lower = text.lower()
    matches = []
    for region_key, config in THAI_TMD_REGION_CANE_AREAS.items():
        idx = lower.find(region_key)
        if idx < 0:
            continue
        window = lower[idx:idx + 900]
        if any(phrase in window for phrase in THAI_TMD_RAIN_PHRASES):
            matches.append(config)
    return matches


def tmd_thai_weather_item_from_text(text: str, report_date: str, source_url: str = TMD_DAILY_FORECAST_URL) -> dict | None:
    lower = text.lower()
    if not any(term in lower for term in ("rain", "thundershower", "thunderstorm", "heavy rain", "drought")):
        return None
    matched = []
    for province in THAI_TMD_CANE_PROVINCES:
        if province.lower() in lower:
            cn = THAI_TMD_PROVINCE_CN[province]
            if cn not in matched:
                matched.append(cn)
    region_matches = [] if matched else tmd_region_rainfall_matches(text)
    if not matched and not region_matches:
        return None

    date_match = re.search(r"Forecast Date:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})", text)
    forecast_date = date_match.group(1) if date_match else report_date
    issue_match = re.search(r"Issued at\s+([0-9.]+\s*[ap]\.m\.)", text, flags=re.IGNORECASE)
    issue_text = f"{forecast_date} {issue_match.group(1)}" if issue_match else forecast_date
    has_heavy = any(term in lower for term in ("isolated heavy rain", "heavy rains", "heavy rain"))
    rain_desc = "雷阵雨并伴有局地大雨" if has_heavy else "雷阵雨或降雨"
    if matched:
        province_text = "、".join(matched[:8])
        if len(matched) > 8:
            province_text += "等"
    else:
        province_text = "；".join(
            f"{match['label']}（{'、'.join(match['areas'])}）"
            for match in region_matches
        )
    news = (
        f"泰国气象局预报（{issue_text}），{province_text}等主要甘蔗产区预计出现{rain_desc}。"
        "当前处于甘蔗生长阶段，强降雨、雷阵雨以及预报大雨均有利于补充产区土壤水分，"
        "促进甘蔗生长和单产形成，提高后期甘蔗及食糖产量预期。"
        f"来源：泰国气象局（{source_url}）"
    )
    return {
        "country_group": "泰国",
        "country": "泰国",
        "title": "泰国主要甘蔗产区预计出现降雨",
        "news": news,
        "impact": "利空：甘蔗生长阶段的降雨有利于补充土壤水分、改善墒情并促进甘蔗生长和单产形成，从而增加未来甘蔗及食糖供应预期，因此利空糖价。",
        "source_name": "泰国气象局",
        "source_url": source_url,
        "published_date_local": report_date,
        "event_date": report_date,
        "date_status": "official_forecast",
        "dedupe_key": f"thailand_cane_weather_tmd_{report_date.replace('-', '')}",
        "importance": 78,
    }


def fetch_tmd_thai_weather_item(report_date: str) -> tuple[dict | None, dict]:
    entry = {
        "country": "泰国",
        "language": "en",
        "keywords": "TMD daily forecast Thailand main sugarcane provinces rainfall",
        "source_url": TMD_DAILY_FORECAST_URL,
        "request_status": "pending",
        "returned_count": 0,
        "retained_count": 0,
        "filtered": [],
        "fixed_step": "Thai main sugarcane area rainfall check",
    }
    try:
        req = Request(TMD_DAILY_FORECAST_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=RSS_AUTOGEN_TIMEOUT_SECONDS) as resp:
            body = resp.read().decode("utf-8", errors="replace")
        text = plain_text_from_html(body)
        entry["request_status"] = "executed"
        entry["returned_count"] = 1
        item = tmd_thai_weather_item_from_text(text, report_date)
        if item:
            entry["retained_count"] = 1
        else:
            entry["filtered"].append({"reason": "No valid rainfall forecast for main Thai sugarcane provinces found in TMD daily forecast."})
        return item, entry
    except Exception as exc:
        entry["request_status"] = "failed"
        entry["error"] = str(exc)[:500]
        return None, entry


def fetch_open_meteo_thai_weather_item(report_date: str) -> tuple[dict | None, dict]:
    entry = {
        "country": "泰国",
        "language": "api",
        "keywords": "Open-Meteo Thailand main sugarcane provinces precipitation forecast",
        "source_url": OPEN_METEO_FORECAST_URL,
        "request_status": "pending",
        "returned_count": 0,
        "retained_count": 0,
        "filtered": [],
        "fixed_step": "Thai main sugarcane area rainfall check fallback",
    }
    lats = ",".join(f"{point['lat']:.4f}" for point in THAI_OPEN_METEO_POINTS)
    lons = ",".join(f"{point['lon']:.4f}" for point in THAI_OPEN_METEO_POINTS)
    url = (
        f"{OPEN_METEO_FORECAST_URL}?latitude={lats}&longitude={lons}"
        "&daily=precipitation_sum,precipitation_probability_max"
        "&timezone=Asia%2FBangkok&forecast_days=7"
    )
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=RSS_AUTOGEN_TIMEOUT_SECONDS) as resp:
            payload = json.loads(resp.read().decode("utf-8", errors="replace"))
        entries = payload if isinstance(payload, list) else [payload]
        entry["request_status"] = "executed"
        entry["returned_count"] = len(entries)
    except Exception as exc:
        entry["request_status"] = "failed"
        entry["error"] = str(exc)[:500]
        return None, entry

    rainy_points = []
    forecast_start = None
    forecast_end = None
    available_days = 0
    for idx, point_payload in enumerate(entries[:len(THAI_OPEN_METEO_POINTS)]):
        daily = point_payload.get("daily") if isinstance(point_payload, dict) else None
        if not isinstance(daily, dict):
            continue
        dates = daily.get("time") or []
        precip = daily.get("precipitation_sum") or []
        probs = daily.get("precipitation_probability_max") or []
        usable = [
            (date, _number(value), _number(probs[i]) if i < len(probs) else None)
            for i, (date, value) in enumerate(zip(dates, precip))
        ]
        usable = [(date, value, prob) for date, value, prob in usable if value is not None]
        if not usable:
            continue
        forecast_start = forecast_start or usable[0][0]
        forecast_end = usable[-1][0]
        available_days = max(available_days, len(usable))
        total_rain = sum(value for _date, value, _prob in usable if value is not None)
        max_prob = max((prob for _date, _value, prob in usable if prob is not None), default=None)
        has_rain = total_rain > 0 or (max_prob is not None and max_prob >= 30)
        if not has_rain:
            continue
        point = THAI_OPEN_METEO_POINTS[idx]
        rainy_points.append({
            "region": point["region"],
            "province": point["province"],
            "total_rain": total_rain,
            "max_prob": max_prob,
        })

    if not rainy_points:
        entry["filtered"].append({"reason": "Open-Meteo returned no rainfall forecast for configured Thai cane points."})
        return None, entry

    by_region: dict[str, list[dict]] = {}
    for point in rainy_points:
        by_region.setdefault(point["region"], []).append(point)
    region_text = "；".join(
        f"{region}（{'、'.join(point['province'] for point in points[:4])}）"
        for region, points in by_region.items()
    )
    top_points = sorted(rainy_points, key=lambda point: point["total_rain"], reverse=True)[:3]
    top_text = "、".join(
        f"{point['province']}约{point['total_rain']:.1f}mm"
        for point in top_points
    )
    period = f"{forecast_start}至{forecast_end}" if forecast_start and forecast_end else f"未来{available_days}日"
    news = (
        f"Open-Meteo预报显示（{period}），泰国{region_text}等主要甘蔗产区存在降雨预报，"
        f"实际可用预报期为{available_days}日，累计预测降雨较高的监测点包括{top_text}。"
        "当前处于甘蔗生长阶段，降雨有利于补充土壤水分、改善墒情并促进甘蔗生长和单产形成，"
        f"提高后期甘蔗及食糖产量预期。来源：Open-Meteo（{OPEN_METEO_FORECAST_URL}）"
    )
    item = {
        "country_group": "泰国",
        "country": "泰国",
        "title": "泰国主要甘蔗产区预计出现降雨",
        "news": news,
        "impact": "利空：甘蔗生长阶段的降雨有利于补充土壤水分、改善墒情并促进甘蔗生长和单产形成，从而增加未来甘蔗及食糖供应预期，因此利空糖价。",
        "source_name": "Open-Meteo",
        "source_url": OPEN_METEO_FORECAST_URL,
        "published_date_local": report_date,
        "event_date": forecast_start or report_date,
        "date_status": "official_forecast",
        "dedupe_key": f"thailand_cane_weather_open_meteo_{report_date.replace('-', '')}",
        "importance": 78,
    }
    entry["retained_count"] = 1
    entry["forecast_start"] = forecast_start
    entry["forecast_end"] = forecast_end
    entry["available_days"] = available_days
    entry["rainy_points"] = rainy_points
    return item, entry


def has_thai_weather_item(items: list[dict]) -> bool:
    for item in items:
        if item.get("country_group") != "泰国":
            continue
        text = " ".join(str(item.get(field, "")) for field in ("title", "news", "impact"))
        if _contains_any(text, THAI_WEATHER_TERMS):
            return True
    return False


def fallback_thai_weather_item_from_verified(report_date: str) -> tuple[dict | None, dict]:
    entry = {
        "request_status": "fallback_checked",
        "source": "recent verified Sugar News files",
        "retained_count": 0,
        "filtered": [],
    }
    root = PROJECT_ROOT / "data" / "verified_news"
    if not root.exists():
        entry["filtered"].append({"reason": "verified news root missing"})
        return None, entry
    candidates = sorted(root.rglob("sugar_news_*.json"), reverse=True)
    for path in candidates[:20]:
        try:
            with path.open("r", encoding="utf-8") as f:
                payload = json.load(f)
        except Exception as exc:
            entry["filtered"].append({"file": project_display_path(path), "reason": str(exc)[:200]})
            continue
        for item in payload.get("items") or []:
            item = normalize_country_fields(dict(item))
            if item.get("published_date_local") != report_date:
                continue
            if item.get("country_group") != "泰国":
                continue
            text = " ".join(str(item.get(field, "")) for field in ("title", "news", "impact"))
            if not _contains_any(text, THAI_WEATHER_TERMS):
                continue
            item["event_date"] = item.get("event_date") or report_date
            item["date_status"] = item.get("date_status") or "official_forecast"
            item["dedupe_key"] = item.get("dedupe_key") or f"thailand_cane_weather_verified_{report_date.replace('-', '')}"
            entry["retained_count"] = 1
            entry["source_file"] = project_display_path(path)
            return item, entry
    entry["filtered"].append({"reason": "no dated Thailand cane-area weather item found in recent verified files"})
    return None, entry


def ensure_thai_weather_item(
    data: dict,
    report_date: str,
    fetcher=fetch_tmd_thai_weather_item,
    open_meteo_fetcher=fetch_open_meteo_thai_weather_item,
) -> tuple[dict, dict]:
    items = data.get("items") or []
    log = {
        "fixed_step": "Thai main sugarcane area rainfall check",
        "target_date": report_date,
        "source_url": TMD_DAILY_FORECAST_URL,
        "status": "pending",
    }
    if has_thai_weather_item(items):
        log["status"] = "skipped"
        log["reason"] = "Thailand cane-area weather item already present in verified news."
        return data, log

    item, fetch_log = fetcher(report_date)
    log.update(fetch_log)
    if not item:
        open_meteo_item, open_meteo_log = open_meteo_fetcher(report_date)
        log["open_meteo_fallback"] = open_meteo_log
        item = open_meteo_item
    if not item:
        fallback_item, fallback_log = fallback_thai_weather_item_from_verified(report_date)
        log["fallback"] = fallback_log
        item = fallback_item
    if item:
        data = dict(data)
        data["items"] = [*items, item]
        log["status"] = "added"
    else:
        log["status"] = "not_added"
        log.setdefault("reason", "No valid Thailand cane-area rainfall forecast item was generated.")
    return data, log


def has_china_item(items: list[dict]) -> bool:
    return any(
        item.get("country_group") == "中国" or item.get("country") == "中国"
        for item in items
    )


def china_monitoring_note(report_date: str) -> dict:
    source_url = "https://github.com/lishi2626/sugar-news/actions"
    return {
        "country_group": "中国",
        "country": "中国",
        "title": "中国糖业每日监测",
        "news": (
            "已完成中国食糖、甘蔗、甜菜糖、进口、库存、现货价格和主产区天气等重点方向监测，"
            "未发现符合收录标准的新增重要糖业事件。"
            "该结果仅表示本次公开来源检索没有可发布的新事件，不代表中国食糖供需或价格没有变化。"
            f"来源：Sugar News中国糖业每日监测日志（{source_url}）"
        ),
        "impact": "中性：本次监测未发现足以改变中国食糖供需、库存、进口或价格预期的新增公开信息。",
        "source_name": "Sugar News中国糖业每日监测日志",
        "source_url": source_url,
        "published_date_local": report_date,
        "event_date": report_date,
        "date_status": "monitoring_completed",
        "dedupe_key": f"china_daily_monitoring_{report_date.replace('-', '')}",
        "importance": 40,
    }


def ensure_china_news_item(data: dict, report_date: str) -> tuple[dict, dict]:
    items = list(data.get("items") or [])
    log = {
        "fixed_step": "China sugar daily priority monitoring",
        "target_date": report_date,
        "status": "pending",
        "retained_count": 0,
    }
    if has_china_item(items):
        log["status"] = "skipped"
        log["reason"] = "China item already present in verified news."
        log["retained_count"] = sum(
            1 for item in items
            if item.get("country_group") == "中国" or item.get("country") == "中国"
        )
        return data, log

    log["status"] = "completed_no_publishable_item"
    log["reason"] = "China search completed without a publishable event; no public placeholder row was added."
    log["retained_count"] = 0
    return data, log


def persist_verified_news(task_root: Path, report_date: str, data: dict) -> Path:
    path = verified_json_path(task_root, report_date)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def autogenerate_verified_from_rss(task_root: Path, date_text: str) -> Path:
    dt = datetime.strptime(date_text, "%Y-%m-%d")
    readable = dt.strftime("%B %-d %Y") if os.name != "nt" else dt.strftime("%B %#d %Y")
    context = {
        "readable": readable,
        "day": dt.day,
        "month_name": dt.strftime("%B"),
        "month_name_pt": PORTUGUESE_MONTH_NAMES[dt.month],
        "month": dt.month,
        "year": dt.year,
        "date_slash": dt.strftime("%d/%m/%Y"),
        "buddhist_year": dt.year + 543,
    }
    country_templates = dict(COUNTRY_SEARCH_TEMPLATES)
    country_templates["印度指标"] = INDIA_PRICE_INVENTORY_SEARCH_TEMPLATES
    country_templates["其他国家"] = OTHER_COUNTRY_SEARCH_TEMPLATES
    items = []
    seen = set()
    structured_candidates = []
    query_window_start, query_window_end, query_window_rule = publication_window_for_target(date_text)
    search_log = {
        "target_date": date_text,
        "run_date": beijing_now().date().isoformat(),
        "search_tool": "Google News RSS autogeneration",
        "note": "Two-stage discovery: broad country/topic/source queries first, structured candidate verification second. Only verified candidates can be published.",
        "root_cause_diagnostics": {
            "previous_priority_country_only": "Old autogeneration executed only Brazil/India/Thailand/China plus one other-country bucket, so concrete other-country templates were not reached.",
            "previous_query_budget": "Old defaults capped each country at 12 queries, total at 72 queries, and only inspected the first 10 RSS rows per query.",
            "previous_exact_date_filter": "Old logic required publication date to equal target_date, missing weekend, timezone, and delayed-publication items.",
            "previous_vague_summary_generator": "Old RSS summaries used source-led generic wording and were then rejected by quality checks before page output.",
        },
        "source_matrix": SOURCE_MATRIX,
        "case_regression_topics": CASE_REGRESSION_TOPICS,
        "query_window": {
            "start": query_window_start.isoformat(),
            "end": query_window_end.isoformat(),
            "rule": query_window_rule,
        },
        "query_budget": {
            "max_queries_per_country": RSS_AUTOGEN_MAX_QUERIES_PER_COUNTRY,
            "max_total_queries": RSS_AUTOGEN_MAX_TOTAL_QUERIES,
            "max_items_per_query": RSS_AUTOGEN_MAX_ITEMS_PER_QUERY,
        },
        "india_price_inventory_sources": INDIA_PRICE_INVENTORY_SOURCE_GUIDE,
        "india_completeness_requirements": {
            "sugar_core": "India sugar production/stocks/prices/mills/sales quota/shortage searched",
            "ethanol_e20": "India ethanol policy/blending/E20/above 20 percent/OMC/feedstock searched",
            "reuters_site_search": "site:reuters.com India sugar/ethanol/E20/sugarcane/molasses searched",
            "weather": "India sugarcane rainfall and core cane-state forecasts searched",
            "no_country_cap": "Autogeneration does not stop after a fixed number of items per country.",
        },
        "thai_weather_requirements": {
            "fixed_step": "After Thailand sugar news discovery, check TMD daily forecast for main sugarcane provinces and add one weather item when a valid rainfall forecast exists.",
            "source": TMD_DAILY_FORECAST_URL,
            "provinces": [THAI_TMD_PROVINCE_CN[name] for name in THAI_TMD_CANE_PROVINCES if name in THAI_TMD_PROVINCE_CN],
        },
        "searches": [],
    }
    total_queries = 0
    for country, templates in country_templates.items():
        retained_for_country = 0
        queries_for_country = 0
        for language, template in templates:
            if queries_for_country >= RSS_AUTOGEN_MAX_QUERIES_PER_COUNTRY or total_queries >= RSS_AUTOGEN_MAX_TOTAL_QUERIES:
                search_log["searches"].append({
                    "country": country,
                    "language": language,
                    "keywords": template.format(**context),
                    "request_status": "skipped",
                    "returned_count": 0,
                    "retained_count": 0,
                    "filtered": [],
                    "reason": "RSS autogeneration query budget reached; daily job will continue with retained verified candidates.",
                })
                break
            query = template.format(**context)
            entry = {"country": country, "language": language, "keywords": query, "request_status": "pending", "returned_count": 0, "retained_count": 0, "filtered": []}
            total_queries += 1
            queries_for_country += 1
            print(f"[sugar-news:rss] {country} {queries_for_country}/{RSS_AUTOGEN_MAX_QUERIES_PER_COUNTRY}: {query}", flush=True)
            try:
                rss_items = fetch_rss(query, timeout=RSS_AUTOGEN_TIMEOUT_SECONDS)
                entry["request_status"] = "executed"
                entry["returned_count"] = len(rss_items)
            except Exception as exc:
                entry["request_status"] = "failed"
                entry["error"] = str(exc)[:500]
                search_log["searches"].append(entry)
                continue
            if country == "印度指标":
                entry["sample_results"] = rss_items[:5]
                for result in rss_items[:5]:
                    entry["filtered"].append({
                        "title": result.get("title"),
                        "news_date": result.get("published"),
                        "source": "Google News RSS",
                        "url": result.get("link"),
                        "stage": "price_inventory_verification",
                        "reason": "Price and stock indicators require source-page date, quote type, unit, and comparable-date verification before dashboard publication.",
                    })
                search_log["searches"].append(entry)
                continue
            for rss in rss_items[:RSS_AUTOGEN_MAX_ITEMS_PER_QUERY]:
                in_window, window_reason, item_date = rss_item_in_publication_window(rss, date_text)
                title_raw = rss.get("title", "").strip()
                title_clean, source = rss_source_from_title(title_raw)
                candidate_record = structured_candidate_from_rss(country, rss, date_text, title_clean, source)
                entry.setdefault("structured_candidates", []).append(candidate_record)
                structured_candidates.append(candidate_record)
                if not in_window:
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = window_reason
                    entry["filtered"].append({"title": title_raw, "reason": window_reason, "published": rss.get("published")})
                    continue
                haystack = f"{title_clean} {rss.get('description', '')}".lower()
                if is_medical_sugar_context(haystack):
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "medical blood-sugar/glucose/diabetes context, not sugar industry"
                    entry["filtered"].append({"title": title_raw, "reason": "medical blood-sugar/glucose/diabetes context, not sugar industry"})
                    continue
                if is_non_industry_sugar_context(haystack):
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "game, fiction, entertainment, recipe, or consumer content; not sugar industry"
                    entry["filtered"].append({"title": title_raw, "reason": "game, fiction, entertainment, recipe, or consumer content; not sugar industry"})
                    continue
                relevant = rss_sugar_relevant(country, haystack)
                if country == "印度" and is_india_indirect_sugar_relevant(haystack):
                    relevant = True
                if not relevant:
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "not sugar/rainfall/ethanol/indirect-sugar relevant"
                    entry["filtered"].append({"title": title_raw, "reason": "not sugar/rainfall/ethanol/indirect-sugar relevant"})
                    continue
                candidate_record["sugar_relevance"] = "relevant"
                concrete_country, country_group = candidate_record["event_country"], (
                    candidate_record["event_country"] if candidate_record["event_country"] in GROUP_ORDER else "其他国家"
                )
                if country not in {"其他国家", "印度指标"}:
                    concrete_country, country_group = infer_core_country(haystack, country)
                if country == "其他国家" and country_group != "其他国家":
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "other-country query found a priority-country item; keep only under the core country"
                    entry["filtered"].append({"title": title_raw, "reason": "other-country query found a priority-country item; keep only under the core country"})
                    continue
                if country == "其他国家" and country_group == "其他国家" and concrete_country in {"其他", "其他国家"}:
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "other-country item lacks concrete country/region; skipped before publication"
                    entry["filtered"].append({"title": title_raw, "reason": "other-country item lacks concrete country/region; skipped before publication"})
                    continue
                if country != "其他国家" and country_group != country:
                    entry.setdefault("reclassified", []).append({"title": title_raw, "from": country, "to": concrete_country, "reason": "core event country differs from search bucket"})
                candidate_record["event_country"] = concrete_country
                country_confidence = candidate_country_confidence(candidate_record, country, haystack, source)
                if not country_confidence:
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "event country not confirmed by title, summary, or trusted country source"
                    entry["filtered"].append({
                        "title": title_raw,
                        "reason": "event country not confirmed by title, summary, or trusted country source",
                    })
                    continue
                candidate_record["country_confidence"] = country_confidence
                fact_ok, fact_reason = candidate_has_verifiable_industry_fact(candidate_record)
                if not fact_ok:
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = fact_reason
                    entry["filtered"].append({"title": title_raw, "reason": fact_reason})
                    continue
                key = event_fingerprint(candidate_record) or re.sub(r"\W+", "", title_clean.lower())[:120]
                if key in seen:
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "duplicate event fingerprint"
                    continue
                seen.add(key)
                link = rss.get("link", "").strip()
                source_url, source_url_resolution = decode_google_news_source_url(link)
                candidate_record["source_url"] = source_url
                candidate_record["source_url_resolution"] = source_url_resolution
                if google_news_article_id(source_url):
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = "Google News RSS link could not be resolved to the original source"
                    entry["filtered"].append({
                        "title": title_raw,
                        "reason": "Google News RSS link could not be resolved to the original source",
                    })
                    continue
                try:
                    news, impact = rss_summary_for_publication(candidate_record)
                    candidate = normalize_country_fields({
                        "country_group": country_group,
                        "country": concrete_country,
                        "title": title_clean[:80],
                        "news": news,
                        "impact": impact,
                        "source_name": source,
                        "source_url": source_url,
                        "published_date_local": item_date or date_text,
                        "event_date": candidate_record["event_date"],
                        "date_status": "verified" if item_date == date_text else "continuing_impact",
                        "structured_candidate": candidate_record,
                        "dedupe_key": f"rss_{key}",
                        "importance": max(50, 90 - retained_for_country * 5),
                    })
                    validate_editorial_quality(candidate, len(items) + 1)
                except Exception as exc:
                    candidate_record["verification_status"] = "不采用"
                    candidate_record["drop_reason"] = f"candidate failed pre-publication quality check: {str(exc)[:300]}"
                    entry["filtered"].append({
                        "title": title_raw,
                        "reason": f"candidate failed pre-publication quality check: {str(exc)[:300]}",
                    })
                    continue
                candidate_record["verification_status"] = "已核实"
                candidate_record["published_to_report"] = True
                items.append(candidate)
                retained_for_country += 1
                entry["retained_count"] += 1
            search_log["searches"].append(entry)
    data_for_china = {
        "target_date": date_text,
        "run_date": beijing_now().date().isoformat(),
        "items": items,
    }
    data_for_china, china_log = ensure_china_news_item(data_for_china, date_text)
    search_log["searches"].append(china_log)
    items = data_for_china["items"]
    data_for_thai_weather = {
        "target_date": date_text,
        "run_date": beijing_now().date().isoformat(),
        "items": items,
    }
    data_for_thai_weather, thai_weather_log = ensure_thai_weather_item(data_for_thai_weather, date_text)
    search_log["searches"].append(thai_weather_log)
    items = data_for_thai_weather["items"]
    if not items:
        raise FileNotFoundError("RSS autogeneration found no publishable Sugar News items")
    status_counts = Counter(candidate.get("verification_status") for candidate in structured_candidates)
    drop_reason_counts = Counter(
        candidate.get("drop_reason")
        for candidate in structured_candidates
        if candidate.get("drop_reason")
    )
    executed_by_country = defaultdict(int)
    retained_by_country = defaultdict(int)
    failed_by_country = defaultdict(int)
    for entry in search_log["searches"]:
        country = entry.get("country") or entry.get("fixed_step") or "unknown"
        if entry.get("request_status") == "executed":
            executed_by_country[country] += 1
        if entry.get("request_status") == "failed":
            failed_by_country[country] += 1
        retained_by_country[country] += int(entry.get("retained_count") or 0)
    search_log["structured_candidates"] = structured_candidates[:1000]
    search_log["completeness_report"] = {
        "executed_queries_by_country": dict(executed_by_country),
        "failed_queries_by_country": dict(failed_by_country),
        "retained_items_by_country": dict(retained_by_country),
        "candidate_verification_status_counts": dict(status_counts),
        "drop_reason_counts": {str(key): value for key, value in drop_reason_counts.items()},
        "found_but_not_output_count": sum(
            1
            for candidate in structured_candidates
            if candidate.get("sugar_relevance") == "relevant" and not candidate.get("published_to_report")
        ),
        "single_source_country_failure": [
            country
            for country, templates in country_templates.items()
            if country not in {"印度指标"}
            and executed_by_country.get(country, 0) == 0
            and failed_by_country.get(country, 0) > 0
            and len(templates) <= 1
        ],
        "final_report_items": [
            {
                "country": item.get("country"),
                "title": item.get("title"),
                "dedupe_key": item.get("dedupe_key"),
            }
            for item in items
        ],
    }
    path = verified_json_path(task_root, date_text)
    path.parent.mkdir(parents=True, exist_ok=True)
    persist_verified_news(task_root, date_text, {
        "target_date": date_text,
        "run_date": beijing_now().date().isoformat(),
        "search_tool": "Google News RSS autogeneration",
        "items": items,
    })
    log_path = search_log_path(task_root, date_text)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    search_log["pipeline_counts"] = {"structured_data_count": len(items), "passed_to_excel": len(items)}
    with log_path.open("w", encoding="utf-8") as f:
        json.dump(search_log, f, ensure_ascii=False, indent=2)
    return path


def load_verified_or_fail(
    task_root: Path,
    date_text: str,
    offline_only: bool,
    allow_rss_autogen: bool = False,
    force_rss_autogen: bool = False,
) -> dict:
    path = verified_json_path(task_root, date_text)
    if force_rss_autogen:
        if offline_only:
            raise ValueError("--force-rss-autogen cannot be combined with --offline-only")
        path = autogenerate_verified_from_rss(task_root, date_text)
    if not path.exists() and not offline_only:
        if allow_rss_autogen:
            path = autogenerate_verified_from_rss(task_root, date_text)
        else:
            fallback_discovery(date_text, task_root)
    if not path.exists():
        raise FileNotFoundError(
            f"Missing verified Sugar News data: {path}. "
            "The job stopped before Excel/dashboard publication to avoid publishing blank or unverified content."
        )
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if data.get("target_date") != date_text:
        raise ValueError(f"target_date mismatch in {path}")
    return data


def _contains_any(text: str, terms: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(term.lower() in lowered for term in terms)


def validate_india_weather_impact(item: dict, idx: int) -> None:
    if item.get("country_group") != "印度":
        return

    fact_text = " ".join(str(item.get(field, "")) for field in ("title", "news"))
    text = f"{fact_text} {item.get('impact', '')}"
    if not _contains_any(text, INDIA_WEATHER_TERMS):
        return

    if _contains_any(fact_text, INDIA_WATER_STRESS_TERMS):
        if not item["impact"].startswith(("偏多糖价：", "利多：")):
            raise ValueError(f"India weather item {idx} indicates water-resource pressure and should be bullish")
        return

    in_main_area = _contains_any(fact_text, INDIA_MAIN_CANE_REGIONS)
    if not in_main_area:
        if not item["impact"].startswith("影响有限："):
            raise ValueError(f"India weather item {idx} is outside main cane regions and should be impact-limited")
        return

    bullish_prefixes = ("偏多糖价：", "利多：")
    bearish_prefixes = ("偏空糖价：", "利空：")

    if _contains_any(fact_text, INDIA_HARVEST_TERMS):
        if not item["impact"].startswith(bullish_prefixes):
            raise ValueError(f"India weather item {idx} indicates harvest/crushing disruption and should be bullish")
        return

    if _contains_any(fact_text, INDIA_DAMAGE_TERMS):
        if not item["impact"].startswith(bullish_prefixes):
            raise ValueError(f"India weather item {idx} indicates confirmed damage and should be bullish")
        return

    if _contains_any(fact_text, INDIA_DROUGHT_TERMS):
        if not item["impact"].startswith(bullish_prefixes):
            raise ValueError(f"India weather item {idx} indicates drought/rain shortage and should be bullish")
        return

    if _contains_any(fact_text, INDIA_RAIN_BENEFIT_TERMS):
        if not item["impact"].startswith(bearish_prefixes):
            raise ValueError(f"India weather item {idx} indicates growing-season rainfall support and should be bearish")


def validate_thai_weather_impact(item: dict, idx: int) -> None:
    if item.get("country_group") != "泰国":
        return

    fact_text = " ".join(str(item.get(field, "")) for field in ("title", "news"))
    text = f"{fact_text} {item.get('impact', '')}"
    if not _contains_any(text, THAI_WEATHER_TERMS):
        return
    if not _contains_any(fact_text, THAI_WEATHER_EVENT_TERMS):
        return

    in_main_area = (
        _contains_any(fact_text, THAI_MAIN_CANE_PROVINCES)
        or "东北部" in fact_text
        or "中部" in fact_text
    )
    if not in_main_area:
        if not item["impact"].startswith("影响有限："):
            raise ValueError(f"Thai weather item {idx} is outside main cane areas and should be impact-limited")
        return

    is_bearish = item["impact"].startswith(("偏空糖价：", "利空："))
    if _contains_any(fact_text, THAI_HARVEST_TERMS):
        if not item["impact"].startswith("偏多糖价："):
            raise ValueError(f"Thai weather item {idx} indicates harvest disruption and should be bullish")
        return

    if _contains_any(fact_text, THAI_DAMAGE_TERMS):
        return

    if _contains_any(fact_text, THAI_DROUGHT_TERMS):
        if not item["impact"].startswith("偏多糖价："):
            raise ValueError(f"Thai weather item {idx} indicates drought/rain shortage and should be bullish")
        return

    if _contains_any(fact_text, THAI_RAIN_INCREASE_TERMS) or _contains_any(fact_text, ("雷阵雨", "阵雨", "大雨", "强降雨")):
        if not is_bearish:
            raise ValueError(f"Thai weather item {idx} indicates growing-season rainfall improvement and should be bearish")


def normalize_items(data: dict) -> list[dict]:
    items = data.get("items") or []
    seen = set()
    normalized = []
    for idx, item in enumerate(items, start=1):
        item = normalize_country_fields(item)
        item = polish_verified_summary_text(item)
        for field in ("country_group", "country", "news", "impact", "source_name", "source_url", "published_date_local"):
            if not item.get(field):
                raise ValueError(f"Verified item {idx} missing {field}")
        if not item["impact"].startswith(IMPACT_PREFIXES):
            raise ValueError(f"Verified item {idx} has invalid impact prefix")
        item["news"] = ensure_news_impact_marker(item["news"], item["impact"])
        validate_news_impact_marker(item, idx)
        if any(text in item["news"] or text in item["impact"] for text in PLACEHOLDERS):
            raise ValueError(f"Verified item {idx} contains placeholder wording")
        if re.search(r"\bLMT\b|lmt", item["news"]):
            raise ValueError(f"Verified item {idx} contains raw LMT/lmt unit")
        if "来源：" not in item["news"] or item["source_url"] not in item["news"]:
            raise ValueError(f"Verified item {idx} missing B-column source link")
        if item["country_group"] == "其他国家" and item["country"] == "其他":
            raise ValueError("Other-country rows must use the concrete country/region name, not 其他")
        if item["country"] == "中国" and item["country_group"] != "中国":
            raise ValueError("China news must use country_group=中国 and must not be stored as other-country news")
        if item["country_group"] == "中国" and item["country"] != "中国":
            raise ValueError("country_group=中国 rows must use country=中国")
        if item["published_date_local"] != data["target_date"] and item.get("date_status") != "continuing_impact":
            raise ValueError(f"Verified item {idx} date is not target date or continuing impact")
        validate_editorial_quality(item, idx)
        validate_india_weather_impact(item, idx)
        validate_thai_weather_impact(item, idx)
        dedupe_key = item.get("dedupe_key") or re.sub(r"\s+", "", item["news"][:100])
        if dedupe_key in seen:
            raise ValueError(f"Duplicate verified news: {dedupe_key}")
        seen.add(dedupe_key)
        row = dict(item)
        row["_order"] = idx
        normalized.append(row)
    return sorted(normalized, key=lambda x: (GROUP_ORDER.get(x["country_group"], 3), -int(x.get("importance", 0)), x["_order"]))


def copy_row_style(source_ws, source_row: int, target_ws, target_row: int) -> None:
    for col in range(1, 4):
        source = source_ws.cell(source_row, col)
        target = target_ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.font = copy(source.font)


def write_excel(task_root: Path, date_text: str, items: list[dict]) -> Path:
    template = task_root / "templates" / "新闻格式.xlsx"
    if not template.exists():
        raise FileNotFoundError(f"Missing template: {template}")
    out = excel_path(task_root, date_text)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, out)

    wb = load_workbook(out)
    ws = wb.active
    if [ws.cell(1, c).value for c in range(1, 4)] != ["国家", "新闻", "影响"]:
        raise ValueError("Excel template headers must be 国家/新闻/影响")
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    template_wb = load_workbook(template)
    template_ws = template_wb.active
    source_row = 2 if template_ws.max_row >= 2 else 1
    for row, item in enumerate(items, start=2):
        copy_row_style(template_ws, source_row, ws, row)
        ws.cell(row, 1).value = item["country"]
        ws.cell(row, 2).value = item["news"]
        ws.cell(row, 3).value = item["impact"]
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = max(72, min(180, 24 + 0.55 * max(len(item["news"]), len(item["impact"]))))
    for col in ("B", "C"):
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, 55)
    wb.save(out)
    return out


def read_excel_rows(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for row in range(2, ws.max_row + 1):
        country = ws.cell(row, 1).value
        news = ws.cell(row, 2).value
        impact = ws.cell(row, 3).value
        if country or news or impact:
            rows.append({"row": row, "country": country, "news": news, "impact": impact})
    return rows


def split_impact(value: str) -> tuple[str, str]:
    for prefix in IMPACT_PREFIXES:
        if value.startswith(prefix):
            return prefix[:-1], value[len(prefix):]
    raise ValueError(f"Invalid impact value: {value}")


def _number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def _round(value, digits: int = 2):
    return None if value is None else round(float(value), digits)


def price_from_quintal(value) -> float | None:
    number = _number(value)
    return _round(number / 100, 4) if number is not None else None


def lakh_tonnes_to_wan_tonnes(value) -> float | None:
    number = _number(value)
    return _round(number * 10, 2) if number is not None else None


def million_tonnes_to_wan_tonnes(value) -> float | None:
    number = _number(value)
    return _round(number * 100, 2) if number is not None else None


def normalize_price_metric(metric: dict | None, metric_type: str) -> dict:
    metric = dict(metric or {})
    status = metric.get("status") or ("ok" if metric.get("priceInrPerQuintal") or metric.get("rangeInrPerQuintal") else "pending")
    result = {
        "metricType": metric_type,
        "status": status,
        "statusText": metric.get("statusText") or ("数据待更新" if status != "ok" else ""),
        "dataDate": metric.get("dataDate") or metric.get("priceDate"),
        "priceDate": metric.get("priceDate") or metric.get("dataDate"),
        "grade": metric.get("grade"),
        "market": metric.get("market"),
        "quoteType": metric.get("quoteType"),
        "unit": metric.get("unit"),
        "rawUnit": metric.get("rawUnit"),
        "displayRange": metric.get("displayRange"),
        "rawRange": metric.get("rawRange"),
        "low": metric.get("low"),
        "high": metric.get("high"),
        "midpoint": metric.get("midpoint"),
        "priceBasis": metric.get("priceBasis"),
        "citiesUsed": metric.get("citiesUsed") or [],
        "cityCount": metric.get("cityCount"),
        "cityPrices": metric.get("cityPrices") or {},
        "rawCityPrices": metric.get("rawCityPrices") or {},
        "includesGst": metric.get("includesGst"),
        "originalUnit": metric.get("originalUnit") or "₹/quintal",
        "sourceName": metric.get("sourceName"),
        "sourceUrl": metric.get("sourceUrl"),
        "previousSourceUrl": metric.get("previousSourceUrl"),
        "yoySourceUrl": metric.get("yoySourceUrl"),
        "yoyComparisonDate": metric.get("yoyComparisonDate"),
        "yoyExactDateMatch": metric.get("yoyExactDateMatch"),
        "dailyMarketUpdateUrl": metric.get("dailyMarketUpdateUrl"),
        "publishedDate": metric.get("publishedDate"),
        "fetchedAt": metric.get("fetchedAt") or beijing_now().isoformat(timespec="seconds"),
        "note": metric.get("note"),
    }
    price_q = _number(metric.get("priceInrPerQuintal"))
    price_kg = _number(metric.get("priceInrPerKg"))
    if price_q is None and price_kg is not None:
        price_q = price_kg * 100
    if price_kg is None and price_q is not None:
        price_kg = price_from_quintal(price_q)
    result["priceInrPerQuintal"] = _round(price_q, 2)
    result["priceInrPerKg"] = _round(price_kg, 4)
    result["retailPriceInrPerKg"] = _round(_number(metric.get("retailPriceInrPerKg")), 2)

    range_q = metric.get("rangeInrPerQuintal") or {}
    low_q = _number(range_q.get("low") if isinstance(range_q, dict) else None)
    high_q = _number(range_q.get("high") if isinstance(range_q, dict) else None)
    if low_q is not None or high_q is not None:
        result["rangeInrPerQuintal"] = {"low": _round(low_q, 2), "high": _round(high_q, 2)}
        result["rangeInrPerKg"] = {"low": price_from_quintal(low_q), "high": price_from_quintal(high_q)}
    previous_range_q = metric.get("previousRangeInrPerQuintal") or {}
    if isinstance(previous_range_q, dict) and (previous_range_q.get("low") is not None or previous_range_q.get("high") is not None):
        prev_low = _number(previous_range_q.get("low"))
        prev_high = _number(previous_range_q.get("high"))
        result["previousRangeInrPerQuintal"] = {"low": _round(prev_low, 2), "high": _round(prev_high, 2)}
        result["previousRangeInrPerKg"] = {"low": price_from_quintal(prev_low), "high": price_from_quintal(prev_high)}

    previous_q = _number(metric.get("previousInrPerQuintal"))
    previous_kg = _number(metric.get("previousInrPerKg"))
    if previous_q is None and previous_kg is not None:
        previous_q = previous_kg * 100
    if previous_kg is None and previous_q is not None:
        previous_kg = price_from_quintal(previous_q)
    if previous_q is not None:
        result["previousInrPerQuintal"] = _round(previous_q, 2)
    result["previousInrPerKg"] = _round(previous_kg, 4)
    change_q = _number(metric.get("changeInrPerQuintal"))
    change_kg = _number(metric.get("changeInrPerKg"))
    if change_q is None and change_kg is not None:
        change_q = change_kg * 100
    if change_kg is None and change_q is not None:
        change_kg = price_from_quintal(change_q)
    if change_q is None and price_q is not None and previous_q is not None:
        change_q = price_q - previous_q
        change_kg = price_from_quintal(change_q)
    if change_kg is None and price_kg is not None and previous_kg is not None:
        change_kg = price_kg - previous_kg
    result["changeInrPerQuintal"] = _round(change_q, 2)
    result["changeInrPerKg"] = _round(change_kg, 4)
    change_pct = _number(metric.get("changePct"))
    if change_pct is None and change_q is not None and previous_q:
        change_pct = change_q / previous_q * 100
    if change_pct is None and change_kg is not None and previous_kg:
        change_pct = change_kg / previous_kg * 100
    result["changePct"] = _round(change_pct, 2)
    result["direction"] = metric.get("direction") or ("up" if change_q and change_q > 0 else "down" if change_q and change_q < 0 else "flat" if change_q == 0 else "unknown")
    result["previousDataDate"] = metric.get("previousDataDate")
    result["previousYearDate"] = metric.get("previousYearDate")
    previous_year_q = _number(metric.get("previousYearInrPerQuintal"))
    previous_year_kg = _number(metric.get("previousYearInrPerKg"))
    if previous_year_q is None and previous_year_kg is not None:
        previous_year_q = previous_year_kg * 100
    if previous_year_kg is None and previous_year_q is not None:
        previous_year_kg = price_from_quintal(previous_year_q)
    result["previousYearInrPerQuintal"] = _round(previous_year_q, 2)
    result["previousYearInrPerKg"] = _round(previous_year_kg, 4)
    yoy_q = _number(metric.get("yearOnYearChangeInrPerQuintal"))
    yoy_kg = _number(metric.get("yearOnYearChangeInrPerKg"))
    if yoy_q is None and yoy_kg is not None:
        yoy_q = yoy_kg * 100
    if yoy_kg is None and yoy_q is not None:
        yoy_kg = price_from_quintal(yoy_q)
    result["yearOnYearChangeInrPerQuintal"] = _round(yoy_q, 2)
    result["yearOnYearChangeInrPerKg"] = _round(yoy_kg, 4)
    result["yearOnYearChangePct"] = _round(_number(metric.get("yearOnYearChangePct")), 2)
    midpoint = _number(metric.get("midpointInrPerQuintal"))
    if midpoint is not None:
        result["midpointInrPerQuintal"] = _round(midpoint, 2)
        result["midpointInrPerKg"] = price_from_quintal(midpoint)
    if metric.get("gstStatus"):
        result["gstStatus"] = metric.get("gstStatus")
    return result


def normalize_stock_metric(metric: dict | None) -> dict:
    metric = dict(metric or {})
    status = metric.get("status") or ("ok" if any(metric.get(k) is not None for k in ("stockWanTonnes", "stockLakhTonnes", "stockMillionTonnes")) else "pending")
    stock_wan = _number(metric.get("stockWanTonnes"))
    if stock_wan is None and metric.get("stockLakhTonnes") is not None:
        stock_wan = lakh_tonnes_to_wan_tonnes(metric.get("stockLakhTonnes"))
    if stock_wan is None and metric.get("stockMillionTonnes") is not None:
        stock_wan = million_tonnes_to_wan_tonnes(metric.get("stockMillionTonnes"))
    previous_wan = _number(metric.get("previousForecastWanTonnes"))
    if previous_wan is None and metric.get("previousForecastLakhTonnes") is not None:
        previous_wan = lakh_tonnes_to_wan_tonnes(metric.get("previousForecastLakhTonnes"))
    yoy_wan = _number(metric.get("yoyChangeWanTonnes"))
    if yoy_wan is None and metric.get("yoyChangeLakhTonnes") is not None:
        yoy_wan = lakh_tonnes_to_wan_tonnes(metric.get("yoyChangeLakhTonnes"))
    revision_wan = _number(metric.get("revisionWanTonnes"))
    if revision_wan is None and stock_wan is not None and previous_wan is not None:
        revision_wan = stock_wan - previous_wan
    return {
        "metricType": "carryoverStock",
        "status": status,
        "statusText": metric.get("statusText") or ("数据待更新" if status != "ok" else ""),
        "dataDate": metric.get("dataDate"),
        "season": metric.get("season"),
        "stockWanTonnes": _round(stock_wan, 2),
        "stockLakhTonnes": _round(stock_wan / 10, 2) if stock_wan is not None else _round(_number(metric.get("stockLakhTonnes")), 2),
        "stockMillionTonnes": _round(stock_wan / 100, 2) if stock_wan is not None else _round(_number(metric.get("stockMillionTonnes")), 2),
        "previousForecastWanTonnes": _round(previous_wan, 2),
        "revisionWanTonnes": _round(revision_wan, 2),
        "yoyChangeWanTonnes": _round(yoy_wan, 2),
        "stockUseRatio": _round(_number(metric.get("stockUseRatio")), 2),
        "consumptionMonths": _round(_number(metric.get("consumptionMonths")), 2),
        "reason": metric.get("reason"),
        "sourceName": metric.get("sourceName"),
        "sourceTier": metric.get("sourceTier"),
        "organization": metric.get("organization") or metric.get("sourceName"),
        "sourceUrl": metric.get("sourceUrl"),
        "publishedDate": metric.get("publishedDate"),
        "previousSeasonWanTonnes": _round(_number(metric.get("previousSeasonWanTonnes")), 2),
        "yearOnYearChangePercent": _round(_number(metric.get("yearOnYearChangePercent")), 2),
        "forecastRevisionPercent": _round(_number(metric.get("forecastRevisionPercent")), 2),
        "fetchedAt": metric.get("fetchedAt") or beijing_now().isoformat(timespec="seconds"),
        "note": metric.get("note"),
    }


def pending_metric(metric_type: str) -> dict:
    base = {
        "metricType": metric_type,
        "status": "pending",
        "statusText": "数据待更新",
        "fetchedAt": beijing_now().isoformat(timespec="seconds"),
        "note": "未获取到已完成日期、口径和来源核验的可靠数据；不编造价格或库存。",
    }
    if metric_type == "carryoverStock":
        base.update({"dataDate": None, "stockWanTonnes": None, "statusText": "等待权威来源更新"})
    else:
        base.update({"dataDate": None, "priceDate": None, "priceInrPerQuintal": None, "priceInrPerKg": None})
    return base


def normalize_market_forecast(item: dict, main_stock_wan: float | None = None) -> dict:
    stock_wan = _number(item.get("closing_stock_ten_thousand_tonnes") or item.get("stockWanTonnes"))
    diff = stock_wan - main_stock_wan if stock_wan is not None and main_stock_wan is not None else None
    return {
        "sourceTier": "market_forecast_comparison_only",
        "organization": item.get("forecast_organization") or item.get("source_name"),
        "season": item.get("season"),
        "forecastDate": item.get("forecast_date") or item.get("data_date"),
        "stockWanTonnes": _round(stock_wan, 2),
        "differenceToMainWanTonnes": _round(diff, 2),
        "sourceUrl": item.get("source_url"),
        "fetchedAt": item.get("fetched_at"),
        "note": item.get("note"),
    }


def latest_previous_india_metrics(date_text: str) -> dict | None:
    reports_root = PUBLIC_DATA_ROOT / "reports"
    if not reports_root.exists():
        return None
    candidates = []
    for path in reports_root.rglob("*.json"):
        try:
            with path.open("r", encoding="utf-8") as f:
                payload = json.load(f)
        except Exception:
            continue
        if payload.get("newsDate", "") >= date_text:
            continue
        metrics = payload.get("indiaMetrics")
        if metrics:
            candidates.append((payload.get("newsDate"), metrics))
    if not candidates:
        return None
    return sorted(candidates, key=lambda pair: pair[0], reverse=True)[0][1]


def latest_india_metrics_snapshot() -> dict | None:
    path = PUBLIC_DATA_ROOT / "india_metrics" / "latest.json"
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def india_metrics_from_snapshot(snapshot: dict | None) -> dict:
    if not snapshot:
        return {}
    wholesale = snapshot.get("domesticWholesalePrice") or snapshot.get("domesticSugarPrice") or {}
    retail = snapshot.get("domesticRetailPrice") or {}
    domestic = snapshot.get("domesticSugarPrice") or wholesale
    up_ex = snapshot.get("upExMillPrice") or {}
    stock = snapshot.get("carryoverStock") or {}
    result = {
        "dataDate": snapshot.get("targetDate"),
        "sourceStatus": "dynamic_fetch",
        "fetchLog": snapshot.get("fetchLog", []),
    }
    if wholesale:
        result["domesticWholesalePrice"] = {
            "status": wholesale.get("status", "ok"),
            "priceDate": wholesale.get("data_date"),
            "grade": wholesale.get("grade") or "M-30；Hyderabad为S-30",
            "market": "ChiniMandi city sample",
            "quoteType": "wholesale sample average",
            "unit": wholesale.get("unit"),
            "priceBasis": wholesale.get("price_basis"),
            "citiesUsed": wholesale.get("cities_used") or [],
            "cityCount": wholesale.get("city_count"),
            "cityPrices": wholesale.get("city_prices") or {},
            "rawCityPrices": wholesale.get("raw_city_prices") or {},
            "includesGst": wholesale.get("includes_gst"),
            "priceInrPerQuintal": wholesale.get("price_inr_per_quintal") or wholesale.get("wholesale_price_inr_per_quintal"),
            "priceInrPerKg": wholesale.get("price_inr_per_kg") or wholesale.get("wholesale_price_inr_per_kg"),
            "previousDataDate": wholesale.get("previous_data_date"),
            "previousInrPerQuintal": wholesale.get("previous_value"),
            "changeInrPerQuintal": wholesale.get("change_value"),
            "changePct": wholesale.get("change_percent"),
            "previousYearDate": wholesale.get("previous_year_date"),
            "previousYearInrPerQuintal": wholesale.get("previous_year_value"),
            "yearOnYearChangeInrPerQuintal": wholesale.get("year_on_year_change"),
            "yearOnYearChangePct": wholesale.get("year_on_year_change_percent"),
            "originalUnit": wholesale.get("unit") or "₹/quintal and ₹/kg",
            "sourceName": wholesale.get("source_name"),
            "sourceUrl": wholesale.get("source_url"),
            "dailyMarketUpdateUrl": wholesale.get("daily_market_update_url"),
            "publishedDate": wholesale.get("data_date"),
            "fetchedAt": wholesale.get("fetched_at"),
        }
        result["domesticSugarPrice"] = result["domesticWholesalePrice"]
    if retail:
        result["domesticRetailPrice"] = {
            "status": retail.get("status", "ok"),
            "priceDate": retail.get("data_date"),
            "grade": retail.get("grade") or "M-30；Hyderabad为S-30",
            "market": "ChiniMandi city sample",
            "quoteType": "retail sample average",
            "unit": retail.get("unit"),
            "priceBasis": retail.get("price_basis"),
            "citiesUsed": retail.get("cities_used") or [],
            "cityCount": retail.get("city_count"),
            "cityPrices": retail.get("city_prices") or {},
            "rawCityPrices": retail.get("raw_city_prices") or {},
            "includesGst": retail.get("includes_gst"),
            "priceInrPerKg": retail.get("price_inr_per_kg"),
            "previousDataDate": retail.get("previous_data_date"),
            "previousInrPerKg": retail.get("previous_value"),
            "changeInrPerKg": retail.get("change_value"),
            "changePct": retail.get("change_percent"),
            "previousYearDate": retail.get("previous_year_date"),
            "previousYearInrPerKg": retail.get("previous_year_value"),
            "yearOnYearChangeInrPerKg": retail.get("year_on_year_change"),
            "yearOnYearChangePct": retail.get("year_on_year_change_percent"),
            "originalUnit": retail.get("unit") or "₹/kg",
            "sourceName": retail.get("source_name"),
            "sourceUrl": retail.get("source_url"),
            "dailyMarketUpdateUrl": retail.get("daily_market_update_url"),
            "publishedDate": retail.get("data_date"),
            "fetchedAt": retail.get("fetched_at"),
        }
    if up_ex:
        result["upExMillPrice"] = {
            "status": up_ex.get("status", "ok"),
            "priceDate": up_ex.get("data_date"),
            "displayRange": up_ex.get("display_range"),
            "rawRange": up_ex.get("raw_range"),
            "low": up_ex.get("low"),
            "high": up_ex.get("high"),
            "midpoint": up_ex.get("midpoint"),
            "currency": up_ex.get("currency"),
            "unit": up_ex.get("unit"),
            "rawUnit": up_ex.get("raw_unit"),
            "grade": up_ex.get("grade") or "M/30",
            "market": "Uttar Pradesh",
            "quoteType": "ex-mill",
            "includesGst": up_ex.get("includes_gst"),
            "rangeInrPerQuintal": {"low": up_ex.get("up_ex_mill_min_inr_per_quintal"), "high": up_ex.get("up_ex_mill_max_inr_per_quintal")},
            "midpointInrPerQuintal": up_ex.get("up_ex_mill_mid_inr_per_quintal"),
            "previousRangeInrPerQuintal": {"low": up_ex.get("previous_min"), "high": up_ex.get("previous_max")},
            "previousDataDate": up_ex.get("previous_data_date"),
            "previousSourceUrl": up_ex.get("previous_source_url"),
            "previousLow": up_ex.get("previous_low"),
            "previousHigh": up_ex.get("previous_high"),
            "previousMidpoint": up_ex.get("previous_midpoint"),
            "previousInrPerQuintal": up_ex.get("previous_mid"),
            "changeInrPerQuintal": up_ex.get("change_value"),
            "dailyChangeAbsolute": up_ex.get("daily_change_absolute"),
            "dailyChangePercent": up_ex.get("daily_change_percent"),
            "changePct": up_ex.get("change_percent"),
            "previousYearDate": up_ex.get("previous_year_date"),
            "yoyComparisonDate": up_ex.get("yoy_comparison_date"),
            "yoySourceUrl": up_ex.get("yoy_source_url"),
            "yoyExactDateMatch": up_ex.get("yoy_exact_date_match"),
            "yoyLow": up_ex.get("yoy_low"),
            "yoyHigh": up_ex.get("yoy_high"),
            "yoyMidpoint": up_ex.get("yoy_midpoint"),
            "previousYearInrPerQuintal": up_ex.get("previous_year_mid"),
            "yearOnYearChangeInrPerQuintal": up_ex.get("year_on_year_change"),
            "yoyChangeAbsolute": up_ex.get("yoy_change_absolute"),
            "yoyChangePercent": up_ex.get("yoy_change_percent"),
            "yearOnYearChangePct": up_ex.get("year_on_year_change_percent"),
            "direction": up_ex.get("change_direction"),
            "gstStatus": up_ex.get("gst_status"),
            "originalUnit": "₹/quintal",
            "sourceName": up_ex.get("source_name"),
            "sourceUrl": up_ex.get("source_url"),
            "publishedDate": up_ex.get("data_date"),
            "fetchedAt": up_ex.get("fetched_at"),
        }
    if stock:
        result["carryoverStock"] = {
            "status": stock.get("status", "ok"),
            "dataDate": stock.get("forecast_date") or stock.get("data_date"),
            "season": stock.get("season"),
            "stockWanTonnes": stock.get("closing_stock_ten_thousand_tonnes"),
            "stockLakhTonnes": stock.get("closing_stock_lakh_tonnes"),
            "stockMillionTonnes": stock.get("closing_stock_million_tonnes"),
            "previousForecastWanTonnes": stock.get("previous_forecast_value"),
            "revisionWanTonnes": stock.get("forecast_revision"),
            "forecastRevisionPercent": stock.get("forecast_revision_percent"),
            "yoyChangeWanTonnes": stock.get("year_on_year_change"),
            "yearOnYearChangePercent": stock.get("year_on_year_change_percent"),
            "previousSeasonWanTonnes": stock.get("previous_season_value"),
            "sourceTier": "authoritative_main",
            "organization": stock.get("forecast_organization") or stock.get("source_name"),
            "sourceName": stock.get("forecast_organization") or stock.get("source_name"),
            "sourceUrl": stock.get("source_url"),
            "publishedDate": stock.get("forecast_date"),
            "fetchedAt": stock.get("fetched_at"),
        }
    forecasts = snapshot.get("carryoverStockForecasts") or []
    if forecasts:
        main_wan = stock.get("closing_stock_ten_thousand_tonnes") if stock else None
        result["carryoverStockForecasts"] = [normalize_market_forecast(item, _number(main_wan)) for item in forecasts]
    if snapshot.get("authorizedCarryoverStockAlternatives"):
        result["authorizedCarryoverStockAlternatives"] = snapshot.get("authorizedCarryoverStockAlternatives")
    return result


def normalize_india_metrics(data: dict, date_text: str) -> dict:
    raw = data.get("indiaMetrics") or data.get("india_metrics")
    snapshot_raw = india_metrics_from_snapshot(latest_india_metrics_snapshot())
    previous = latest_previous_india_metrics(date_text) or {}
    source = raw if isinstance(raw, dict) else snapshot_raw if snapshot_raw else previous if isinstance(previous, dict) else {}
    wholesale = source.get("domesticWholesalePrice") if isinstance(source, dict) else None
    retail = source.get("domesticRetailPrice") if isinstance(source, dict) else None
    domestic = source.get("domesticSugarPrice") if isinstance(source, dict) else None
    if wholesale is None:
        wholesale = domestic
    up_ex_mill = source.get("upExMillPrice") if isinstance(source, dict) else None
    stock = source.get("carryoverStock") if isinstance(source, dict) else None
    payload = {
        "title": "印度糖价",
        "dataDate": (source.get("dataDate") if isinstance(source, dict) else None) or date_text,
        "updatedAt": beijing_now().isoformat(timespec="seconds"),
        "sourceStatus": "verified" if raw else "dynamic_fetch" if snapshot_raw else "carried_forward" if previous else "pending",
        "domesticWholesalePrice": normalize_price_metric(wholesale, "domesticWholesalePrice") if wholesale else pending_metric("domesticWholesalePrice"),
        "domesticRetailPrice": normalize_price_metric(retail, "domesticRetailPrice") if retail else pending_metric("domesticRetailPrice"),
        "domesticSugarPrice": normalize_price_metric(wholesale, "domesticSugarPrice") if wholesale else pending_metric("domesticSugarPrice"),
        "upExMillPrice": normalize_price_metric(up_ex_mill, "upExMillPrice") if up_ex_mill else pending_metric("upExMillPrice"),
        "carryoverStock": normalize_stock_metric(stock) if stock else pending_metric("carryoverStock"),
    }
    if not raw and snapshot_raw:
        payload["note"] = source.get("note") or "印度糖价指标来自本次动态抓取；库存候选仅保留在后台日志，不用于价格看板展示。"
        if source.get("fetchLog"):
            payload["fetchLog"] = source.get("fetchLog")
        if source.get("carryoverStockForecasts"):
            payload["carryoverStockForecasts"] = source.get("carryoverStockForecasts")
    elif not raw and previous:
        payload["note"] = "本期未发现新的已核验印度糖价数据，沿用最近一期有效数据并保留原始数据日期。"
    elif not raw:
        payload["note"] = "本期未获取到已完成日期、口径和来源核验的印度糖价数据。"
    else:
        payload["note"] = source.get("note")
    return payload


def latest_brazil_metrics_snapshot() -> dict | None:
    path = PUBLIC_DATA_ROOT / "brazil_metrics" / "latest.json"
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def first_log_url(metric: dict | None) -> tuple[str | None, str | None]:
    if not isinstance(metric, dict):
        return None, None
    for entry in metric.get("fetchLog") or []:
        if entry.get("url"):
            return entry.get("source"), entry.get("url")
    return metric.get("source_name") or metric.get("dataset_name"), metric.get("source_url")


def normalize_brazil_metric(metric: dict | None, metric_type: str, display_date: str) -> dict:
    metric = metric if isinstance(metric, dict) else {}
    status = metric.get("status") or "pending"
    source_name, source_url = first_log_url(metric)
    source_data_date = (
        metric.get("source_data_date")
        or metric.get("data_date")
        or metric.get("reference_date")
        or metric.get("reference_period")
        or metric.get("published_at")
    )
    base = {
        "metricType": metric_type,
        "status": status if status in {"ok", "pending", "stale"} else "pending",
        "statusText": "",
        "dataDate": source_data_date,
        "sourceDataDate": source_data_date,
        "refreshDate": display_date,
        "fetchedAt": metric.get("fetched_at") or beijing_now().isoformat(timespec="seconds"),
        "sourceName": metric.get("source_name") or metric.get("dataset_name") or source_name,
        "sourceUrl": metric.get("source_url") or source_url,
        "datasetName": metric.get("dataset_name"),
        "previousYearValue": metric.get("previous_year_value"),
        "previousYearSourceUrl": metric.get("previous_year_source_url"),
        "yoyTargetDate": metric.get("yoy_target_date"),
        "yoyComparisonDate": metric.get("yoy_comparison_date"),
        "yoyExactDateMatch": metric.get("yoy_exact_date_match"),
        "yoyDateGapDays": metric.get("yoy_date_gap_days"),
        "yearOnYearChange": metric.get("year_on_year_change"),
        "yearOnYearChangePercent": metric.get("year_on_year_change_percent"),
        "yoyStatus": metric.get("yoy_status") or ("ok" if metric.get("year_on_year_change_percent") is not None else "insufficient"),
        "note": metric.get("note"),
    }
    if metric_type == "sugarPremium":
        value = metric.get("premium_discount_cents_per_lb")
        base.update({
            "product": metric.get("product"),
            "port": metric.get("port"),
            "pricingBasis": metric.get("pricing_basis"),
            "futuresContract": metric.get("futures_contract"),
            "importPremiumDiscountCentsPerLb": metric.get("import_premium_discount_cents_per_lb"),
            "premiumDiscountCentsPerLb": value,
            "premiumLabel": "升水" if isinstance(value, (int, float)) and value >= 0 else "贴水" if isinstance(value, (int, float)) else None,
            "unit": "美分/磅",
            "previousDataDate": metric.get("previous_data_date"),
            "previousValue": metric.get("previous_value"),
            "dailyChange": metric.get("daily_change"),
            "dailyChangePercent": metric.get("daily_change_percent"),
            "previousYearDate": metric.get("previous_year_date"),
            "articleId": metric.get("article_id"),
            "articleTitle": metric.get("article_title"),
            "articlePublishedAt": metric.get("article_published_at"),
            "imageUrl": metric.get("image_url"),
            "ocrBackend": metric.get("ocr_backend"),
        })
    elif metric_type == "sugarStock":
        base.update({
            "product": metric.get("product") or "食糖",
            "stockValue": metric.get("sugar_stock_value"),
            "stockUnit": metric.get("stock_unit"),
            "datasetName": metric.get("dataset_name"),
            "season": metric.get("season"),
            "referenceDate": metric.get("reference_date") or metric.get("reference_period"),
            "referenceDateRaw": metric.get("reference_date_raw"),
            "referenceDateSource": metric.get("reference_date_source"),
            "stockTotalTonnes": metric.get("stock_total_tonnes"),
            "stockTotalTenThousandTonnes": metric.get("stock_total_ten_thousand_tonnes"),
            "previousPeriodDate": metric.get("previous_period_date"),
            "previousPeriodStock": metric.get("previous_period_stock"),
            "halfMonthChange": metric.get("half_month_change"),
            "halfMonthChangePercent": metric.get("half_month_change_percent"),
            "previousYearDate": metric.get("previous_year_date"),
            "previousYearStock": metric.get("previous_year_stock"),
            "documentNumber": metric.get("document_number"),
            "documentTitle": metric.get("document_title"),
            "fileHash": metric.get("file_hash"),
        })
    else:
        base.update({
            "ethanolType": metric.get("ethanol_type"),
            "stockType": metric.get("stock_type"),
            "hydrousEthanolStock": metric.get("hydrous_ethanol_stock"),
            "anhydrousEthanolStock": metric.get("anhydrous_ethanol_stock"),
            "totalEthanolStock": metric.get("total_ethanol_stock"),
            "stockCubicMetres": metric.get("stock_cubic_metres"),
            "stockTenThousandCubicMetres": metric.get("stock_ten_thousand_cubic_metres"),
            "stockUnit": metric.get("stock_unit"),
            "datasetName": metric.get("dataset_name"),
            "season": metric.get("season"),
            "referenceDate": metric.get("reference_date") or metric.get("reference_period"),
            "reportUpdatedAt": metric.get("report_updated_at"),
            "previousPeriodDate": metric.get("previous_period_date"),
            "previousPeriodStock": metric.get("previous_period_stock"),
            "halfMonthChange": metric.get("half_month_change"),
            "halfMonthChangePercent": metric.get("half_month_change_percent"),
            "previousYearDate": metric.get("previous_year_date"),
            "previousYearStock": metric.get("previous_year_stock"),
            "yearOnYearChange": metric.get("year_on_year_change"),
            "yearOnYearChangePercent": metric.get("year_on_year_change_percent"),
            "yoyStatus": metric.get("yoy_status"),
            "sourcePageUrl": metric.get("source_page_url"),
            "reportUrl": metric.get("report_url"),
            "sourceFileName": metric.get("source_file_name"),
            "fileHash": metric.get("file_hash"),
        })
    return base


def normalize_brazil_metrics(date_text: str) -> dict:
    snapshot = latest_brazil_metrics_snapshot() or {}
    return {
        "title": "巴西糖价与库存",
        "dataDate": date_text,
        "snapshotTargetDate": snapshot.get("targetDate"),
        "updatedAt": snapshot.get("updatedAt") or beijing_now().isoformat(timespec="seconds"),
        "sourceStatus": "dynamic_fetch" if snapshot else "pending",
        "sugarPremium": normalize_brazil_metric(snapshot.get("sugarPremium"), "sugarPremium", date_text),
        "sugarStock": normalize_brazil_metric(snapshot.get("sugarStock"), "sugarStock", date_text),
        "ethanolStock": normalize_brazil_metric(snapshot.get("ethanolStock"), "ethanolStock", date_text),
        "note": None,
        "fetchLog": snapshot.get("fetchLog", []),
    }


def strip_public_fetch_logs(value):
    if isinstance(value, dict):
        return {k: strip_public_fetch_logs(v) for k, v in value.items() if k != "fetchLog"}
    if isinstance(value, list):
        return [strip_public_fetch_logs(item) for item in value]
    return value


def build_dashboard_payload(date_text: str, items: list[dict], excel_file: Path, verified_data: dict | None = None) -> dict:
    grouped: dict[str, list[dict]] = defaultdict(list)
    country_order: list[tuple[int, int, str]] = []
    for item in items:
        impact_type, impact_text = split_impact(item["impact"])
        grouped[item["country"]].append({
            "news": re.sub(r"\s*来源：.*$", "", item["news"]).strip(),
            "impactType": impact_type,
            "impact": impact_text.strip(),
            "impactLabel": impact_marker_from_impact(item["impact"]),
            "sourceName": item["source_name"],
            "sourceUrl": item["source_url"],
            "publishedDateLocal": item["published_date_local"],
            "eventDate": item.get("event_date"),
        })
        country_order.append((GROUP_ORDER.get(item["country_group"], 3), -int(item.get("importance", 0)), item["country"]))

    countries = []
    seen = set()
    for _, _, country in sorted(country_order, key=lambda pair: (pair[0], pair[1], country_order.index(pair))):
        if country in seen:
            continue
        seen.add(country)
        if grouped[country]:
            countries.append({"country": country, "items": grouped[country]})

    return {
        "newsDate": date_text,
        "updatedAt": beijing_now().isoformat(timespec="seconds"),
        "timezone": "Asia/Shanghai",
        "excelFile": project_display_path(excel_file),
        "globalHighlights": global_summary_for_report(items, verified_data),
        "brazilMetrics": strip_public_fetch_logs(normalize_brazil_metrics(date_text)),
        "indiaMetrics": strip_public_fetch_logs(normalize_india_metrics(verified_data or {}, date_text)),
        "countries": countries,
    }


def preserve_existing_dashboard_metrics(date_text: str, payload: dict) -> dict:
    report_path = public_report_path(date_text)
    if not report_path.exists():
        return payload
    with report_path.open("r", encoding="utf-8") as f:
        existing = json.load(f)
    preserved = dict(payload)
    for field in ("brazilMetrics", "indiaMetrics"):
        if isinstance(existing.get(field), dict):
            preserved[field] = strip_public_fetch_logs(existing[field])
    return preserved


def write_dashboard_data(date_text: str, payload: dict) -> tuple[Path, Path]:
    report_path = public_report_path(date_text)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    reports = []
    reports_root = PUBLIC_DATA_ROOT / "reports"
    if reports_root.exists():
        for path in reports_root.rglob("*.json"):
            with path.open("r", encoding="utf-8") as f:
                entry = json.load(f)
            reports.append({
                "newsDate": entry["newsDate"],
                "updatedAt": entry.get("updatedAt"),
                "path": "/" + str(path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                "count": sum(len(c.get("items", [])) for c in entry.get("countries", [])),
            })
    reports.sort(key=lambda x: x["newsDate"], reverse=True)
    index = {
        "latestNewsDate": reports[0]["newsDate"] if reports else None,
        "updatedAt": beijing_now().isoformat(timespec="seconds"),
        "reports": reports,
    }
    index_path = public_index_path()
    index_path.parent.mkdir(parents=True, exist_ok=True)
    with index_path.open("w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    return report_path, index_path


def validate_all(date_text: str, items: list[dict], excel_file: Path, report_path: Path, index_path: Path) -> dict:
    excel_rows = read_excel_rows(excel_file)
    expected_pairs = {(item["country"], item["news"]) for item in items}
    actual_pairs = {(row["country"], row["news"]) for row in excel_rows}
    if expected_pairs != actual_pairs:
        missing = expected_pairs - actual_pairs
        extra = actual_pairs - expected_pairs
        raise ValueError(f"Excel mismatch: missing={missing}; extra={extra}")

    with report_path.open("r", encoding="utf-8") as f:
        report = json.load(f)
    with index_path.open("r", encoding="utf-8") as f:
        index = json.load(f)
    dashboard_count = sum(len(c.get("items", [])) for c in report.get("countries", []))
    expected_china = sum(1 for item in items if item["country_group"] == "中国" or item["country"] == "中国")
    actual_china = sum(
        len(c.get("items", []))
        for c in report.get("countries", [])
        if c.get("country") == "中国"
    )
    if dashboard_count != len(items):
        raise ValueError(f"Dashboard count mismatch: {dashboard_count} != {len(items)}")
    if report.get("newsDate") != date_text:
        raise ValueError("Dashboard report date mismatch")
    if index.get("latestNewsDate") < date_text:
        raise ValueError("Dashboard index latest date is older than target date")
    if any(not c.get("items") for c in report.get("countries", [])):
        raise ValueError("Dashboard contains empty country section")
    if any(c.get("country") == "其他" for c in report.get("countries", [])):
        raise ValueError("Dashboard must not collapse other countries into a single 其他 section")
    if actual_china != expected_china:
        raise ValueError(f"China dashboard count mismatch: {actual_china} != {expected_china}")
    global_summary = validate_global_summary(str(report.get("globalHighlights", "")))
    brazil_metrics = report.get("brazilMetrics")
    if not isinstance(brazil_metrics, dict):
        raise ValueError("Dashboard missing brazilMetrics")
    if brazil_metrics.get("dataDate") != date_text:
        raise ValueError("Brazil dashboard date must match Sugar News date")
    for field in ("sugarPremium", "sugarStock", "ethanolStock"):
        metric = brazil_metrics.get(field)
        if not isinstance(metric, dict):
            raise ValueError(f"brazilMetrics missing {field}")
        if metric.get("status") not in {"ok", "pending", "stale"}:
            raise ValueError(f"brazilMetrics {field} has invalid status")
        if metric.get("refreshDate") != date_text:
            raise ValueError(f"brazilMetrics {field} refresh date must match Sugar News date")
        if metric.get("status") == "ok":
            if not metric.get("sourceDataDate"):
                raise ValueError(f"brazilMetrics {field} must preserve sourceDataDate")
            if metric.get("dataDate") != metric.get("sourceDataDate"):
                raise ValueError(f"brazilMetrics {field} card date must use sourceDataDate")
            if field == "sugarPremium" and metric.get("premiumDiscountCentsPerLb") is None:
                raise ValueError("sugarPremium ok status requires premiumDiscountCentsPerLb")
            if field == "sugarPremium" and "HiSugar" not in str(metric.get("datasetName")):
                raise ValueError("sugarPremium must use HiSugar import cost estimate")
            if field == "sugarStock" and metric.get("stockValue") is None:
                raise ValueError("sugarStock ok status requires stockValue")
            if field == "sugarStock" and "MAPA" not in str(metric.get("sourceName")):
                raise ValueError("sugarStock must use MAPA, not ANP")
            if field == "ethanolStock" and metric.get("totalEthanolStock") is None:
                raise ValueError("ethanolStock ok status requires totalEthanolStock")
            if field == "ethanolStock" and "MAPA" not in str(metric.get("sourceName")):
                raise ValueError("ethanolStock must use MAPA as the dashboard source")
            if field == "ethanolStock" and metric.get("stockType") != "physical":
                raise ValueError("ethanolStock must use physical stock")
    india_metrics = report.get("indiaMetrics")
    if not isinstance(india_metrics, dict):
        raise ValueError("Dashboard missing indiaMetrics")
    for field in ("domesticWholesalePrice", "domesticRetailPrice", "upExMillPrice"):
        metric = india_metrics.get(field)
        if not isinstance(metric, dict):
            raise ValueError(f"indiaMetrics missing {field}")
        if metric.get("status") not in {"ok", "pending", "stale"}:
            raise ValueError(f"indiaMetrics {field} has invalid status")
        if metric.get("status") == "ok":
            if metric.get("priceInrPerQuintal") is None and metric.get("priceInrPerKg") is None and not metric.get("rangeInrPerQuintal"):
                raise ValueError(f"{field} ok status requires price or range")
            if metric.get("previousDataDate") is None:
                raise ValueError(f"{field} ok status requires previousDataDate for daily change comparison")
            if metric.get("changePct") is None:
                raise ValueError(f"{field} ok status requires daily change percent")
            if field in {"domesticWholesalePrice", "domesticRetailPrice"}:
                expected_url = "https://www.chinimandi.com/wholesale-sugar-prices/" if field == "domesticWholesalePrice" else "https://www.chinimandi.com/retail-prices/"
                if metric.get("sourceName") != "ChiniMandi":
                    raise ValueError(f"{field} must use ChiniMandi")
                if metric.get("sourceUrl") != expected_url:
                    raise ValueError(f"{field} sourceUrl mismatch")
                if metric.get("includesGst") is not True:
                    raise ValueError(f"{field} must mark includesGst true")
                if not metric.get("citiesUsed") or not metric.get("cityCount"):
                    raise ValueError(f"{field} requires ChiniMandi city sample metadata")
            if field == "upExMillPrice":
                if metric.get("sourceName") != "ChiniMandi — Daily Sugar Market Update":
                    raise ValueError("upExMillPrice must use ChiniMandi Daily Sugar Market Update")
                if metric.get("market") != "Uttar Pradesh":
                    raise ValueError("upExMillPrice must not use destination spot prices")
                if metric.get("grade") != "M/30":
                    raise ValueError("upExMillPrice must use M/30")
                if metric.get("includesGst") is not False:
                    raise ValueError("upExMillPrice must be excluding GST")
                if not metric.get("sourceUrl") or "daily-sugar-market-update-by-vizzie" not in metric.get("sourceUrl"):
                    raise ValueError("upExMillPrice requires Daily Sugar Market Update sourceUrl")
                if not metric.get("previousSourceUrl") or not metric.get("yoySourceUrl"):
                    raise ValueError("upExMillPrice requires previous and yoy source links")

    group_positions = []
    for row in excel_rows:
        if row["country"] == "巴西":
            group_positions.append(0)
        elif row["country"] == "印度":
            group_positions.append(1)
        elif row["country"] == "泰国":
            group_positions.append(2)
        elif row["country"] == "中国":
            group_positions.append(4)
        else:
            group_positions.append(3)
    checks = {
        "verified_count": len(items),
        "excel_count": len(excel_rows),
        "dashboard_count": dashboard_count,
        "excel_matches_verified": True,
        "dashboard_matches_verified": True,
        "country_order_ok": group_positions == sorted(group_positions),
        "global_summary_sentences": len(split_cn_sentences(global_summary)),
        "no_empty_country_sections": True,
        "counts_by_country": dict(Counter(item["country"] for item in items)),
        "china_count": expected_china,
        "other_country_count": sum(1 for item in items if item["country_group"] == "其他国家"),
    }
    return checks


def china_monitoring_log(items: list[dict]) -> dict:
    china_items = [item for item in items if item.get("country_group") == "中国" or item.get("country") == "中国"]
    query_templates = [template for _language, template in COUNTRY_SEARCH_TEMPLATES.get("中国", ())]
    return {
        "status": "completed",
        "mode": "verified_items_checked_after_load",
        "retained_count": len(china_items),
        "query_template_count": len(query_templates),
        "query_templates": query_templates,
        "priority_sources": [
            "中国糖业协会",
            "农业农村部 / CASDE",
            "海关总署",
            "国家统计局",
            "国家发展改革委 / 商务部",
            "广西、云南等主产区农业、气象和糖业主管部门",
            "云糖网 / 沐甜科技",
            "郑州商品交易所",
            "权威期货公司、研究机构及糖厂公告",
        ],
        "column_required": False,
        "note": "China sugar monitoring completed; public China output is present" if china_items else "China sugar monitoring completed; no publishable China item found",
    }


def write_status(date_text: str, status: str, details: dict, error: str | None = None) -> None:
    path = public_status_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "latestNewsDate": date_text if status == "success" else None,
        "lastRunAt": beijing_now().isoformat(timespec="seconds"),
        "lastRunStatus": status,
        "timezone": "Asia/Shanghai",
        "details": details,
    }
    if error:
        payload["error"] = error[:1000]
    if path.exists() and status != "success":
        with path.open("r", encoding="utf-8") as f:
            old = json.load(f)
        payload["latestNewsDate"] = old.get("latestNewsDate")
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def write_task_log(task_root: Path, date_text: str, payload: dict) -> None:
    path = write_log_path(task_root, date_text)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def run_metric_refresh(script_name: str, date_text: str, latest_path: Path) -> dict:
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    try:
        result = subprocess.run(
            [sys.executable, str(PROJECT_ROOT / "scripts" / script_name), "--date", date_text],
            cwd=str(PROJECT_ROOT),
            env=env,
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=True,
            timeout=METRIC_REFRESH_TIMEOUT_SECONDS,
        )
        if result.returncode != 0:
            return {
                "status": "failed",
                "error": (result.stderr or result.stdout or "")[-1000:],
                "timeoutSeconds": METRIC_REFRESH_TIMEOUT_SECONDS,
            }
        updated_at = None
        if latest_path.exists():
            with latest_path.open("r", encoding="utf-8") as f:
                updated_at = json.load(f).get("updatedAt")
        return {"status": "success", "snapshotUpdatedAt": updated_at}
    except subprocess.TimeoutExpired as exc:
        return {
            "status": "timeout",
            "error": f"{script_name} exceeded {METRIC_REFRESH_TIMEOUT_SECONDS}s",
            "stdoutTail": (exc.stdout or "")[-500:] if isinstance(exc.stdout, str) else "",
            "stderrTail": (exc.stderr or "")[-500:] if isinstance(exc.stderr, str) else "",
        }
    except Exception as exc:
        return {"status": "failed", "error": str(exc)[:1000]}


def refresh_india_metrics(date_text: str) -> dict:
    return run_metric_refresh("india_sugar_metrics.py", date_text, PUBLIC_DATA_ROOT / "india_metrics" / "latest.json")


def refresh_brazil_metrics(date_text: str) -> dict:
    return run_metric_refresh("brazil_sugar_metrics.py", date_text, PUBLIC_DATA_ROOT / "brazil_metrics" / "latest.json")


def main() -> int:
    args = parse_args()
    date_text = target_date(args.date)
    task_root = task_root_from_args(args.task_root)
    ensure_task_dirs(task_root, date_text)

    if args.skip_if_success and success_exists(date_text):
        print(json.dumps({"status": "skipped", "reason": "already_success", "newsDate": date_text}, ensure_ascii=False))
        return 0

    try:
        editorial_skill = load_editorial_skill_metadata()
        print(f"[sugar-news] editorial skill loaded: {editorial_skill['path']} {editorial_skill['sha256'][:12]}", flush=True)
        if args.skip_metric_refresh:
            brazil_metrics_refresh = {"status": "skipped", "reason": "standalone workflow refresh; consuming latest snapshot"}
            india_metrics_refresh = {"status": "skipped", "reason": "standalone workflow refresh; consuming latest snapshot"}
            print(f"[sugar-news] consume workflow-refreshed metric snapshots for {date_text}", flush=True)
        else:
            print(f"[sugar-news] refresh Brazil metrics for {date_text}", flush=True)
            brazil_metrics_refresh = refresh_brazil_metrics(date_text)
            print(f"[sugar-news] Brazil metrics: {brazil_metrics_refresh.get('status')}", flush=True)
            print(f"[sugar-news] refresh India metrics for {date_text}", flush=True)
            india_metrics_refresh = refresh_india_metrics(date_text)
            print(f"[sugar-news] India metrics: {india_metrics_refresh.get('status')}", flush=True)
        print(f"[sugar-news] load verified/autogenerate news for {date_text}", flush=True)
        data = load_verified_or_fail(
            task_root,
            date_text,
            offline_only=args.offline_only,
            allow_rss_autogen=args.allow_rss_autogen,
            force_rss_autogen=args.force_rss_autogen,
        )
        print(f"[sugar-news] ensure required China sugar item for {date_text}", flush=True)
        data, china_monitoring_check = ensure_china_news_item(data, date_text)
        print(f"[sugar-news] ensure Thailand cane-area weather item for {date_text}", flush=True)
        data, thai_weather_log = ensure_thai_weather_item(data, date_text)
        print(f"[sugar-news] normalize/write outputs for {date_text}", flush=True)
        items = normalize_items(data)
        data["items"] = [{key: value for key, value in item.items() if key != "_order"} for item in items]
        persist_verified_news(task_root, date_text, data)
        excel_file = write_excel(task_root, date_text, items)
        payload = build_dashboard_payload(date_text, items, excel_file, data)
        if args.preserve_existing_metrics:
            payload = preserve_existing_dashboard_metrics(date_text, payload)
        report_path, index_path = write_dashboard_data(date_text, payload)
        checks = validate_all(date_text, items, excel_file, report_path, index_path)
        log_payload = {
            "target_date": date_text,
            "generated_at": beijing_now().isoformat(timespec="seconds"),
            "status": "success",
            "verified_news_file": str(verified_json_path(task_root, date_text)),
            "excel_file": str(excel_file),
            "dashboard_report": str(report_path),
            "dashboard_index": str(index_path),
            "editorial_skill": editorial_skill,
            "thai_weather_check": thai_weather_log,
            "china_monitoring_check": {
                **china_monitoring_log(items),
                "ensure_step": china_monitoring_check,
            },
            "brazil_metrics_refresh": brazil_metrics_refresh,
            "india_metrics_refresh": india_metrics_refresh,
            "checks": checks,
        }
        write_task_log(task_root, date_text, log_payload)
        write_status(date_text, "success", checks)
        print(json.dumps(log_payload, ensure_ascii=False, indent=2))
        return 0
    except Exception as exc:
        error = str(exc)
        details = {"target_date": date_text, "task_root": str(task_root)}
        write_task_log(task_root, date_text, {
            "target_date": date_text,
            "generated_at": beijing_now().isoformat(timespec="seconds"),
            "status": "failed",
            "error": error,
        })
        write_status(date_text, "failed", details, error=error)
        print(json.dumps({"status": "failed", "newsDate": date_text, "error": error}, ensure_ascii=False, indent=2), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
