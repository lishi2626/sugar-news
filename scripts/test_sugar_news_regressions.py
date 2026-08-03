from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from brazil_sugar_metrics import (
    HISUGAR_IMPORT_COST_LIST_URL,
    article_available_for_target,
    build_snapshot,
    hisugar_query_list_page,
    parse_hisugar_list_articles,
    parse_hisugar_noisy_ocr_rows,
    stock_rows_from_pdf,
)
from sugar_news_pipeline import (
    CASE_REGRESSION_TOPICS,
    COUNTRY_SEARCH_TEMPLATES,
    OTHER_COUNTRY_SEARCH_TEMPLATES,
    SOURCE_MATRIX,
    candidate_country_confidence,
    candidate_has_verifiable_industry_fact,
    classify_sugar_topic,
    extract_metrics,
    infer_core_country,
    is_title_only_low_quality_context,
    is_india_indirect_sugar_relevant,
    is_medical_sugar_context,
    ensure_china_news_item,
    ensure_thai_weather_item,
    normalize_brazil_metrics,
    normalize_items,
    preserve_existing_dashboard_metrics,
    publication_window_for_target,
    rss_source_from_title,
    rss_summary_for_publication,
    success_exists,
    structured_candidate_from_rss,
    rss_sugar_relevant,
    tmd_thai_weather_item_from_text,
    validate_editorial_quality,
)
from verify_sugar_news_dashboard import verify_payload


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TASK_ROOT = PROJECT_ROOT
TARGET_DATE = "2026-07-20"
REQUIRED_DEDUPE_KEYS = {
    "india_aista_sugar_supply_shortage_unwarranted_20260720",
    "india_no_plan_ethanol_blend_above_e20_20260720",
    "india_cane_states_monsoon_heavy_rain_forecast_20260720",
}


def read_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def test_india_relevance_helpers() -> None:
    assert is_india_indirect_sugar_relevant(
        "India says no plan now to raise ethanol mix in gasoline above 20%; "
        "maize accounts for 37% of ethanol feedstock while molasses and sugar syrup remain cane-based inputs."
    )
    assert is_india_indirect_sugar_relevant(
        "OMC ethanol procurement tender includes B-heavy molasses, C-heavy molasses and sugar syrup supplies."
    )


def test_india_search_templates_cover_e20_reuters() -> None:
    queries = "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["印度"])
    for expected in (
        "India E20 petrol",
        "India ethanol above 20 percent",
        "India grain ethanol",
        "India maize ethanol",
        "India OMC ethanol tender",
        "site:reuters.com India E20",
        "site:reuters.com India ethanol",
        "site:reuters.com India molasses",
    ):
        assert expected in queries


def test_china_daily_monitoring_skill_and_templates() -> None:
    skill = (PROJECT_ROOT / ".codex" / "skills" / "sugar-news-editorial-rules" / "SKILL.md").read_text(encoding="utf-8")
    assert "中国糖业新闻每日重点监测" in skill
    assert "食糖进口" in skill
    assert "郑糖" in skill
    assert "糖料产区" in skill

    queries = "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["中国"])
    for expected in (
        "中国白糖",
        "中国食糖",
        "中国甘蔗",
        "中国甜菜糖",
        "广西糖业 甘蔗",
        "云南糖业 甘蔗",
        "食糖库存 产销率",
        "原糖进口 到港量 进口成本",
        "糖浆预混粉进口",
        "甘蔗收购价 种植补贴",
        "郑糖主力合约",
        "糖料产区 降雨 暴雨 干旱 台风",
    ):
        assert expected in queries


def test_brazil_metrics_daily_refresh_skill_and_workflow() -> None:
    skill = (PROJECT_ROOT / ".codex" / "skills" / "sugar-news-editorial-rules" / "SKILL.md").read_text(encoding="utf-8")
    assert "巴西糖价与库存每日刷新" in skill
    assert "brazil_sugar_metrics.py" in skill
    assert "Vercel" in skill

    workflow = (PROJECT_ROOT / ".github" / "workflows" / "sugar-news.yml").read_text(encoding="utf-8")
    pipeline_lines = [line for line in workflow.splitlines() if "scripts/sugar_news_pipeline.py" in line]
    assert pipeline_lines
    assert "scripts/brazil_sugar_metrics.py --date" in workflow
    assert "scripts/india_sugar_metrics.py --date" in workflow
    assert "--skip-metric-refresh" in "\n".join(pipeline_lines)
    assert "--skip-if-success" not in "\n".join(pipeline_lines)
    assert '".codex/skills/**"' in workflow
    assert "set -euo pipefail" in workflow
    assert 'SUGAR_NEWS_METRIC_REFRESH_TIMEOUT: "300"' in workflow
    assert "[brazil-only]" in workflow

    source = (PROJECT_ROOT / "scripts" / "sugar_news_pipeline.py").read_text(encoding="utf-8")
    assert "def refresh_brazil_metrics" in source
    assert "brazil_metrics_refresh = refresh_brazil_metrics(date_text)" in source
    assert "if args.preserve_existing_metrics:" in source
    assert "other-country item lacks concrete country/region; skipped before publication" in source

    assert "sourceDataDate" in skill
    assert "automatic monitoring + latest valid value retention" in skill
    assert "--preserve-existing-metrics" in skill


def test_other_country_rss_queries_are_concrete() -> None:
    queries = "\n".join(template for _language, template in OTHER_COUNTRY_SEARCH_TEMPLATES)
    for expected in ("Indonesia", "Pakistan", "Philippines", "Vietnam", "Russia", "Cameroon"):
        assert expected in queries
    assert "global sugar industry news" not in queries
    source = (PROJECT_ROOT / "scripts" / "sugar_news_pipeline.py").read_text(encoding="utf-8")
    assert "candidate failed pre-publication quality check" in source


def test_expanded_search_matrix_covers_user_case_topics() -> None:
    assert len(CASE_REGRESSION_TOPICS) == 16
    assert "UNICA" in SOURCE_MATRIX["巴西"]
    assert "Datagro" in SOURCE_MATRIX["巴西"]
    assert "Vasantdada Sugar Institute" in SOURCE_MATRIX["印度"]
    assert "Thai Meteorological Department" in SOURCE_MATRIX["泰国"]
    assert "淀粉糖行业数据" in SOURCE_MATRIX["中国"]
    assert "USDA / EIA / American Sugar Alliance" in SOURCE_MATRIX["其他国家"]

    brazil_queries = "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["巴西"])
    for expected in ("mistura etanol", "estoque de etanol", "preço da cana", "Datagro consumo global de açúcar"):
        assert expected in brazil_queries

    india_queries = "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["印度"])
    for expected in ("Lok Sabha sugar", "Rajya Sabha ethanol", "Vasantdada Sugar Institute", "red rot white grub", "ethanol procurement price"):
        assert expected in india_queries

    china_queries = "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["中国"])
    for expected in ("淀粉糖 玉米消耗量 产能利用率", "糖浆 白砂糖预混粉", "广西 食糖销量 工业库存"):
        assert expected in china_queries

    assert "EIA ethanol production" in "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["美国"])
    assert "raw sugar import allocation" in "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["印度尼西亚"])
    assert "Philippines sugarcane pest" in "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["菲律宾"])


def test_case_titles_are_classified_as_sugar_industry_candidates() -> None:
    cases = {
        "Brazil raises anhydrous ethanol blend in gasoline from 30% to 32%": "ethanol_policy",
        "Datagro sees global sugar consumption growth slowing to 1.8%": "supply_demand",
        "Lok Sabha ethanol programme saved foreign exchange worth Rs 1.98 lakh crore": "ethanol_capacity",
        "Vasantdada Sugar Institute releases sugarcane varieties with yield of 143.72 tonnes per hectare": "variety_research",
        "Mindanao planters sound alarm as sugarcane pest spreads to 4 provinces": "weather_pest",
        "Indonesia raw sugar import allocation for refineries revised by government regulation": "trade_policy",
        "EIA ethanol production averaged 1.113 million barrels per day and stocks rose": "ethanol_capacity",
        "Guangxi sugar sales reached 550000 tonnes and industrial inventory stood at 2.10 million tonnes": "supply_demand",
        "China starch sugar corn use and capacity utilization rose in July": "starch_sugar_substitute",
    }
    for title, expected_topic in cases.items():
        assert classify_sugar_topic(title) == expected_topic
        assert rss_sugar_relevant("其他国家", title) or "starch sugar" in title.lower()


def test_monday_publication_window_keeps_weekend_items() -> None:
    start, end, rule = publication_window_for_target("2026-08-02")
    assert rule == "monday_weekend_window_friday_16_to_monday_06"
    assert start.isoformat().startswith("2026-07-31T16:00:00")
    assert end.isoformat().startswith("2026-08-03T06:00:00")


def test_structured_rss_candidate_summary_is_specific() -> None:
    rss = {
        "title": "Mindanao planters sound alarm as sugarcane pest spreads to 4 provinces - Inquirer",
        "link": "https://example.test/philippines-pest",
        "published": "Sun, 02 Aug 2026 18:00:00 GMT",
        "description": "",
    }
    title, source = rss_source_from_title(rss["title"])
    candidate = structured_candidate_from_rss("菲律宾", rss, "2026-08-02", title, source)
    assert candidate["event_country"] == "菲律宾"
    assert candidate["topic"] == "weather_pest"
    assert "4 provinces" in candidate["metrics"]
    news, impact = rss_summary_for_publication(candidate)
    item = {
        "country_group": "其他国家",
        "country": "菲律宾",
        "title": title,
        "news": news,
        "impact": impact,
        "source_name": source,
        "source_url": rss["link"],
        "published_date_local": "2026-08-02",
        "event_date": "2026-08-02",
        "date_status": "verified",
        "dedupe_key": "philippines_pest_case",
        "importance": 80,
    }
    validate_editorial_quality(item, 1)
    assert "消息涉及" not in news
    assert "4 provinces" in news


def test_rss_publication_filters_reject_health_plugin_and_wrong_country_fallback() -> None:
    health = "Sugar Rationing in Early Life May Reduce Dementia Risk by 23% - News and Statistics"
    plugin = "August Mega Bundle: Sugar Bytes Factory synth and other plugins and sounds for $30"
    wrong_country = {
        "source_title": "Between destruction and neglect: This is the state of the Argentina sugar mill",
        "event_country": "菲律宾",
        "topic": "mill_operations",
        "metrics": [],
        "publisher": "CiberCuba",
    }
    assert is_medical_sugar_context(health)
    assert not rss_sugar_relevant("巴西", health)
    assert is_title_only_low_quality_context(plugin)
    assert not rss_sugar_relevant("俄罗斯", plugin)
    assert candidate_country_confidence(wrong_country, "菲律宾", wrong_country["source_title"], "CiberCuba") is None


def test_rss_publication_filters_keep_trusted_specific_industry_items() -> None:
    ban_rss = {
        "title": "India’s sugar export ban fuels smuggling ahead of peak festival season - The Kathmandu Post",
        "link": "https://example.test/ban",
        "published": "Sat, 01 Aug 2026 15:32:26 GMT",
        "description": "",
    }
    title, source = rss_source_from_title(ban_rss["title"])
    ban_candidate = structured_candidate_from_rss("印度", ban_rss, "2026-08-02", title, source)
    assert candidate_country_confidence(ban_candidate, "印度", title, source) == "explicit_country_or_region_in_title_or_source"
    ok, reason = candidate_has_verifiable_industry_fact(ban_candidate)
    assert ok, reason

    quota_candidate = {
        "source_title": "State-wise monthly sugar quota for sale in August 2026",
        "event_country": "印度",
        "topic": "trade_policy",
        "metrics": [],
    }
    ok, reason = candidate_has_verifiable_industry_fact(quota_candidate)
    assert not ok
    assert "policy direction or quantity" in reason

    consumer_e20_candidate = {
        "source_title": "Kejriwal demands choice between pure petrol and E20, seeks lower fuel prices at town hall",
        "event_country": "印度",
        "topic": "ethanol_policy",
        "metrics": ["E20"],
    }
    ok, reason = candidate_has_verifiable_industry_fact(consumer_e20_candidate)
    assert not ok
    assert "without sugar-feedstock linkage" in reason

    finance_candidate = {
        "source_title": "Dhampur Sugar Mills Latest Results: PAT rises to ₹65.33 crore, revenue up 5.69% YoY",
        "event_country": "印度",
        "topic": "mill_operations",
        "metrics": ["₹65.33 crore", "5.69%"],
    }
    ok, reason = candidate_has_verifiable_industry_fact(finance_candidate)
    assert not ok
    assert "title-only" in reason

    assert "5MT" not in extract_metrics("Petrol would have hit Rs 125/litre without ethanol blend ?oc=5MT")
    assert "Rs 125" in extract_metrics("Petrol would have hit Rs 125/litre without ethanol blend")


def test_skip_if_success_requires_report_and_index() -> None:
    assert not success_exists("2099-01-01")


def test_brazil_import_premium_hisugar_source_and_date_rule() -> None:
    assert HISUGAR_IMPORT_COST_LIST_URL == "https://www.hisugar.com/home/newListMore?parentId=49&level=3&childId=143&menuTap0"
    query_url = hisugar_query_list_page(2)
    assert "parentId=49" in query_url
    assert "categoryId=143" in query_url
    assert "pageNo=2" in query_url

    sample_html = """
    <li>
      <p class="more"><a href="/home/articleContent?id=2026072308523084118748">查看文章</a></p>
      <dd>2026-07-23</dd>
      <h3>20260722食糖进口成本及利润估算</h3>
    </li>
    <li>
      <p class="more"><a href="/home/articleContent?id=2026072415105108451386">查看文章</a></p>
      <dd>2026-07-24</dd>
      <h3>20260723食糖进口成本及利润估算</h3>
    </li>
    """
    articles = parse_hisugar_list_articles(sample_html, "https://www.hisugar.com/home/newListMore")
    assert [article["title_date"] for article in articles] == ["2026-07-22", "2026-07-23"]
    eligible = [article for article in articles if article["title_date"] <= "2026-07-22"]
    assert len(eligible) == 1
    assert eligible[0]["article_title"] == "20260722食糖进口成本及利润估算"

    same_day_late = {
        "title_date": "2026-07-23",
        "article_published_at": "2026-07-24 17:34:43",
    }
    previous_day_available = {
        "title_date": "2026-07-22",
        "article_published_at": "2026-07-23 10:44:45",
    }
    six_am = datetime.fromisoformat("2026-07-24T06:00:00+08:00")
    late_refresh = datetime.fromisoformat("2026-07-24T18:00:00+08:00")
    assert not article_available_for_target(same_day_late, "2026-07-23", six_am)
    assert article_available_for_target(same_day_late, "2026-07-23", late_refresh)
    assert article_available_for_target(previous_day_available, "2026-07-23", six_am)

    late_workflow = (PROJECT_ROOT / ".github" / "workflows" / "brazil-dashboard-late-refresh.yml").read_text(
        encoding="utf-8"
    )
    assert 'cron: "0 2 * * *"' in late_workflow
    assert "refresh_brazil_dashboard_report.py" in late_workflow
    assert "--premium-only" in late_workflow
    assert "sugar_news_pipeline.py" not in late_workflow

    ocr_text = (
        "20260720 20260721 20260722 20260723 20260724 "
        "5L50 ． 0 ． 30 5L50 ． 0 ． 30 5L50 ． 0 ． 30 "
        "5L50 ． 0 ． 30 5L75 ． 0 ． 55 进口升贴水"
    )
    parsed_rows = parse_hisugar_noisy_ocr_rows(
        ocr_text,
        {
            "article_id": "dynamic-test",
            "article_title": "20260724食糖进口成本及利润估算",
            "article_published_at": "2026-07-27 08:56:27",
            "title_date": "2026-07-24",
            "source_url": "https://example.test/article",
        },
        "https://example.test/table.png",
        "test_ocr",
    )
    assert [(row["data_date"], row["premium_discount_cents_per_lb"]) for row in parsed_rows] == [
        ("2026-07-20", -0.3),
        ("2026-07-21", -0.3),
        ("2026-07-22", -0.3),
        ("2026-07-23", -0.3),
        ("2026-07-24", -0.55),
    ]


def test_thailand_weather_templates_and_tmd_item_generation() -> None:
    queries = "\n".join(template for _language, template in COUNTRY_SEARCH_TEMPLATES["泰国"])
    for expected in (
        "Thailand sugarcane rainfall forecast",
        "Udon Thani Khon Kaen Nakhon Ratchasima",
        "Nakhon Sawan Kanchanaburi Lopburi Chai Nat",
    ):
        assert expected in queries

    sample = (
        "Forecast Date: July 23, 2026 Daily Weather Forecast Issued at 5.00 a.m. "
        "During 23 - 24 Jul, the strong southwest monsoon prevails over Thailand. "
        "Northeastern: Scattered thundershowers and isolated heavy rains mostly in "
        "Loei, Udon Thani and Khon Kaen. Central: Fairly widespread thundershowers "
        "and isolated heavy rains in Kanchanaburi. Eastern: thundershowers in Sa Kaeo and Chon Buri."
    )
    item = tmd_thai_weather_item_from_text(sample, "2026-07-23")
    assert item is not None
    assert item["country_group"] == "泰国"
    assert item["impact"].startswith("利空：")
    assert "幅度有限" not in item["impact"]
    assert "乌隆他尼" in item["news"]
    assert "孔敬" in item["news"]
    assert "北碧" in item["news"]
    assert "来源：泰国气象局" in item["news"]

    regional_sample = (
        "Forecast Date: July 23, 2026 Daily Weather Forecast Issued at 5.00 a.m. "
        "Northeastern: Scattered thundershowers and isolated heavy rains. "
        "Northern: Isolated thundershowers. Central: Fairly widespread thundershowers. "
        "Eastern: Isolated thundershowers."
    )
    regional_item = tmd_thai_weather_item_from_text(regional_sample, "2026-07-23")
    assert regional_item is not None
    assert "东北部核心甘蔗产区" in regional_item["news"]
    assert "北部甘蔗产区" in regional_item["news"]
    assert "中部及西部甘蔗产区" in regional_item["news"]
    assert regional_item["published_date_local"] == "2026-07-23"


def test_thailand_weather_is_added_to_existing_verified_news() -> None:
    data = {
        "target_date": "2026-07-23",
        "items": [
            {
                "country_group": "巴西",
                "country": "巴西",
                "title": "巴西糖厂运营调整",
                "news": "巴西圣保罗州一家糖厂宣布7月下旬暂停一条压榨线检修，日压榨能力临时减少约1万吨。压榨减少会推迟糖料入榨和食糖生产释放，但检修时间较短，对全国供应影响暂为中性。来源：Test（https://example.test/brazil）",
                "impact": "中性：短期检修推迟区域压榨，但暂不直接改变巴西全国供应判断。",
                "source_name": "Test",
                "source_url": "https://example.test/brazil",
                "published_date_local": "2026-07-23",
                "dedupe_key": "test_brazil",
            }
        ],
    }

    def fake_fetch(report_date: str) -> tuple[dict | None, dict]:
        sample = (
            "Forecast Date: July 23, 2026 Daily Weather Forecast Issued at 5.00 a.m. "
            "Northeastern: Scattered thundershowers and isolated heavy rains."
        )
        return tmd_thai_weather_item_from_text(sample, report_date), {"request_status": "executed"}

    updated, log = ensure_thai_weather_item(data, "2026-07-23", fetcher=fake_fetch)
    assert log["status"] == "added"
    assert len(updated["items"]) == 2
    assert any(item["country_group"] == "泰国" for item in updated["items"])
    assert len(normalize_items(updated)) == 2


def test_thailand_weather_fallback_recovers_existing_dated_item() -> None:
    data = {
        "target_date": "2026-07-23",
        "items": [],
    }

    def failed_fetch(report_date: str) -> tuple[dict | None, dict]:
        return None, {"request_status": "failed", "error": "403"}

    updated, log = ensure_thai_weather_item(
        data,
        "2026-07-23",
        fetcher=failed_fetch,
        open_meteo_fetcher=failed_fetch,
    )
    assert log["status"] == "added"
    assert log["fallback"]["retained_count"] == 1
    item = updated["items"][0]
    assert item["country_group"] == "泰国"
    assert item["published_date_local"] == "2026-07-23"
    assert item["source_url"] in item["news"]
    assert "降雨" in item["news"]


def test_no_fixed_country_cap_in_autogen() -> None:
    source = (PROJECT_ROOT / "scripts" / "sugar_news_pipeline.py").read_text(encoding="utf-8")
    assert "retained_for_country >= 2" not in source


def test_non_industry_sugar_titles_are_filtered() -> None:
    assert not rss_sugar_relevant("其他国家", "Palm Sugar: A Village Story launches on Windows PC")
    assert not rss_sugar_relevant("印度", "Bengaluru author debuts novel The Burnt Sugar Club")


def test_editorial_country_reclassification_rules() -> None:
    indonesia_country, indonesia_group = infer_core_country(
        "ChiniMandi reports Indonesia sugar import policy and domestic sugar supply",
        "巴西",
    )
    assert indonesia_country == "印度尼西亚"
    assert indonesia_group == "其他国家"

    cameroon_country, cameroon_group = infer_core_country(
        "ChiniMandi reports Cameroon sugar production and import policy",
        "印度",
    )
    assert cameroon_country == "喀麦隆"
    assert cameroon_group == "其他国家"


def test_medical_sugar_news_is_excluded() -> None:
    sample = "Blood sugar monitoring improves diabetes treatment with insulin guidance"
    assert is_medical_sugar_context(sample)
    assert not rss_sugar_relevant("其他国家", sample)


def test_valid_brazil_cane_sugar_ethanol_news_is_allowed() -> None:
    sample = "Brazil sugarcane crushing and sugar production rise while mills adjust ethanol output"
    assert rss_sugar_relevant("巴西", sample)
    country, group = infer_core_country(sample, "巴西")
    assert country == "巴西"
    assert group == "巴西"


def test_india_water_resource_pressure_is_bullish() -> None:
    data = {
        "target_date": "2026-07-23",
        "items": [
            {
                "country_group": "印度",
                "country": "印度",
                "title": "印度半干旱地区甘蔗单产提升伴随水资源风险",
                "news": "Mongabay India称，印度半干旱地区甘蔗单产提升的同时，仍面临气候和水资源压力。水分约束可能限制甘蔗可持续扩产，并增加未来糖料供应波动风险。来源：Mongabay India（https://example.test/water）",
                "impact": "利多：水资源压力和水分约束可能限制甘蔗扩产和单产稳定性，未来糖料供应存在下降风险。",
                "source_name": "Mongabay India",
                "source_url": "https://example.test/water",
                "published_date_local": "2026-07-23",
                "dedupe_key": "india_water_resource_pressure",
            }
        ],
    }
    assert len(normalize_items(data)) == 1
    data["items"][0]["impact"] = "影响有限：报道未明确对应印度核心甘蔗主产区。"
    try:
        normalize_items(data)
    except ValueError as exc:
        assert "water-resource pressure" in str(exc) or "transmission logic" in str(exc)
    else:
        raise AssertionError("water-resource pressure should require bullish impact")


def test_editorial_quality_rejects_publication_date_formula_and_accepts_key_dates() -> None:
    bad = {
        "country_group": "印度",
        "country": "印度",
        "title": "India sugar policy",
        "news": "2026-07-23 ChiniMandi报道：印度糖厂甘蔗款支付改善。甘蔗款支付改善有助于稳定未来糖料供应。来源：ChiniMandi（https://example.test/a）",
        "impact": "利空：甘蔗款支付改善有助于稳定未来糖料供应。",
    }
    try:
        validate_editorial_quality(bad, 1)
    except ValueError as exc:
        assert "publication-date" in str(exc) or "reporting formula" in str(exc)
    else:
        raise AssertionError("publication date formula should be rejected")

    good = {
        "country_group": "印度",
        "country": "印度",
        "title": "India sugar policy",
        "news": "印度政府公布2026/27榨季甘蔗款支付安排，政策执行期关系到糖厂现金流和蔗农交售节奏。若付款秩序改善，蔗农种植积极性和后续糖料供应预期将得到支撑。来源：ChiniMandi（https://example.test/a）",
        "impact": "利空：甘蔗款支付改善有助于稳定未来糖料供应。",
    }
    validate_editorial_quality(good, 2)


def test_summary_must_be_two_or_three_chinese_sentences() -> None:
    data = {
        "target_date": "2026-07-23",
        "items": [
            {
                "country_group": "巴西",
                "country": "巴西",
                "title": "Brazil sugarcane",
                "news": "巴西中南部甘蔗压榨进度改善，糖产量释放速度加快。供应增加可能提高国际市场可用糖源，对原糖价格形成压力。来源：Test（https://example.test/b）",
                "impact": "利空：糖产量释放增加可能压制国际糖价。",
                "source_name": "Test",
                "source_url": "https://example.test/b",
                "published_date_local": "2026-07-23",
                "dedupe_key": "test_brazil_cane",
            }
        ],
    }
    assert len(normalize_items(data)) == 1

    data["items"][0]["news"] = "巴西中南部甘蔗压榨进度改善。来源：Test（https://example.test/b）"
    try:
        normalize_items(data)
    except ValueError as exc:
        assert "2-3" in str(exc)
    else:
        raise AssertionError("one-sentence summary should be rejected")


def test_editorial_quality_rejects_vague_fallback_summary() -> None:
    bad = {
        "country_group": "巴西",
        "country": "巴西",
        "title": "Brazil sugar prices",
        "news": "ChiniMandi消息涉及巴西食糖价格或市场流通变化。价格变化会影响贸易商采购、终端补库和政策调控预期，对短期糖价走势具有参考意义。来源：ChiniMandi（https://example.test/brazil-price）",
        "impact": "中性：该信息需要继续跟踪，短期对当期糖产量和出口量的直接影响有限。",
        "source_name": "ChiniMandi",
        "source_url": "https://example.test/brazil-price",
    }
    try:
        validate_editorial_quality(bad, 1)
    except ValueError as exc:
        assert "vague" in str(exc) or "media source" in str(exc)
    else:
        raise AssertionError("vague source-led fallback summary should be rejected")


def test_editorial_quality_requires_concrete_action_direction_and_impact_path() -> None:
    bad = {
        "country_group": "印度",
        "country": "印度",
        "title": "India sugar market",
        "news": "印度糖业出现新的市场情况，行业各方仍在关注相关变化。该信息对食糖供需和糖价的影响还需要继续观察。来源：Test（https://example.test/india）",
        "impact": "中性：信息仍需跟踪。",
        "source_name": "Test",
        "source_url": "https://example.test/india",
    }
    try:
        validate_editorial_quality(bad, 2)
    except ValueError as exc:
        assert "vague" in str(exc) or "clear event action" in str(exc)
    else:
        raise AssertionError("summary without concrete action and direction should be rejected")


def test_current_report_contains_china_section_after_thailand() -> None:
    report = read_json(PROJECT_ROOT / "public" / "sugar-news" / "data" / "reports" / "2026" / "07" / "2026-07-23.json")
    countries = [country["country"] for country in report["countries"]]
    assert countries[:4] == ["巴西", "印度", "泰国", "中国"]
    china = next(country for country in report["countries"] if country["country"] == "中国")
    assert len(china["items"]) == 4
    text = "\n".join(item["news"] for item in china["items"])
    for phrase in ("进口甘蔗", "工业库存约449万吨", "云南制糖集团报价", "文山、红河、玉溪、普洱"):
        assert phrase in text


def test_brazil_india_metric_value_is_under_absolute_column() -> None:
    html = (PROJECT_ROOT / "public" / "sugar-news" / "index.html").read_text(encoding="utf-8")
    assert '["字段", "绝对值", "（%）"]' in html
    assert 'appendValueRow("取值", config.value());' in html
    assert 'const td = document.createElement("td");' in html
    assert 'const pctTd = document.createElement("td");' in html
    assert html.index('td.className = "brazil-metric-main";') < html.index('pctTd.className = "metric-change na";')


def test_ist_utc_beijing_date_handling() -> None:
    try:
        ist = ZoneInfo("Asia/Kolkata")
    except Exception:
        ist = timezone(timedelta(hours=5, minutes=30), name="Asia/Kolkata")
    try:
        shanghai = ZoneInfo("Asia/Shanghai")
    except Exception:
        shanghai = timezone(timedelta(hours=8), name="Asia/Shanghai")
    local_article_time = datetime(2026, 7, 20, 0, 30, tzinfo=ist)
    assert local_article_time.date().isoformat() == TARGET_DATE
    assert local_article_time.astimezone(timezone.utc).date().isoformat() == "2026-07-19"
    assert datetime(2026, 7, 21, 6, 0, tzinfo=shanghai).date().isoformat() == "2026-07-21"


def test_verified_news_contains_required_india_items() -> None:
    data = read_json(TASK_ROOT / "data" / "verified_news" / "2026" / "07" / f"sugar_news_{TARGET_DATE}.json")
    keys = {item.get("dedupe_key") for item in data["items"]}
    assert REQUIRED_DEDUPE_KEYS <= keys
    country_counts = Counter(item.get("country_group") for item in data["items"])
    assert country_counts["印度"] >= 4


def test_excel_dashboard_consistency() -> None:
    report = read_json(PROJECT_ROOT / "public" / "sugar-news" / "data" / "reports" / "2026" / "07" / f"{TARGET_DATE}.json")
    excel_path = TASK_ROOT / "reports" / "2026" / "07" / f"Sugar News {TARGET_DATE}.xlsx"
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    excel_rows = [
        (row[0], row[1], row[2])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0] and row[1]
    ]
    dashboard_rows = [
        (country["country"], item["news"], f'{item["impactType"]}：{item["impact"]}')
        for country in report["countries"]
        for item in country["items"]
    ]
    assert len(excel_rows) == len(dashboard_rows)
    assert Counter(row[0] for row in excel_rows)["印度"] >= 4
    dashboard_text = "\n".join(news for _country, news, _impact in dashboard_rows)
    for phrase in ("AISTA", "E20以上", "Hardoi和Unnao"):
        assert phrase in dashboard_text


def test_india_metrics_price_changes_and_stock_source_rules() -> None:
    report = read_json(PROJECT_ROOT / "public" / "sugar-news" / "data" / "reports" / "2026" / "07" / f"{TARGET_DATE}.json")
    metrics = report["indiaMetrics"]
    expected_chinimandi_urls = {
        "domesticWholesalePrice": "https://www.chinimandi.com/wholesale-sugar-prices/",
        "domesticRetailPrice": "https://www.chinimandi.com/retail-prices/",
    }
    for field in ("domesticWholesalePrice", "domesticRetailPrice", "upExMillPrice"):
        metric = metrics[field]
        assert metric["status"] == "ok"
        assert metric["previousDataDate"]
        assert metric["changePct"] is not None
        if field in expected_chinimandi_urls:
            assert metric["sourceName"] == "ChiniMandi"
            assert metric["sourceUrl"] == expected_chinimandi_urls[field]
            assert metric["includesGst"] is True
            assert metric["citiesUsed"]
            assert metric["cityCount"] == len(metric["citiesUsed"])
        if field == "domesticRetailPrice":
            assert metric["changeInrPerKg"] is not None
        else:
            assert metric["changeInrPerQuintal"] is not None
        if field == "upExMillPrice":
            low = metric["rangeInrPerQuintal"]["low"]
            high = metric["rangeInrPerQuintal"]["high"]
            assert metric["midpointInrPerQuintal"] == (low + high) / 2
            assert metric["sourceName"] == "ChiniMandi — Daily Sugar Market Update"
            assert metric["market"] == "Uttar Pradesh"
            assert metric["grade"] == "M/30"
            assert metric["includesGst"] is False
            assert "daily-sugar-market-update-by-vizzie" in metric["sourceUrl"]
            assert metric["previousSourceUrl"]
            assert metric["yoySourceUrl"]

    stock = metrics["carryoverStock"]
    if stock["status"] == "ok":
        source = stock.get("organization") or stock.get("sourceName") or ""
        assert any(token in source for token in ("Government of India", "Department of Food", "ISMA", "NFCSF", "印度政府"))
    for forecast in metrics.get("carryoverStockForecasts", []):
        assert forecast.get("sourceTier") == "market_forecast_comparison_only"


def test_brazil_sugar_stock_date_comes_from_acumulado_ate() -> None:
    sample_text = "BRASIL 1.000 2.000 3.450.164 Acumulado ate: 30/06/2026"
    rows = stock_rows_from_pdf(
        sample_text,
        "2026/2027",
        {
            "title": "ESTOQUES DE AÇÚCAR POR TIPO - SAFRA 2026-27",
            "url": "https://example.test/009ESTOQUESDEACARPORTIPOSAFRA20262027_20072026.pdf",
            "document_number": "009",
            "published_at": "2026-07-20",
        },
        "test-hash",
    )
    assert len(rows) == 1
    assert rows[0]["reference_date"] == "2026-06-30"
    assert rows[0]["reference_date_raw"] == "30/06/2026"
    assert rows[0]["reference_date_source"] == "pdf_acumulado_ate"
    assert rows[0]["document_title"] == "ESTOQUES DE AÇÚCAR POR TIPO - SAFRA 2026-27"
    assert rows[0]["stock_total_tonnes"] == 3450164


def test_brazil_dashboard_does_not_show_fetch_time_or_report_as_date() -> None:
    html = (PROJECT_ROOT / "public" / "sugar-news" / "index.html").read_text(encoding="utf-8")
    assert "发布日期/报告" not in html
    assert "last fetched" not in html.lower()
    assert "fetched_at" not in html
    assert "数据日期：" in html
    assert "抓取日期：" not in html
    for forbidden in ("沿用上一期数据", "暂无最新数据", "等待更新", "未抓取到数据", "数据未同步", "数据待更新"):
        assert forbidden not in html


def test_brazil_dashboard_refresh_date_matches_news_and_cards_use_source_dates() -> None:
    metrics = normalize_brazil_metrics("2026-07-25")
    assert metrics["dataDate"] == "2026-07-25"
    for field in ("sugarPremium", "sugarStock", "ethanolStock"):
        assert metrics[field]["refreshDate"] == "2026-07-25"
        assert metrics[field]["dataDate"] == metrics[field]["sourceDataDate"]
        assert metrics[field]["sourceDataDate"]


def test_brazil_snapshot_retains_previous_success_when_history_is_empty() -> None:
    previous = {
        "sugarPremium": {
            "indicator": "brazil_sugar_premium",
            "status": "ok",
            "data_date": "2026-07-23",
            "premium_discount_cents_per_lb": -0.3,
        },
        "sugarStock": {
            "indicator": "brazil_sugar_stock",
            "status": "ok",
            "reference_date": "2026-06-30",
            "sugar_stock_value": 345.0164,
        },
        "ethanolStock": {
            "indicator": "brazil_ethanol_stock",
            "status": "ok",
            "reference_date": "2026-07-01",
            "total_ethanol_stock": 291.3832,
        },
    }
    snapshot = build_snapshot(
        {"version": 1, "records": [], "lastUpdatedAt": None},
        "2026-07-26",
        [],
        previous,
    )
    assert snapshot["displayDate"] == "2026-07-26"
    assert snapshot["sugarPremium"]["premium_discount_cents_per_lb"] == -0.3
    assert snapshot["sugarStock"]["sugar_stock_value"] == 345.0164
    assert snapshot["ethanolStock"]["total_ethanol_stock"] == 291.3832
    for field in ("sugarPremium", "sugarStock", "ethanolStock"):
        assert snapshot[field]["status"] == "ok"
        assert snapshot[field]["refresh_date"] == "2026-07-26"
        assert snapshot[field]["source_data_date"]


def test_china_column_is_mandatory_every_day() -> None:
    skill = (PROJECT_ROOT / ".codex" / "skills" / "sugar-news-editorial-rules" / "SKILL.md").read_text(encoding="utf-8")
    assert "The China column is mandatory in every daily Sugar News report." in skill
    updated, log = ensure_china_news_item(
        {"target_date": "2099-01-01", "items": []},
        "2099-01-01",
    )
    china_items = [item for item in updated["items"] if item["country_group"] == "中国"]
    assert len(china_items) == 1
    assert china_items[0]["title"] == "中国糖业每日监测"
    assert china_items[0]["impact"].startswith("中性：")
    assert log["status"] == "added_monitoring_note"
    validate_editorial_quality(china_items[0], 1)


def test_confirmed_china_items_are_added_for_20260725() -> None:
    verified = read_json(
        PROJECT_ROOT / "data" / "verified_news" / "2026" / "07" / "sugar_news_2026-07-25.json"
    )
    china_items = [item for item in verified["items"] if item["country_group"] == "中国"]
    assert len(china_items) == 4
    assert {item["title"] for item in china_items} == {
        "中国进口甘蔗规模快速增长",
        "国内库存偏高，后期食糖进口预计增加",
        "国内糖市购销平稳，云南报价小幅下调",
        "云南甘蔗产区出现强降雨预报",
    }
    for index, item in enumerate(china_items, start=1):
        validate_editorial_quality(item, index)


def test_news_only_repair_preserves_existing_dashboard_metrics() -> None:
    existing = read_json(
        PROJECT_ROOT / "public" / "sugar-news" / "data" / "reports" / "2026" / "07" / "2026-07-25.json"
    )
    candidate = {
        "newsDate": "2026-07-25",
        "brazilMetrics": {"dataDate": "changed"},
        "indiaMetrics": {"dataDate": "changed"},
    }
    preserved = preserve_existing_dashboard_metrics("2026-07-25", candidate)
    assert preserved["brazilMetrics"] == existing["brazilMetrics"]
    assert preserved["indiaMetrics"] == existing["indiaMetrics"]


def test_dashboard_verifier_requires_china_output() -> None:
    report = read_json(
        PROJECT_ROOT / "public" / "sugar-news" / "data" / "reports" / "2026" / "07" / "2026-07-25.json"
    )
    result = verify_payload(report, "2026-07-25")
    assert result["chinaItemCount"] == 4
    missing_china = dict(report)
    missing_china["countries"] = [
        country for country in report["countries"] if country["country"] != "中国"
    ]
    try:
        verify_payload(missing_china, "2026-07-25")
    except AssertionError as exc:
        assert "China section" in str(exc)
    else:
        raise AssertionError("Production verifier accepted a report without China output")


def main() -> None:
    tests = [
        test_india_relevance_helpers,
        test_india_search_templates_cover_e20_reuters,
        test_china_daily_monitoring_skill_and_templates,
        test_brazil_metrics_daily_refresh_skill_and_workflow,
        test_other_country_rss_queries_are_concrete,
        test_expanded_search_matrix_covers_user_case_topics,
        test_case_titles_are_classified_as_sugar_industry_candidates,
        test_monday_publication_window_keeps_weekend_items,
        test_structured_rss_candidate_summary_is_specific,
        test_rss_publication_filters_reject_health_plugin_and_wrong_country_fallback,
        test_rss_publication_filters_keep_trusted_specific_industry_items,
        test_thailand_weather_templates_and_tmd_item_generation,
        test_thailand_weather_is_added_to_existing_verified_news,
        test_thailand_weather_fallback_recovers_existing_dated_item,
        test_no_fixed_country_cap_in_autogen,
        test_non_industry_sugar_titles_are_filtered,
        test_editorial_country_reclassification_rules,
        test_medical_sugar_news_is_excluded,
        test_valid_brazil_cane_sugar_ethanol_news_is_allowed,
        test_india_water_resource_pressure_is_bullish,
        test_editorial_quality_rejects_publication_date_formula_and_accepts_key_dates,
        test_summary_must_be_two_or_three_chinese_sentences,
        test_editorial_quality_rejects_vague_fallback_summary,
        test_editorial_quality_requires_concrete_action_direction_and_impact_path,
        test_current_report_contains_china_section_after_thailand,
        test_brazil_india_metric_value_is_under_absolute_column,
        test_ist_utc_beijing_date_handling,
        test_verified_news_contains_required_india_items,
        test_excel_dashboard_consistency,
        test_india_metrics_price_changes_and_stock_source_rules,
        test_brazil_sugar_stock_date_comes_from_acumulado_ate,
        test_brazil_dashboard_does_not_show_fetch_time_or_report_as_date,
        test_brazil_dashboard_refresh_date_matches_news_and_cards_use_source_dates,
        test_brazil_snapshot_retains_previous_success_when_history_is_empty,
        test_china_column_is_mandatory_every_day,
        test_confirmed_china_items_are_added_for_20260725,
        test_news_only_repair_preserves_existing_dashboard_metrics,
        test_dashboard_verifier_requires_china_output,
    ]
    for test in tests:
        test()
        print(f"PASS {test.__name__}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"FAIL {exc}", file=sys.stderr)
        raise
